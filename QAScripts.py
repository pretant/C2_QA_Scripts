import math
import os
import queue
import re
import shutil
import subprocess
import sys
import tempfile
import threading
import time
import tkinter as tk
import traceback
import warnings
import webbrowser
from collections import Counter, defaultdict
from datetime import date, datetime
from tkinter import messagebox, filedialog, simpledialog, ttk
from typing import List

import customtkinter as ctk
import exifread
import exiftool
import geopy.distance
import pandas as pd
import piexif
import psutil
import requests
from PIL import Image, ExifTags, ImageFile
from dateutil.parser import parse
from geopy.distance import geodesic
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from thefuzz import process, fuzz

ImageFile.LOAD_TRUNCATED_IMAGES = True
warnings.simplefilter(action='ignore', category=UserWarning)

exe_name = "QAScripts.exe"
version_url = "https://pretant.github.io/C2_QA_Scripts/version.txt"
version_history_url = "https://pretant.github.io/C2_QA_Scripts/versionhistory/"


def print_to_widget(print_text, newline=True):
    text_space.configure(state='normal')
    if newline:
        text_space.insert(tk.END, str(print_text) + '\n')
    else:
        text_space.insert(tk.END, str(print_text))
    text_space.configure(state='disabled')
    text_space.see(tk.END)


def display_exception():
    exc_type, exc_value, exc_traceback = sys.exc_info()
    error_msg = "".join(traceback.format_exception(exc_type, exc_value, exc_traceback))

    # Calculate text_widget dimensions based on the error message
    lines = error_msg.splitlines()
    max_width = max(len(line) for line in lines)
    num_lines = len(lines)

    error_window = ctk.CTkToplevel()
    error_window.title("Error")

    error_label = ctk.CTkLabel(error_window, text="An unexpected error occurred:", anchor="w")
    error_label.pack(padx=10, pady=10, anchor="w")

    text_widget = tk.Text(error_window, wrap=tk.NONE, padx=10, pady=10, bg="#18191A", fg="#B3B3B3",
                          insertbackground="#B3B3B3", selectbackground="#7a7a7a",
                          width=max_width, height=num_lines)
    text_widget.insert(tk.END, error_msg)
    text_widget.config(state=tk.DISABLED)  # Prevent print_text editing
    text_widget.pack(expand=True, fill=tk.BOTH)

    def copy_error_message():
        error_window.clipboard_clear()
        error_window.clipboard_append(error_msg)
        tool_tip.show_tip()
        error_window.after(1500, tool_tip.hide_tip)

    copy_button2 = ctk.CTkButton(error_window, text="Copy", command=copy_error_message)
    copy_button2.pack(pady=10)

    tool_tip = ToolTip(copy_button2, "Copied to clipboard")


def get_current_version():
    return "4.2.0"


def open_version_history(event):
    webbrowser.open(version_history_url)


def get_latest_version():
    response = requests.get(version_url)

    if response.status_code == 200:
        return response.text.strip()
    else:
        print_to_widget(f"Failed to fetch latest version.")
        return None


def start_update_script():
    response = requests.get("https://pretant.github.io/C2_QA_Scripts/UpdateQAScripts.exe")
    if response.status_code == 200:
        with open("UpdateQAScripts.exe", "wb") as f:
            f.write(response.content)
        return subprocess.Popen(["UpdateQAScripts.exe"], creationflags=subprocess.CREATE_NEW_CONSOLE)
    else:
        print_to_widget("Could not download update script.\n")
        return None


def check_for_updates():
    current_version = get_current_version()
    latest_version = get_latest_version()

    if latest_version:
        if latest_version != current_version:
            print_to_widget(f"New version available: v{latest_version}")
            print_to_widget("Do you want to update?")
            response = messagebox.askyesno("Update Available", "Do you want to update?")
            if response:
                print_to_widget("Downloading update...")
                for proc in psutil.process_iter():
                    try:
                        if proc.name() == exe_name:
                            print_to_widget("Closing app...")
                            start_update_script()
                            time.sleep(5)  # Give the update script some time to start before killing the main app
                            proc.kill()
                            break
                    except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
                        pass
            else:
                print_to_widget("\nUpdate canceled.\n")
        else:
            print_to_widget(
                f"You are running the latest version.\nClick \"Version {latest_version}\" to see version history.\n")
    else:
        print_to_widget("Could not check for updates.\n")


# ToolTip class
class ToolTip:
    def __init__(self, widget, tooltip_text):
        self.widget = widget
        self.tip_window = None
        self.text = tooltip_text

    def show_tip(self):
        if self.tip_window or not self.text:
            return
        x, y, cx, cy = self.widget.bbox("insert")
        x += self.widget.winfo_rootx() + 20
        y += self.widget.winfo_rooty() + 30

        # Adjust position if tooltip goes off-screen
        screen_width = self.widget.winfo_screenwidth()
        screen_height = self.widget.winfo_screenheight()
        tip_width = 180  # Approximate width of the tooltip
        tip_height = 50  # Approximate height of the tooltip

        if y + tip_height > screen_height:
            y = self.widget.winfo_rooty() - tip_height
        if x + tip_width > screen_width:
            x = self.widget.winfo_rootx() - tip_width

        self.tip_window = tw = ctk.CTkToplevel(self.widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry("+%d+%d" % (x, y))
        tooltip_label = ctk.CTkLabel(tw, text=self.text, justify=tk.LEFT, font=("tahoma", 10.5, "normal"))
        tooltip_label.pack(ipadx=1)

    def hide_tip(self):
        tw = self.tip_window
        self.tip_window = None
        if tw:
            tw.destroy()


def create_tooltip(widget, tooltip_text):
    tool_tip = ToolTip(widget, tooltip_text)

    def enter(event):
        tool_tip.show_tip()

    def leave(event):
        tool_tip.hide_tip()

    widget.bind('<Enter>', enter)
    widget.bind('<Leave>', leave)


# Function to get "date taken" metadata of an image.
def get_date_taken(file_path):
    image_name = os.path.basename(file_path)
    with open(file_path, 'rb') as f:
        tags = exifread.process_file(f, stop_tag='EXIF DateTimeOriginal')
        if 'EXIF DateTimeOriginal' in tags:
            date_taken = tags['EXIF DateTimeOriginal'].values
            try:
                date_taken = datetime.strptime(date_taken, '%Y:%m:%d %H:%M:%S').strftime('%Y%m%d')
            except ValueError:
                date_taken = None
                date_str_queue = queue.Queue()
        else:
            date_taken = None
            date_str_queue = queue.Queue()

            def ask_date():
                selected_date = tk.simpledialog.askstring("Enter Date",
                                                          f"Unable to find \"datetaken\" metadata for {image_name}. "
                                                          f"Please enter the flight date for this structure ("
                                                          f"YYYYMMDD)", parent=root)
                date_str_queue.put(selected_date)

            while date_taken is None:
                root.after(0, ask_date)
                date_str = date_str_queue.get()  # This will block until a result is available
                try:
                    date_taken = datetime.strptime(date_str, '%Y%m%d').strftime('%Y%m%d')
                except ValueError:
                    print_to_widget("Invalid date format, please try again.")
                    continue
    return date_taken


# Function to extract GPS data from images using PIL library
def get_gps_from_image(filepath):
    try:
        with Image.open(filepath) as img:
            # Get EXIF data from the image
            exif_data = img._getexif()
            if exif_data:
                # Extract only necessary EXIF tags using PIL's TAGS dictionary
                exif_data = {
                    ExifTags.TAGS[k]: v
                    for k, v in exif_data.items()
                    if k in ExifTags.TAGS
                }
                # Extract GPSInfo tag and its corresponding sub-tags using PIL's GPSTAGS dictionary
                gps_info = exif_data.get('GPSInfo', {})
                if gps_info:
                    gps_info = {
                        ExifTags.GPSTAGS.get(key, key): value
                        for key, value in gps_info.items()
                    }
                    # You can process gps_info here

                    # Extract latitude and longitude data from GPSInfo
                    lat = gps_info.get('GPSLatitude')
                    lat_ref = gps_info.get('GPSLatitudeRef', 'N')
                    lng = gps_info.get('GPSLongitude')
                    lng_ref = gps_info.get('GPSLongitudeRef', 'E')
                    if lat and lng:
                        # Convert latitude and longitude from degrees, minutes, seconds to decimal degrees
                        lat_decimal = lat[0].numerator / lat[0].denominator + (
                                lat[1].numerator / lat[1].denominator) / 60 + (
                                              lat[2].numerator / lat[2].denominator) / 3600
                        lng_decimal = lng[0].numerator / lng[0].denominator + (
                                lng[1].numerator / lng[1].denominator) / 60 + (
                                              lng[2].numerator / lng[2].denominator) / 3600
                        # Apply negative sign to latitude and/or longitude if necessary
                        if lat_ref == 'S':
                            lat_decimal *= -1
                        if lng_ref == 'W':
                            lng_decimal *= -1
                        # Return latitude and longitude as a tuple
                        return lat_decimal, lng_decimal
    except (IOError, OSError, KeyError, AttributeError):
        pass
    # Return None if GPS data is not found in the image or there is an error extracting the data
    return None


# function to read GPS image direction metadata and convert to directional letter
def get_gps_direction_letter(image_path):
    directions = ['N', 'NE', 'E', 'SE', 'S', 'SW', 'W', 'NW', 'N']
    exif_data = piexif.load(image_path)
    gps_data = exif_data.get('GPS', {})
    direction = gps_data.get(piexif.GPSIFD.GPSImgDirection, None)
    direction_ref = gps_data.get(piexif.GPSIFD.GPSImgDirectionRef, None)
    if direction is not None and direction_ref is not None:
        degrees = math.floor(direction[0] / direction[1])
        if direction_ref == 'W':
            degrees = 360 - degrees
        index = int((degrees + 22.5) / 45)
        return directions[index]
    else:
        return None


def get_yaw_direction_letter(image_path):
    with exiftool.ExifTool() as et:
        try:
            metadata = et.execute_json('-XMP:all', image_path)
        except (FileNotFoundError, PermissionError, OSError) as e:
            display_exception()
            return None

    if metadata:
        xmp_data = metadata[0]
        flight_yaw_degree = float(xmp_data.get('XMP:FlightYawDegree'))
        if flight_yaw_degree:
            if flight_yaw_degree < 0:
                flight_yaw_degree += 360
            directions = ['N', 'NE', 'E', 'SE', 'S', 'SW', 'W', 'NW', 'N']
            index = int((flight_yaw_degree + 22.5) / 45)
            flight_yaw_direction = directions[index]
            return flight_yaw_direction
        else:
            print_to_widget("XMP:FlightYawDegree not found in the image.")
            return None
    else:
        print_to_widget("No XMP data found in the image.")
        return None


def set_image_title(image_path, title):
    try:
        exif_data = piexif.load(image_path)
    except Exception as e:
        print_to_widget(f"Failed to load EXIF data from {image_path}: {e}")
        print_to_widget(traceback.format_exc())
        return
    exif_data['0th'][piexif.ImageIFD.ImageDescription] = title.encode('utf-8')
    exif_bytes = piexif.dump(exif_data)
    piexif.insert(exif_bytes, image_path)


def get_camera_maker(image_path):
    # Open the image file
    img = Image.open(image_path)

    # Get the exif data
    exif_data = img._getexif()

    # Loop through the exif data and look for the Camera Maker tag
    for tag_id, value in exif_data.items():
        tag_name = ExifTags.TAGS.get(tag_id, tag_id)
        if tag_name == 'Make':
            return value
    return None


def watermark_prep(base_directory):
    total_images = 0  # Initialize counter for total number of images
    for _, _, files in os.walk(base_directory):
        for file in files:
            if file.lower().endswith(('.png', '.jpg', '.jpeg')):
                total_images += 1

    images_processed = 0
    for root_dir, dirs, files in os.walk(base_directory):
        folder_name = os.path.basename(root_dir)
        print_to_widget(f"Processing folder: {folder_name}")
        for file in files:
            if file.lower().endswith(('.png', '.jpg', '.jpeg')):
                images_processed += 1
                image_path = os.path.join(root_dir, file)

                camera_maker = get_camera_maker(image_path)

                if camera_maker is None:
                    print_to_widget(f"No recognizable camera maker found for {file}.")
                    continue

                if camera_maker.lower() == 'sony':
                    direction_letters = get_gps_direction_letter(image_path)
                elif camera_maker.lower() == 'dji' or camera_maker.lower() == 'hasselblad':
                    direction_letters = get_yaw_direction_letter(image_path)
                else:
                    print_to_widget(f"\nUnknown camera maker for {file}: {camera_maker}.")
                    continue

                if direction_letters is None:
                    print_to_widget(f"No GPS image direction found for {file}.")
                else:
                    set_image_title(image_path, direction_letters)
                    print_to_widget(f"{file}'s Title set to {direction_letters}. {images_processed}/{total_images}")


# # function to apply watermark on an image
# def add_watermark(image, watermark_text, exif_bytes=None):
#     watermark = Image.new('RGBA', image.size, (0, 0, 0, 0))
#     font = ImageFont.truetype('calibri.ttf', 700)
#     draw = ImageDraw.Draw(watermark)
#     draw.textbbox((0, 0), watermark_text, font=font)
#     x = 200
#     y = 5400
#     text_pos = (x, y)
#     draw.text(text_pos, watermark_text, font=font, fill=(0, 0, 0, 230))
#     alpha = 0.3
#     image_with_watermark = Image.alpha_composite(image.convert('RGBA'), watermark)
#     output = BytesIO()
#     image_with_watermark.convert('RGB').save(output, format='JPEG', quality=100, exif=exif_bytes)
#     output.seek(0)
#     watermarked_image = Image.open(BytesIO(output.getvalue()))
#     return watermarked_image


def count_images(dir_path):
    # Count the total number of images in all folders within the directory and its subdirectories
    total_image_count = 0
    for subdir, _, files in os.walk(dir_path):
        for file in files:
            if file.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp')):
                total_image_count += 1
    return total_image_count


def folder_contains_images(folder_path):
    for subdir, _, files in os.walk(folder_path):
        for file in files:
            if file.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp')):
                return True
    return False


def find_folder_by_name(base_path, folder_name):
    for subdir, _, _ in os.walk(base_path):
        if folder_name in os.path.basename(subdir):
            return subdir
    return None


def package_data(file_path, dir_path):
    # Read the Excel file using pandas
    try:
        df = pd.read_excel(file_path, sheet_name='C2-Distribution')
    except ValueError:
        df = pd.read_excel(file_path, sheet_name='Known')
    # Iterate through the DataFrame rows
    for df_index, df_row in df.iterrows():
        # Get the folder name from the "SCE_STRUCT" column
        folder_name = df_row.loc["SCE_STRUCT"]
        no_oh_folder_name = folder_name.replace("OH-", "")
        if pd.isna(folder_name):
            continue

        # Get the new directory name from the "Photo_Location" column
        new_directory = df_row["Photo_Location"]
        if pd.isna(new_directory):
            print_to_widget(f"{no_oh_folder_name} is a refly. Skipping...")
            continue

        # Get the new root directory name
        new_new_directory = df_row["Photo_Location"].replace("Known", "")

        # Create the new "AssetPhotoLoc" directory if it doesn't exist
        os.makedirs(os.path.join(dir_path, new_new_directory, new_directory), exist_ok=True)

        # Set the source and destination paths
        source_path = find_folder_by_name(dir_path, str(no_oh_folder_name))
        destination_path = os.path.join(dir_path, new_new_directory, new_directory, str(no_oh_folder_name))

        # Move the folder to the new directory
        if source_path and os.path.exists(source_path):
            if folder_contains_images(source_path):
                shutil.move(source_path, destination_path)
                print_to_widget(f"Folder {no_oh_folder_name} moved to {new_new_directory}/{new_directory}")
            else:
                print_to_widget(f"Folder {no_oh_folder_name} does not contain images. Skipping...")
        else:
            print_to_widget(f"Folder {no_oh_folder_name} not found in source directory. Skipping...")

    # Count the total number of images in all folders within each directory in dir_path and its subdirectories
    grand_total = 0
    for directory in os.listdir(dir_path):
        directory_path = os.path.join(dir_path, directory)
        if os.path.isdir(directory_path):
            total_image_count = count_images(directory_path)
            print_to_widget(f"{directory} contains {total_image_count} images.")
            grand_total += total_image_count
    print_to_widget(f"Total images in all directories: {grand_total}")

    def move_images_to_parent_and_remove_folders(root_dir_path):
        for root_dir, dirs, files in os.walk(root_dir_path, topdown=False):
            for name in files:
                if name.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp')):
                    src_file = os.path.join(root_dir, name)
                    dst_file = os.path.join(os.path.dirname(root_dir), name)
                    if not os.path.exists(dst_file):  # Prevents overwriting files with the same name
                        shutil.move(src_file, dst_file)
            for name in dirs:
                subdir_path = os.path.join(root_dir, name)
                if not os.listdir(subdir_path):  # Check if the folder is empty
                    os.rmdir(subdir_path)

    # Prompt the user to confirm if they want to move all images to their parent directories
    message = ("The next step will remove the images from their structure folders. "
               "Verify that all images were named properly first.\n\n"
               "Click \"Yes\" to continue.\n"
               "Click \"No\" to manually remove them.\n")
    title = "Hold Up!"

    user_choice = messagebox.askyesno(title, message)

    if user_choice:
        # Call the move_images_to_parent_and_remove_folders function
        move_images_to_parent_and_remove_folders(dir_path)
        print_to_widget(
            f"All images have been removed from their structure folders and empty folders have been deleted.")
    else:
        print_to_widget(f"Images have NOT been removed from their respective structure folders.")


def package_helo_data(file_path, dir_path):
    # Read the Excel file using pandas
    try:
        df = pd.read_excel(file_path, sheet_name='FilteredTraveler')
    except ValueError:
        print_to_widget("'FilteredTraveler' sheet not found. Make sure 'Filtered_Helo_Traveler...' file is loaded.")
        df = None
    # Iterate through the DataFrame rows
    for df_index, df_row in df.iterrows():
        # Get the folder name from the "FLOC" column
        folder_name = df_row.loc["FLOC"]
        if pd.isna(folder_name):
            continue

        # Get the new path_directory name from the "Trans/Dist" column
        new_directory = df_row["Trans/Dist"]
        if pd.isna(new_directory):
            print_to_widget(f"{folder_name} is a refly. Skipping...")
            continue

        # Get the new root path_directory name
        new_new_directory = df_row["PhotoLoc"].replace("Known", "")
        new_new_new_directory = df_row["PhotoLoc"]

        # Create the new "Trans/Dist" path_directory if it doesn't exist
        os.makedirs(os.path.join(dir_path, new_directory, new_new_directory, new_new_new_directory), exist_ok=True)

        # Set the source and destination paths
        source_path = find_folder_by_name(dir_path, str(folder_name))
        destination_path = os.path.join(dir_path, new_directory, new_new_directory, new_new_new_directory,
                                        str(folder_name))

        # Move the folder to the new path_directory
        if source_path and os.path.exists(source_path):
            if folder_contains_images(source_path):
                shutil.move(source_path, destination_path)
                print_to_widget(
                    f"Folder {folder_name} moved to {new_new_new_directory}/{new_new_directory}/{new_directory}")
            else:
                print_to_widget(f"Folder {folder_name} does not contain images. Skipping...")
        else:
            print_to_widget(f"Folder {folder_name} not found in source path_directory. Skipping...")

    # Count the total number of images in all folders within each path_directory in dir_path and its subdirectories
    grand_total = 0
    for directory in os.listdir(dir_path):
        directory_path = os.path.join(dir_path, directory)
        if os.path.isdir(directory_path):
            total_image_count = count_images(directory_path)
            print_to_widget(f"{directory} contains {total_image_count} images.")
            grand_total += total_image_count
    print_to_widget(f"Total images in all directories: {grand_total}")

    def move_images_to_parent_and_remove_folders(path_directory):
        for dir_root, dirs, files in os.walk(path_directory, topdown=False):
            for name in files:
                if name.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp')):
                    src_file = os.path.join(dir_root, name)
                    dst_file = os.path.join(os.path.dirname(dir_root), name)
                    if not os.path.exists(dst_file):  # Prevents overwriting files with the same name
                        shutil.move(src_file, dst_file)
            for name in dirs:
                subdir_path = os.path.join(dir_root, name)
                if not os.listdir(subdir_path):  # Check if the folder is empty
                    os.rmdir(subdir_path)

    # Prompt the user to confirm if they want to move all images to their parent directories
    # Show the message box
    message = ("The next step will remove the images from their structure folders. "
               "Verify that all images were named properly first.\n\n"
               "Click \"Yes\" to continue.\n"
               "Click \"No\" to manually remove them.\n")
    title = "Hold Up!"

    user_choice = messagebox.askyesno(title, message)

    if user_choice:
        # Call the move_images_to_parent_and_remove_folders function
        move_images_to_parent_and_remove_folders(dir_path)
        print_to_widget(
            f"All images have been removed from their structure folders and empty folders have been deleted.")
    else:
        print_to_widget(f"Images have NOT been removed from their respective structure folders.")


def sort_images_by_structure_id(dir_path):
    # Iterate over all the files in the selected directory and its subdirectories
    for subdir, sub_dirs, files in os.walk(dir_path):
        for file in files:
            # Split the file name to get the StructureID
            structure_id = file.split('_')[0]

            # Create a new directory for the StructureID if it doesn't exist
            structure_dir = os.path.join(dir_path, structure_id)
            if not os.path.exists(structure_dir):
                os.makedirs(structure_dir)

            # Move the file to the appropriate directory
            old_file_path = os.path.join(subdir, file)
            new_file_path = os.path.join(structure_dir, file)
            if not os.path.exists(new_file_path):
                shutil.move(old_file_path, new_file_path)
            else:
                print_to_widget(f"File {file} already exists in {structure_dir}")

    # Iterate over all the subdirectories and delete the empty ones
    for root_dir, sub_dirs, files in os.walk(dir_path, topdown=False):
        for sub_dir in sub_dirs:
            dir_path = os.path.join(root_dir, sub_dir)
            if not os.listdir(dir_path):
                os.rmdir(dir_path)


# Function to calculate distance between two coordinates in feet
def distance_calculator(coord1, coord2):
    if coord1 is None or coord2 is None:
        raise ValueError("Invalid input coordinates: None")

    # Check if coordinates are strings
    if any(isinstance(c, str) for c in coord1) or any(isinstance(c, str) for c in coord2):
        return None

    if any(math.isnan(c) for c in coord1) or any(math.isnan(c) for c in coord2):
        raise ValueError("Invalid input coordinates: NaN")

    lat1, lon1 = coord1
    lat2, lon2 = coord2

    if not (-90 <= lat1 <= 90) or not (-90 <= lat2 <= 90):
        raise ValueError(f"Invalid latitude values: {lat1}, {lat2}")

    if not (-180 <= lon1 <= 180) or not (-180 <= lon2 <= 180):
        raise ValueError(f"Invalid longitude values: {lon1}, {lon2}")

    return round(geopy.distance.distance(coord1, coord2).feet, 2)


# Function to check if a row has missing values in required columns
def has_missing_values(m_row):
    # Check for column presence and if it has a missing value
    missing_mapped_lat = 'Mapped_Lat' in m_row and (pd.isna(m_row['Mapped_Lat']) or m_row['Mapped_Lat'] == 'N/A')
    missing_mapped_lon = 'Mapped_Lon' in m_row and (pd.isna(m_row['Mapped_Lon']) or m_row['Mapped_Lon'] == 'N/A')
    missing_field_lat = 'FieldLat' in m_row and (pd.isna(m_row['FieldLat']) or m_row['FieldLat'] == 'N/A')
    missing_field_long = 'FieldLong' in m_row and (pd.isna(m_row['FieldLong']) or m_row['FieldLong'] == 'N/A')

    # Return True if any of the columns have a missing value
    return missing_mapped_lat or missing_mapped_lon or missing_field_lat or missing_field_long


# Function to format a cell with the specified value, fill, and font
def format_cell(cell, value, fill, font):
    cell.value = value
    cell.fill = fill
    cell.font = font
    cell.alignment = Alignment(horizontal='right')


# Define a function to find the closest match for a folder name
def find_closest_match(folder_name, folder_path, choices, dataframe, verbose=True, resolve=False):
    # Try to find an exact match first
    choice_match = None
    for choice in choices:
        if folder_name == choice and verbose:
            choice_match = choice
            print_to_widget(f"{folder_name} found.")
            choices.remove(choice)  # Remove the matched option
            break

    if choice_match is not None:
        return folder_name, choice_match, choices

    # If an exact match is not found, proceed with the following steps
    # Find the image that ends with n.jpg
    for file in os.listdir(folder_path):
        if file.lower().endswith("n.jpg"):
            image_path = os.path.join(folder_path, file)
            break
    else:
        # If no image found, use fuzzywuzzy to find the closest match
        closest_match = process.extractOne(folder_name, choices)[0]
        choices.remove(closest_match)  # Remove the matched option
        if verbose:
            print_to_widget(f"{os.path.basename(folder_path)} not found in GIS. Nadir image no 'N'."
                            f" Finding closest ID match instead...")
            print_to_widget(f"Closest ID match: {closest_match}")
        return folder_name, closest_match, choices

    # Get GPS coordinates from the image
    lat, lon = get_gps_from_image(image_path)
    if lat is None or lon is None:
        closest_match = process.extractOne(folder_name, choices)[0]
        choices.remove(closest_match)  # Remove the matched option
        if verbose:
            print_to_widget(f"{os.path.basename(folder_path)} not found in GIS. Nadir image does not have gps data."
                            f" Finding closest ID match instead...")
            print_to_widget(f"Closest ID match: {closest_match}")
        return folder_name, choice_match, choices

    # Calculate the distance between the image's coordinates and each structure coordinate in the date_taken_df
    distances = []
    for dataframe_index, dataframe_row in dataframe.iterrows():
        if pd.isna(dataframe_row['Mapped_Lat']) or pd.isna(dataframe_row['Mapped_Lon']):
            continue
        dist_col_name = "Structure_" if "Structure_" in dataframe.columns else "FLOC"
        mapped_lat, mapped_lon = dataframe_row["Mapped_Lat"], dataframe_row["Mapped_Lon"]
        distance = distance_calculator((lat, lon), (mapped_lat, mapped_lon))
        distances.append((dataframe_row[dist_col_name], distance))

    # Find the closest distance and the associated structure ID
    if distances:
        closest_match = min(distances, key=lambda x: x[1])[0]
        closest_distance = min(distances, key=lambda x: x[1])[1]
        if resolve:
            choices.remove(closest_match)  # Remove the matched option
        if verbose:
            print_to_widget(f"{folder_name} not found in GIS. Closest nadir match: {closest_match}. Distance from "
                            f"nadir: {closest_distance} feet.")
        return folder_name, closest_match, choices


def resolve_duplicates(matches, available_choices, structure_dict, dataframe):
    resolved_duplicates = {}
    while any(isinstance(value, list) for value in matches.values()):
        duplicates = {key: value for key, value in matches.items() if isinstance(value, list)}
        new_matches = {key: value for key, value in matches.items() if not isinstance(value, list)}

        for match, folders in duplicates.items():
            # Compute scores for all folders against the match
            scores = {folder: fuzz.ratio(folder, match) for folder in folders}

            # Identify the most accurate folder (highest score)
            most_accurate_folder = max(scores, key=scores.get)
            new_matches[match] = most_accurate_folder
            resolved_duplicates[match] = most_accurate_folder
            available_choices.remove(match)

            # All other folders will use find_closest_match
            other_folders = [folder for folder in folders if folder != most_accurate_folder]
            for folder in other_folders:
                # Get the folder path from the dictionary
                folder_path = structure_dict[folder]

                # Filter the dataframe based on the available options in 'FLOC'
                filtered_dataframe = dataframe[(dataframe['FLOC'].isin(available_choices))]
                filtered_dataframe = filtered_dataframe[filtered_dataframe['FLOC'] != most_accurate_folder]
                _, closest_match, available_options = find_closest_match(folder, folder_path, available_choices,
                                                                         filtered_dataframe, verbose=False,
                                                                         resolve=True)
                if closest_match in new_matches and closest_match != match:
                    if isinstance(new_matches[closest_match], list):
                        new_matches[closest_match].append(folder)
                    else:
                        new_matches[closest_match] = [new_matches[closest_match], folder]
                else:
                    new_matches[closest_match] = folder
                    resolved_duplicates[closest_match] = folder

        matches = new_matches  # Update matches for the next iteration

    print_to_widget("\nNew closest nadir matches found:")
    for match, folder in resolved_duplicates.items():
        print_to_widget(f"{folder} --> {match}")
    print_to_widget("Make sure that these are correct before proceeding to your next step.")
    return matches, available_choices


def display_duplicates_and_get_selection(duplicates_df, columns_to_display):
    duplicates_root = ctk.CTk()
    duplicates_root.title("Select Row to KEEP")

    duplicates_frame = ctk.CTkFrame(duplicates_root)
    frame.pack(padx=10, pady=10)

    # Continue using ttk.Treeview as there is no ctk equivalent
    tree = ttk.Treeview(duplicates_frame, columns=columns_to_display, show="headings")
    tree.pack(side="left")

    for column in columns_to_display:
        tree.heading(column, text=column)
        tree.column(column, width=100)

    scrollbar = ctk.CTkScrollbar(duplicates_frame, command=tree.yview)
    tree.configure(yscrollcommand=scrollbar.set)
    scrollbar.pack(side="right", fill="y")

    for duplicates_idx, duplicates_row in duplicates_df.iterrows():
        values_to_insert = [duplicates_row[duplicates_col] for duplicates_col in columns_to_display]
        tree.insert("", "end", values=values_to_insert, tags=(duplicates_idx,))

    selected_idx = tk.StringVar()

    def on_ok():
        selected_item = tree.selection()[0]
        selected_idx.set(tree.item(selected_item, "tags")[0])
        duplicates_root.destroy()

    ok_button = ctk.CTkButton(duplicates_root, text="OK", command=on_ok)
    ok_button.pack(pady=10)

    duplicates_root.mainloop()

    return selected_idx.get()


# Function to find the column index based on the header name
def find_col_index_by_header(ws, header_name):
    for column in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=column).value == header_name:
            return column
    return None


options_match = None


# Function to filter distro extract
def filter_extract_distro(file_path, dir_path):
    # Delete hidden Mac files
    global options_match
    print_to_widget(f"\nDeleting hidden Mac files...")
    found_mac_files = False
    for dirpath, dirnames, filenames in os.walk(dir_path):
        for filename in filenames:
            if filename.startswith("._"):
                filepath = os.path.join(dirpath, filename)
                os.remove(filepath)
                print_to_widget(f"{filename} deleted.")
                found_mac_files = True
    if not found_mac_files:
        print_to_widget("No hidden Mac files found in the folder.")
    else:
        print_to_widget(f"All hidden Mac files have been deleted.")

    structure_names = []
    structure_paths = []
    structure_dict = {}

    print_to_widget("\nExtracting folder names and paths...")
    for dirpath, dirs, files in os.walk(dir_path):
        # Check if any of the files in the directory are image files
        if any(file.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp')) for file in files):
            structure_name = os.path.basename(dirpath)
            structure_path = dirpath  # Full directory path

            # Add to the lists
            structure_names.append(structure_name)
            structure_paths.append(structure_path)

            # Add to the dictionary
            structure_dict[structure_name] = structure_path

    # Read the Excel file into a pandas dataframe
    print_to_widget(f"\nCounting images and extracting GPS data for each image...")
    df = pd.read_excel(file_path)

    # Check if the Structure_ID column is present in the dataframe
    if 'FLOC' not in df.columns:
        print_to_widget(
            "Error: \"FLOC\" column not found in Excel file. Make sure the right DISTRIBUTION extract is "
            "selected. Rerun the script...")
        sys.exit(1)
    else:
        # Make a copy of the dataframe before filtering
        df_copy = df.copy()

        # Remove the first three characters from each cell in the FLOC column
        df_copy['FLOC'] = df_copy['FLOC'].str.replace('^OH-', '', regex=True)

        # Create a dictionary to store the photo count for each folder
        photo_count = {}

        # Create a dictionary to store the latitude and longitude of image for each folder
        latlong_date_N = {}

        # Loop through each folder in the directory
        for folder in structure_paths:
            folder_name = os.path.basename(folder)  # Extract the name of the folder from the path
            # Get the list of files in the folder
            files = os.listdir(folder)
            # Count the number of files in the folder
            count = len(files)
            # Add the photo count to the dictionary
            photo_count[folder_name] = count
            # Initialize the lat and long variables to N/A.
            lat = 'N/A'
            lng = 'N/A'
            date_taken = 'N/A'
            # Initialize variable to track whether an "N" image was found
            n_found = False
            # Loop through each file in the folder
            print_to_widget(f"{folder_name} has {count} images.")
            for file in files:
                # Check if the file ends with "N" and is an image file
                if file.lower().endswith(("n.jpg", "n.jpeg", "n.png", "n.bmp", "n.JPG")):
                    # Get the full path of the image file
                    image_path = os.path.join(folder, file)
                    # Extract the date taken from the image file
                    date_taken = get_date_taken(image_path)
                    # Extract the GPS coordinates from the image file
                    coords = get_gps_from_image(image_path)
                    # Update the lat/long variables with the GPS coordinates, or blank if no GPS coordinates were found
                    if coords:
                        lat = coords[0]
                        lng = coords[1]
                    else:
                        lat = 'Metadata Issue'
                        lng = 'Metadata Issue'
                        print_to_widget(f"Warning: {folder_name}'s nadir has no GPS coordinates.\n")
                    # Set variable to True if an "N" image is found
                    n_found = True
            # Print a message if no "N" image was found
            if not n_found:
                print_to_widget(f"Warning: {folder_name} has no image that ends with \"N\".")
                lat = 'Nadir missing N'
                lng = 'Nadir missing N'

                # Loop again to find the first image with a date
                for file in files:
                    # Ensure we only consider image files (using the same extensions as before)
                    if file.lower().endswith((".jpg", ".jpeg", ".png", ".bmp", ".JPG")):
                        image_path = os.path.join(folder, file)
                        date_taken = get_date_taken(image_path)

                        # Check if a date was found
                        if date_taken != 'N/A':
                            break
                else:
                    # No image with a date was found
                    print_to_widget(f"Warning: {folder_name} has no date taken metadata.\n")

            # Add the lat, long, and date taken to the dictionary
            latlong_date_N[folder_name] = (lat, lng, date_taken)

        # Find the closest match for each folder name in the dataframe
        # Filter df by flight dates
        df_copy['Inspection_Date1'] = pd.to_datetime(df_copy['Inspection_Date1'])
        df_copy['Inspection_Date2'] = pd.to_datetime(df_copy['Inspection_Date2'])
        # Subtract 8 hours using pd.Timedelta
        df_copy['Inspection_Date1'] -= pd.Timedelta(hours=8)
        df_copy['Inspection_Date2'] -= pd.Timedelta(hours=8)
        # Extract date taken values from latlong_date_N dictionary
        date_taken_values = [details[2] for details in latlong_date_N.values()]
        date_taken_dates = [pd.to_datetime(date_taken).date() for date_taken in date_taken_values]
        # Filter dataframe based on adjusted inspection dates
        date_taken_df = df_copy[
            ((df_copy['Inspection_Date1'].dt.date.isin(date_taken_dates)) |
             (df_copy['Inspection_Date2'].dt.date.isin(date_taken_dates)))
        ]
        # Dictionary to store matches
        matches = {}
        folder_name_options = list(df_copy['FLOC'])  # Convert to list for manipulation
        print_to_widget("\nVerifying if folder names exist in GIS...")
        for structure_name, structure_path in zip(structure_names, structure_paths):
            result = find_closest_match(structure_name, structure_path, folder_name_options, date_taken_df)
            if result is None:
                print_to_widget(f"No match found for {structure_name}. Make sure it is part of distribution scope.")
                continue  # Skip to the next iteration of the loop, or use 'return' or 'break' as needed
            else:
                structure_name_matched, closest_match, options_match = result

            # Check if match already exists
            if closest_match in matches:
                if not isinstance(matches[closest_match], list):
                    matches[closest_match] = [matches[closest_match]]
                matches[closest_match].append(structure_name_matched)
            else:
                matches[closest_match] = structure_name_matched

        # Create a temporary duplicates dictionary for printing purposes
        duplicates = {key: value for key, value in matches.items() if isinstance(value, list)}

        if duplicates:
            output = "\nThe same match(es) found for the following folders:"
            for key, values in duplicates.items():
                output += f"\n{', '.join(values)} --> {key}"

            print_to_widget(f"{output}\nFinding new matches...")

            # Resolve duplicates
            matches, available_options = resolve_duplicates(matches, options_match, structure_dict, date_taken_df)

        # Filter the dataframe based on the closest matches
        print_to_widget("\nFiltering extract based on found structure ID matches...")
        closest_matches = list(matches.keys())
        filtered_df = df_copy[df_copy['FLOC'].isin(closest_matches)].copy()
        filtered_df = filtered_df.assign(
            closest_match=pd.Categorical(filtered_df['FLOC'], categories=closest_matches, ordered=True))
        filtered_df = filtered_df.sort_values('closest_match')
        filtered_df.drop('closest_match', axis=1, inplace=True)

        # Check for duplicates in the extract's 'FLOC' column
        duplicates = filtered_df.duplicated('FLOC', keep=False)
        columns_to_show = ['FLOC', 'Eq_ObjType', 'Block_ID', 'Team_Number', 'Inspection_Date1', 'Inspection_Date2']

        if any(duplicates):
            duplicate_values = filtered_df.loc[duplicates, 'FLOC'].unique()
            print_to_widget(
                f"Warning: The following Structure IDs have duplicates in the extract: {duplicate_values}."
                f" Please choose which row(s) to keep.")
            for dup_value in duplicate_values:
                duplicate_rows = filtered_df[filtered_df['FLOC'] == dup_value]
                selected_index = display_duplicates_and_get_selection(duplicate_rows, columns_to_show)

                # Drop all rows except the selected one
                drop_indices = [idx for idx in duplicate_rows.index if idx != int(selected_index)]
                filtered_df.drop(drop_indices, inplace=True)

            print_to_widget("All extract duplicates have been effectively removed.")

        print_to_widget("\nAdding new columns to the extract...")

        folder_names_update = list(matches.values())

        # Add a new column to the filtered dataframe indicating the folder name for each row's "FLOC" value
        filtered_df.insert(filtered_df.columns.get_loc('FLOC') + 1, 'FolderName', folder_names_update)

        # # Sort table in ascending order by Structure ID
        print_to_widget(f"\nSorting extract in ascending Structure ID order...")
        filtered_df.sort_values(by=["FLOC"], inplace=True)

        # Add a new column to the filtered dataframe indicating the photo count for each row's "FLOC" value
        filtered_df.insert(filtered_df.columns.get_loc('FLOC') + 2, 'PhotoCount',
                           filtered_df['FolderName'].apply(lambda x: photo_count[x]))

        # Add a new column to the filtered dataframe indicating the latitude for each row's "FLOC" value
        filtered_df.insert(filtered_df.columns.get_loc('FLOC') + 3, 'FieldLat',
                           filtered_df['FolderName'].apply(lambda x: latlong_date_N[x][0]))

        # Add a new column to the filtered dataframe indicating the longitude for each row's "FLOC" value
        filtered_df.insert(filtered_df.columns.get_loc('FLOC') + 4, 'FieldLong',
                           filtered_df['FolderName'].apply(lambda x: latlong_date_N[x][1]))

        # Add "Team " to Team_Number values
        filtered_df['Team_Number'] = filtered_df['Team_Number'].apply(
            lambda x: 'Team ' + str(int(x)) if pd.notna(x) else 'Team ')

        # Add a new column to the filtered dataframe indicating the flight date for each row's "FLOC" value
        filtered_df['FlightDate'] = filtered_df['FolderName'].apply(
            lambda x: datetime.strptime(latlong_date_N[x][2], '%Y%m%d').date()
            if x in latlong_date_N and latlong_date_N[x][2] != 'N/A' else 'N/A')
        filtered_df.insert(filtered_df.columns.get_loc('Team_Number') + 1, 'FlightDate', filtered_df.pop('FlightDate'))

        # Add a new column to the filtered dataframe indicating the photo location for each row's "FLOC" value
        filtered_df.insert(filtered_df.columns.get_loc('Team_Number') + 1, 'PhotoLoc',
                           filtered_df['Block_ID'].astype(str) + '_KnownAssets_' + pd.to_datetime(
                               filtered_df['FlightDate'], errors='coerce').dt.strftime('%Y%m%d').fillna(''))

        # Calculate distances and add them as a new column to the DataFrame
        print_to_widget("\nCalculating structure distance from GIS location...")
        distances = []
        for index, filtered_df_row in filtered_df.iterrows():
            if not has_missing_values(filtered_df_row):
                dist = distance_calculator((filtered_df_row['Mapped_Lat'], filtered_df_row['Mapped_Lon']),
                                           (filtered_df_row['FieldLat'], filtered_df_row['FieldLong']))
                if dist is not None and not math.isnan(dist):
                    distances.append(round(dist, 2))
                else:
                    distances.append('N/A')
            else:
                distances.append('N/A')
        filtered_df.insert(filtered_df.columns.get_loc('Mapped_Lon') + 1, 'DistFromGIS (ft)', distances)

        # Calculate the farthest image distance from the nadir as a new column to the DataFrame
        print_to_widget("\nCalculating the farthest distance between images and the nadir...")
        nadir_distances = get_farthest_from_nadir(dir_path)
        filtered_df.insert(filtered_df.columns.get_loc('DistFromGIS (ft)') + 1, 'FarthestDistFromNadir (ft)',
                           filtered_df['FolderName'].apply(
                               lambda x: round(nadir_distances.get(x), 2) if nadir_distances.get(
                                   x) is not None else 'N/A'))

        # Extract the directory path from dir_path
        directory = os.path.dirname(file_path)

        # Extract the base name of the file without extension
        file_name = os.path.splitext(os.path.basename(file_path))[0]

        # Search for a date in the file name using regex
        match = re.search(r"_(\d{4}\d{2}\d{2})_|(\d{4}\d{2}\d{2})$", file_name)

        # If a date is found, parse it to a datetime object
        if match:
            file_date = match.group(0)

        else:
            # If no date is found, use the current date as a fallback
            file_date = datetime.now().strftime("%Y%m%d")

        # Now you can use file_date to name your new file
        new_file_name = f"D_Filtered_GIS_Extract_{file_date}.xlsx"

        # Construct the new file path for saving
        new_file_path = os.path.join(directory, new_file_name)

        # Initialize the counter
        counter = 1

        # Check if the file already exists
        while os.path.exists(new_file_path):
            # Append the counter to the base file name
            modified_file_name = f"{new_file_name.split('.')[0]} ({counter}).{new_file_name.split('.')[-1]}"
            new_file_path = os.path.join(directory, modified_file_name)
            counter += 1

        # Write the filtered dataframe to the original Excel file
        filtered_df.to_excel(new_file_path, index=False, sheet_name='FilteredExtract')

        # Open the Excel file with openpyxl and format the cells in the "Match" and "FLOC" columns that are False
        wb = load_workbook(new_file_path)
        ws = wb.active

        # Delete unnecessary columns from the worksheet
        print_to_widget("\nDeleting unnecessary columns...")
        # List of header names to delete
        headers_to_delete = [
            "FID", "AOC_Area_Name", "Inspection", "Vendor_ID", "Scope_Area", "XY_Source", "ODI_Bundle", "Agency_Nam",
            "Agency_Are", "ESA", "ObjectId", "GlobalID", "x", "y"
        ]

        # Find indices and delete columns
        for header in headers_to_delete:
            col_index = find_col_index_by_header(ws, header)
            if col_index:
                ws.delete_cols(col_index)

        # Define a PatternFill object
        light_red_fill = PatternFill(start_color='FFFFC7CE', end_color='FFFFC7CE', fill_type='solid')
        dark_red_font = Font(color='9C0006')

        # Find the index of the 'Distance', 'FieldLat', and 'FieldLong' columns
        distance_column_index = find_col_index_by_header(ws, "DistFromGIS (ft)")
        field_lat_column_index = find_col_index_by_header(ws, "FieldLat")
        field_long_column_index = find_col_index_by_header(ws, "FieldLong")
        nadir_max_distance_col = find_col_index_by_header(ws, "FarthestDistFromNadir (ft)")
        photo_count_col_index = find_col_index_by_header(ws, "PhotoCount")

        # Apply cell formatting to the appropriate cells based on the specified conditions
        print_to_widget("\nHighlighting potential issues...")
        for ws_row in range(2, ws.max_row + 1):
            distance_cell = ws.cell(row=ws_row, column=distance_column_index)
            field_lat_cell = ws.cell(row=ws_row, column=field_lat_column_index)
            field_long_cell = ws.cell(row=ws_row, column=field_long_column_index)
            farthest_cell = ws.cell(row=ws_row, column=nadir_max_distance_col)
            photo_count_cell = ws.cell(row=ws_row, column=photo_count_col_index)
            distance_value = distance_cell.value
            if distance_value == 'N/A':
                format_cell(field_lat_cell, field_lat_cell.value, light_red_fill, dark_red_font)
                format_cell(field_long_cell, field_long_cell.value, light_red_fill, dark_red_font)
                format_cell(distance_cell, distance_value, light_red_fill, dark_red_font)
            elif distance_value > 200:
                format_cell(distance_cell, distance_value, light_red_fill, dark_red_font)
            else:
                farthest_cell.alignment = Alignment(horizontal='right')
                distance_cell.alignment = Alignment(horizontal='right')
            if farthest_cell.value == 'N/A' or int(farthest_cell.value) > 300:
                format_cell(farthest_cell, farthest_cell.value, light_red_fill, dark_red_font)
            if photo_count_cell.value == 0:
                format_cell(photo_count_cell, photo_count_cell.value, light_red_fill, dark_red_font)

            # Highlight cells that are not a matching
            for ws_iter_row in ws.iter_rows(min_row=2, min_col=1, max_col=33):
                date1 = ws_iter_row[22].value
                date2 = ws_iter_row[25].value
                if not isinstance(date1, datetime):
                    date1 = datetime.strptime(date1, '%Y-%m-%d')
                if not isinstance(date2, datetime):
                    date2 = datetime.strptime(date2, '%Y-%m-%d %H:%M:%S')
                # Compare folder name and GIS structure ID
                if ws_iter_row[0].value != ws_iter_row[1].value:
                    ws_iter_row[0].fill = light_red_fill
                    ws_iter_row[0].font = dark_red_font
                    ws_iter_row[1].fill = light_red_fill
                    ws_iter_row[1].font = dark_red_font
                # Compare flight date and inspection date
                if date1.date() != date2.date():
                    ws_iter_row[25].fill = light_red_fill
                    ws_iter_row[25].font = dark_red_font

        # Save the modified Excel file
        wb.save(new_file_path)

        # Create a new DataFrame to compare GIS Structure IDs and Folder Names
        print_to_widget("\nAdding new tabs for each team...")

        # Filter the GIS data based on 'Inspection_Date1' or 'Inspection_Date2' and 'Team_Number'
        filtered_df_drop = filtered_df.copy()
        filtered_df_drop['Team_Number'] = filtered_df_drop['Team_Number'].astype(str)
        filtered_df_drop['Team_Float'] = filtered_df_drop['Team_Number'].str.extract(r'(\d+)').astype(float)

        # Maintain a list of created sheet names
        created_sheet_names = []

        with pd.ExcelWriter(new_file_path, engine='openpyxl', mode='a') as writer:
            # Iterate over each unique team number
            for team in filtered_df_drop['Team_Float'].unique():
                # Filter data for the current team
                filtered_df_team = filtered_df_drop[filtered_df_drop['Team_Float'] == team]
                mask_team = (df_copy['Team_Number'] == team)

                df_copy['Inspection_Date'] = df_copy['Inspection_Date2'].combine_first(df_copy['Inspection_Date1'])
                # Extract only the date part from the datetime objects
                df_copy_date_only = df_copy['Inspection_Date'].dt.date
                mask_date = df_copy_date_only.isin(filtered_df_team['FlightDate'])
                gis_df = df_copy[mask_team & mask_date].copy()

                filtered_df_sorted = filtered_df_team.iloc[
                    filtered_df_team['FolderName'].apply(natural_sort_key).argsort()]
                gis_df_sorted = gis_df.iloc[gis_df['FLOC'].apply(natural_sort_key).argsort()]

                structure_ids_df = pd.DataFrame({
                    'Folder_ID': filtered_df_sorted['FolderName'],
                    'GIS_ID': gis_df_sorted['FLOC'],
                    'Team_Number': gis_df_sorted['Team_Number'],
                    'Inspection_Date': gis_df_sorted['Inspection_Date2'].combine_first(gis_df_sorted[
                                                                                           'Inspection_Date1']),
                    'Mapped_Lat': gis_df_sorted['Mapped_Lat'],
                    'Mapped_Long': gis_df_sorted['Mapped_Lon'],
                    'Vendor_Notes': gis_df_sorted['Vendor_Notes'],
                    'Existing_Notes': gis_df_sorted['Existing_Notes']
                })
                structure_ids_df.sort_values(by=["GIS_ID"], inplace=True)
                # Write the filtered dataframe for the current team to Excel
                try:
                    sheet_name = f"Team {int(team)}"
                except (ValueError, TypeError):  # Handle both empty strings and None values, among others
                    sheet_name = "Team NaN"
                structure_ids_df.to_excel(writer, index=False, sheet_name=sheet_name)
                created_sheet_names.append(sheet_name)

        # Load workbook
        wb = load_workbook(new_file_path)

        for sheet_name in created_sheet_names:  # Loop through only the recently created sheets
            ws = wb[sheet_name]

            # The rest remains the same...
            folder_ids = set(cell.value for cell in ws['A'] if cell.value is not None)
            gis_ids = set(cell.value for cell in ws['B'] if cell.value is not None)
            unmatched_folder_ids = folder_ids - gis_ids
            unmatched_gis_ids = gis_ids - folder_ids

            for cell in ws['A']:
                if cell.row != 1 and cell.value in unmatched_folder_ids:
                    cell.fill = light_red_fill
                    cell.font = dark_red_font

            for cell in ws['B']:
                if cell.row != 1 and cell.value in unmatched_gis_ids:
                    cell.fill = light_red_fill
                    cell.font = dark_red_font

        # Save the workbook
        wb.save(new_file_path)

    return new_file_path


# Function to filter trans extract
def filter_extract_trans(file_path, dir_path):
    # Delete hidden Mac files
    global options_match
    print_to_widget(f"\nDeleting hidden Mac files...")
    found_mac_files = False
    for dirpath, dirnames, filenames in os.walk(dir_path):
        for filename in filenames:
            if filename.startswith("._"):
                filepath = os.path.join(dirpath, filename)
                os.remove(filepath)
                print_to_widget(f"{filename} deleted.")
                found_mac_files = True
    if not found_mac_files:
        print_to_widget("No hidden Mac files found in the folder.")
    else:
        print_to_widget(f"All hidden Mac files have been deleted.")

    structure_names = []
    structure_paths = []
    structure_dict = {}

    print_to_widget("\nExtracting team numbers based on the folder names...")
    for dir_root, dirs, files in os.walk(dir_path):
        # Check if any of the files in the directory are image files
        if any(file.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp')) for file in files):
            structure_name = os.path.basename(dir_root)  # Name of the current root
            structure_path = dir_root  # Full directory path

            # Add to the lists
            structure_names.append(structure_name)
            structure_paths.append(structure_path)

            # Add to the dictionary
            structure_dict[structure_name] = structure_path

    # Read the Excel file into a pandas dataframe
    print_to_widget(f"\nCounting images and extracting GPS data...")
    df = pd.read_excel(file_path)

    # Check if the FLOC column is present in the dataframe
    if 'FLOC' not in df.columns:
        print_to_widget(
            "Error: \"FLOC\" column not found in Excel file. Make sure the right TRANSMISSION extract is selected."
            " Rerun the script...")
        sys.exit(1)
    else:
        # Make a copy of the dataframe before filtering
        df_copy = df.copy()

        # Remove the first three characters from each cell in the FLOC column
        df_copy['FLOC'] = df_copy['FLOC'].str.replace('^OH-', '', regex=True)

        # Create a dictionary to store the photo count for each folder
        photo_count = {}

        # Create a dictionary to store the latitude and longitude of image for each folder
        latlong_date_N = {}

        # Loop through each folder in the directory
        for folder in structure_paths:
            folder_name = os.path.basename(folder)  # Extract the name of the folder from the path
            # Get the list of files in the folder
            files = os.listdir(folder)
            # Count the number of files in the folder
            count = len(files)
            # Add the photo count to the dictionary
            photo_count[folder_name] = count
            # Initialize the lat and long variables to N/A.
            lat = 'N/A'
            lng = 'N/A'
            date_taken = 'N/A'
            # Initialize variable to track whether an "N" image was found
            n_found = False
            # Loop through each file in the folder
            print_to_widget(f"{folder_name} has {count} images.")
            for file in files:
                # Check if the file ends with "N" and is an image file
                if file.lower().endswith(("n.jpg", "n.jpeg", "n.png", "n.bmp", "n.JPG")):
                    # Get the full path of the image file
                    image_path = os.path.join(folder, file)
                    # Extract the date taken from the image file
                    date_taken = get_date_taken(image_path)
                    # Extract the GPS coordinates from the image file
                    coords = get_gps_from_image(image_path)
                    # Update the lat and long variables with the GPS coordinates, or blank if no GPS found
                    if coords:
                        lat = coords[0]
                        lng = coords[1]
                    else:
                        lat = 'Metadata Issue'
                        lng = 'Metadata Issue'
                        print_to_widget(f"Warning: {folder_name}'s nadir has no GPS coordinates.\n")
                    # Set variable to True if an "N" image is found
                    n_found = True
            # Print a message if no "N" image was found
            if not n_found:
                print_to_widget(f"Warning: {folder_name} has no image that ends with \"N\".")
                lat = 'Nadir missing N'
                lng = 'Nadir missing N'

                # Loop again to find the first image with a date
                for file in files:
                    # Ensure we only consider image files (using the same extensions as before)
                    if file.lower().endswith((".jpg", ".jpeg", ".png", ".bmp", ".JPG")):
                        image_path = os.path.join(folder, file)
                        date_taken = get_date_taken(image_path)

                        # Check if a date was found
                        if date_taken != 'N/A':
                            break
                else:
                    # No image with a date was found
                    print_to_widget(f"Warning: {folder_name} has no image with a date.\n")

            # Add the lat, long, and date taken to the dictionary
            latlong_date_N[folder_name] = (lat, lng, date_taken)

        # Find the closest match for each folder name in the dataframe
        # Filter df by flight dates
        df_copy['Inspection_Date1'] = pd.to_datetime(df_copy['Inspection_Date1'])
        df_copy['Inspection_Date2'] = pd.to_datetime(df_copy['Inspection_Date2'])
        # Subtract 8 hours using pd.Timedelta
        df_copy['Inspection_Date1'] -= pd.Timedelta(hours=8)
        df_copy['Inspection_Date2'] -= pd.Timedelta(hours=8)
        # Extract date taken values from latlong_date_N dictionary
        date_taken_values = [details[2] for details in latlong_date_N.values()]
        date_taken_dates = [pd.to_datetime(date_taken).date() for date_taken in date_taken_values]
        # Filter dataframe based on adjusted inspection dates
        date_taken_df = df_copy[
            ((df_copy['Inspection_Date1'].dt.date.isin(date_taken_dates)) |
             (df_copy['Inspection_Date2'].dt.date.isin(date_taken_dates)))
        ]
        # Dictionary to store matches
        matches = {}
        folder_name_options = list(df_copy['FLOC'])  # Convert to list for manipulation
        print_to_widget("\nVerifying if folder names exist in GIS...")
        for structure_name, structure_path in zip(structure_names, structure_paths):
            structure_name_matched, closest_match, options_match = (
                find_closest_match(structure_name, structure_path, folder_name_options, date_taken_df))

            # Check if match already exists
            if closest_match in matches:
                if not isinstance(matches[closest_match], list):
                    matches[closest_match] = [matches[closest_match]]
                matches[closest_match].append(structure_name_matched)
            else:
                matches[closest_match] = structure_name_matched

        # Create a temporary duplicates dictionary for printing purposes
        duplicates = {key: value for key, value in matches.items() if isinstance(value, list)}

        if duplicates:
            output = "\nThe same match(es) found for the following folders:"
            for key, values in duplicates.items():
                output += f"\n{', '.join(values)} --> {key}"

            print_to_widget(f"{output}\nFinding new matches...")

            # Resolve duplicates
            matches, available_options = resolve_duplicates(matches, options_match, structure_dict, date_taken_df)

        # Filter the dataframe based on the closest matches
        print_to_widget("\nFiltering extract based on found structure ID matches...")
        closest_matches = list(matches.keys())
        filtered_df = df_copy[df_copy['FLOC'].isin(closest_matches)].copy()
        filtered_df = filtered_df.assign(
            closest_match=pd.Categorical(filtered_df['FLOC'], categories=closest_matches, ordered=True))
        filtered_df = filtered_df.sort_values('closest_match')
        filtered_df.drop('closest_match', axis=1, inplace=True)

        # Check for duplicates in the extract's 'FLOC' column
        duplicates = filtered_df.duplicated('FLOC', keep=False)
        columns_to_show = ['FLOC', 'Eq_ObjType', 'Block_ID', 'Team_Number', 'Inspection_Date1', 'Inspection_Date2']

        if any(duplicates):
            duplicate_values = filtered_df.loc[duplicates, 'FLOC'].unique()
            print_to_widget(
                f"Warning: The following Structure IDs have duplicates in the extract: {duplicate_values}."
                f" Please choose which row(s) to KEEP.")
            for dup_value in duplicate_values:
                duplicate_rows = filtered_df[filtered_df['FLOC'] == dup_value]
                selected_index = display_duplicates_and_get_selection(duplicate_rows, columns_to_show)

                # Drop all rows except the selected one
                drop_indices = [idx for idx in duplicate_rows.index if idx != int(selected_index)]
                filtered_df.drop(drop_indices, inplace=True)

            print_to_widget("All extract duplicates have been effectively removed.")

        print_to_widget("\nAdding new columns to the extract...")

        folder_names_update = list(matches.values())

        # Add a new column to the filtered dataframe indicating the folder name for each row's "FLOC" value
        filtered_df.insert(filtered_df.columns.get_loc('FLOC') + 1, 'FolderName', folder_names_update)

        # Sort table in ascending order by Structure ID
        print_to_widget(f"\nSorting extract in ascending Structure ID order...")
        filtered_df.sort_values(by=["FLOC"], inplace=True)

        # Add a new column to the filtered dataframe indicating the photo count for each row's "FLOC" value
        filtered_df.insert(filtered_df.columns.get_loc('FLOC') + 2, 'PhotoCount',
                           filtered_df['FolderName'].apply(lambda x: photo_count[x]))

        # Add a new column to the filtered dataframe indicating the latitude for each row's "FLOC" value
        filtered_df.insert(filtered_df.columns.get_loc('FLOC') + 3, 'FieldLat',
                           filtered_df['FolderName'].apply(lambda x: latlong_date_N[x][0]))

        # Add a new column to the filtered dataframe indicating the longitude for each row's "FLOC" value
        filtered_df.insert(filtered_df.columns.get_loc('FLOC') + 4, 'FieldLong',
                           filtered_df['FolderName'].apply(lambda x: latlong_date_N[x][1]))

        # Add "Team " to Team_Number values
        filtered_df['Team_Number'] = filtered_df['Team_Number'].apply(
            lambda x: 'Team ' + str(int(x)) if pd.notna(x) else 'Team ')

        # Add a new column to the filtered dataframe indicating the flight date for each row's "FLOC" value
        filtered_df['FlightDate'] = filtered_df['FolderName'].apply(
            lambda x: datetime.strptime(latlong_date_N[x][2], '%Y%m%d').date()
            if x in latlong_date_N and latlong_date_N[x][2] != 'N/A' else 'N/A')
        filtered_df.insert(filtered_df.columns.get_loc('Team_Number') + 1, 'FlightDate', filtered_df.pop('FlightDate'))

        # Add a new column to indicating the KnownAsset photo location for each row's "FLOC" value
        filtered_df.insert(filtered_df.columns.get_loc('Team_Number') + 1, 'PhotoLoc',
                           filtered_df['Block_ID'].astype(str) + '_KnownAssets_' + pd.to_datetime(
                               filtered_df['FlightDate'], errors='coerce').dt.strftime('%Y%m%d').fillna(''))

        # Calculate distances and add them as a new column to the DataFrame
        print_to_widget("\nCalculating structure distance from GIS location...")
        distances = []
        for index, ws_row in filtered_df.iterrows():
            if not has_missing_values(ws_row):
                dist = distance_calculator((ws_row['Mapped_Lat'], ws_row['Mapped_Lon']), (ws_row['FieldLat'],
                                                                                          ws_row['FieldLong']))
                if dist is not None and not math.isnan(dist):
                    distances.append(round(dist, 2))
                else:
                    distances.append('N/A')
            else:
                distances.append('N/A')
        filtered_df.insert(filtered_df.columns.get_loc('Mapped_Lon') + 1, 'DistFromGIS (ft)', distances)

        # Calculate the farthest image distance from the nadir as a new column to the DataFrame
        print_to_widget("\nCalculating the farthest distance of images from the nadir...")
        nadir_distances = get_farthest_from_nadir(dir_path)
        filtered_df.insert(filtered_df.columns.get_loc('DistFromGIS (ft)') + 1, 'FarthestDistFromNadir (ft)',
                           filtered_df['FolderName'].apply(
                               lambda x: round(nadir_distances.get(x), 2) if nadir_distances.get(
                                   x) is not None else 'N/A'))

        # Extract the directory path from dir_path
        directory = os.path.dirname(file_path)

        # Extract the base name of the file without extension
        file_name = os.path.splitext(os.path.basename(file_path))[0]

        # Search for a date in the file name using regex
        match = re.search(r"_(\d{4}\d{2}\d{2})_|(\d{4}\d{2}\d{2})$", file_name)

        # If a date is found, parse it to a datetime object
        if match:
            file_date = match.group(0)
        else:
            # If no date is found, use the current date as a fallback
            file_date = datetime.now().strftime("%Y%m%d")

        # Now you can use file_date to name your new file
        new_file_name = f"T_Filtered_GIS_Extract_{file_date}.xlsx"

        # Construct the new file path for saving
        new_file_path = os.path.join(directory, new_file_name)

        # Initialize the counter
        counter = 1

        # Check if the file already exists
        while os.path.exists(new_file_path):
            # Append the counter to the base file name
            modified_file_name = f"{new_file_name.split('.')[0]} ({counter}).{new_file_name.split('.')[-1]}"
            new_file_path = os.path.join(directory, modified_file_name)
            counter += 1

        # Write the filtered dataframe to a new Excel file
        filtered_df.to_excel(new_file_path, index=False, sheet_name='FilteredExtract')

        # Open the Excel file with openpyxl and format the cells in the "Match" and "FLOC" columns that are False
        wb = load_workbook(new_file_path)
        ws = wb.active

        # Delete unnecessary columns from the worksheet
        print_to_widget("\nDeleting unnecessary columns...")
        # List of header names to delete
        headers_to_delete = [
            "Unique_ID", "FID", "AOC_Area_Name", "FLOC_Plann", "Date_Released", "Added_Removed", "Inspection",
            "Vendor_ID", "Scope_Area", "TransGrid_", "XY_Source", "SAP_City", "Bundle", "Agency_Nam", "Agency_Are",
            "ESA", "ObjectId", "GlobalID", "x", "y"
        ]
        # Find indices and delete columns
        for header in headers_to_delete:
            col_index = find_col_index_by_header(ws, header)
            if col_index:
                ws.delete_cols(col_index)

        # Define a PatternFill object
        light_red_fill = PatternFill(start_color='FFFFC7CE', end_color='FFFFC7CE', fill_type='solid')
        dark_red_font = Font(color='9C0006')

        # Find the index of the 'Distance', 'FarthestDistFromNadir (ft)', 'FieldLat', and 'FieldLong' columns
        distance_column_index = find_col_index_by_header(ws, "DistFromGIS (ft)")
        field_lat_column_index = find_col_index_by_header(ws, "FieldLat")
        field_long_column_index = find_col_index_by_header(ws, "FieldLong")
        nadir_max_distance_col = find_col_index_by_header(ws, "FarthestDistFromNadir (ft)")
        photo_count_col_index = find_col_index_by_header(ws, "PhotoCount")
        flight_date_col_index = find_col_index_by_header(ws, "FlightDate")

        # Apply cell formatting to the appropriate cells based on the specified conditions
        print_to_widget("\nHighlighting potential issues...")
        for ws_row in range(2, ws.max_row + 1):
            distance_cell = ws.cell(row=ws_row, column=distance_column_index)
            field_lat_cell = ws.cell(row=ws_row, column=field_lat_column_index)
            field_long_cell = ws.cell(row=ws_row, column=field_long_column_index)
            farthest_cell = ws.cell(row=ws_row, column=nadir_max_distance_col)
            photo_count_cell = ws.cell(row=ws_row, column=photo_count_col_index)
            distance_value = distance_cell.value
            if distance_value == 'N/A':
                format_cell(field_lat_cell, field_lat_cell.value, light_red_fill, dark_red_font)
                format_cell(field_long_cell, field_long_cell.value, light_red_fill, dark_red_font)
                format_cell(distance_cell, distance_value, light_red_fill, dark_red_font)
            elif distance_value > 200:
                format_cell(distance_cell, distance_value, light_red_fill, dark_red_font)
            else:
                farthest_cell.alignment = Alignment(horizontal='right')
                distance_cell.alignment = Alignment(horizontal='right')
            if farthest_cell.value == 'N/A' or int(farthest_cell.value) > 300:
                format_cell(farthest_cell, farthest_cell.value, light_red_fill, dark_red_font)
            if photo_count_cell.value == 0:
                format_cell(photo_count_cell, photo_count_cell.value, light_red_fill, dark_red_font)

        # Highlight cells that are not a matching
        for ws_row in ws.iter_rows(min_row=2, min_col=1, max_col=33):
            date1 = ws_row[24].value
            date2 = ws_row[27].value
            if not isinstance(date1, datetime):
                date1 = datetime.strptime(date1, '%Y-%m-%d')
            if not isinstance(date2, datetime):
                date2 = datetime.strptime(date2, '%Y-%m-%d %H:%M:%S')
            # Compare folder name and GIS structure ID
            if ws_row[0].value != ws_row[1].value:
                ws_row[0].fill = light_red_fill
                ws_row[0].font = dark_red_font
                ws_row[1].fill = light_red_fill
                ws_row[1].font = dark_red_font
            # Compare flight date and inspection date
            if date1.date() != date2.date():
                ws_row[27].fill = light_red_fill
                ws_row[27].font = dark_red_font

        # Save the modified Excel file
        wb.save(new_file_path)

        # Create a new DataFrame to compare GIS Structure IDs and Folder Names
        print_to_widget("\nAdding new tabs for each team...")

        # Filter the GIS data based on 'Inspection_Date1' or 'Inspection_Date2' and 'Team_Number'
        filtered_df_drop = filtered_df.copy()
        filtered_df_drop['Team_Number'] = filtered_df_drop['Team_Number'].astype(str)
        filtered_df_drop['Team_Float'] = filtered_df_drop['Team_Number'].str.extract(r'(\d+)').astype(float)

        # Initialize a list of created sheet names
        created_sheet_names = []

        with pd.ExcelWriter(new_file_path, engine='openpyxl', mode='a') as writer:
            # Iterate over each unique team number
            for team in filtered_df_drop['Team_Float'].unique():
                # Filter data for the current team
                filtered_df_team = filtered_df_drop[filtered_df_drop['Team_Float'] == team]
                mask_team = (df_copy['Team_Number'] == team)

                df_copy['Inspection_Date'] = df_copy['Inspection_Date2'].combine_first(df_copy['Inspection_Date1'])
                # Extract only the date part from the datetime objects
                df_copy_date_only = df_copy['Inspection_Date'].dt.date
                mask_date = df_copy_date_only.isin(filtered_df_team['FlightDate'])
                gis_df = df_copy[mask_team & mask_date].copy()

                filtered_df_sorted = filtered_df_team.iloc[
                    filtered_df_team['FolderName'].apply(natural_sort_key).argsort()]
                gis_df_sorted = gis_df.iloc[gis_df['FLOC'].apply(natural_sort_key).argsort()]

                structure_ids_df = pd.DataFrame({
                    'Folder_ID': filtered_df_sorted['FolderName'],
                    'GIS_ID': gis_df_sorted['FLOC'],
                    'Team_Number': gis_df_sorted['Team_Number'],
                    'Inspection_Date': gis_df_sorted['Inspection_Date2'].combine_first(gis_df_sorted[
                                                                                           'Inspection_Date1']),
                    'Mapped_Lat': gis_df_sorted['Mapped_Lat'],
                    'Mapped_Long': gis_df_sorted['Mapped_Lon'],
                    'Vendor_Notes': gis_df_sorted['Vendor_Notes'],
                    'GIS_Notes': gis_df_sorted['GIS_Notes']
                })
                structure_ids_df.sort_values(by=["GIS_ID"], inplace=True)
                # Write the filtered dataframe for the current team to Excel
                try:
                    sheet_name = f"Team {int(team)}"
                except (ValueError, TypeError):  # Handle both empty strings and None values, among others
                    sheet_name = "Team NaN"
                structure_ids_df.to_excel(writer, index=False, sheet_name=sheet_name)
                created_sheet_names.append(sheet_name)

        # Load workbook
        wb = load_workbook(new_file_path)

        for sheet_name in created_sheet_names:  # Loop through only the recently created sheets
            ws = wb[sheet_name]

            # The rest remains the same...
            folder_ids = set(cell.value for cell in ws['A'] if cell.value is not None)
            gis_ids = set(cell.value for cell in ws['B'] if cell.value is not None)
            unmatched_folder_ids = folder_ids - gis_ids
            unmatched_gis_ids = gis_ids - folder_ids

            for cell in ws['A']:
                if cell.row != 1 and cell.value in unmatched_folder_ids:
                    cell.fill = light_red_fill
                    cell.font = dark_red_font

            for cell in ws['B']:
                if cell.row != 1 and cell.value in unmatched_gis_ids:
                    cell.fill = light_red_fill
                    cell.font = dark_red_font

        # Save the workbook
        wb.save(new_file_path)

    return new_file_path


def append_issues_sheet(file_path):
    # read FilteredExtract into a pandas DataFrame
    df = pd.read_excel(file_path, sheet_name='FilteredExtract')

    ## filter rows based on your criteria
    print_to_widget(f"\nFinding issues...")

    # filter missing N and GPS metadata issues
    field_lat_issue_df = df[(df['FieldLat'] == 'Nadir missing N') | (df['FieldLat'] == 'Metadata Issue')].copy()

    # filter incorrect folder names issue
    structure_floc = df['Structure_'] if 'Structure_' in df.columns else df[
        'FLOC'] if 'FLOC' in df.columns else None
    if structure_floc is None:
        print_to_widget('"FLOC" column not found. Skipping this file.')
        return
    folder_name_issue_df = df[
        df['FolderName'].astype(str).str.strip() != structure_floc.astype(str).str.strip()].copy()

    # GIS notes issue
    inspection_status_issue_df = df[
        df[['Inspection_Date1', 'Inspection_Date2']].isnull().all(axis=1) | df['Vendor_Status'].isnull()].copy()

    # Inspection/Flight date issue
    df['FlightDate'] = pd.to_datetime(df['FlightDate']).dt.date
    df['Inspection_Date1'] = pd.to_datetime(df['Inspection_Date1']).dt.date
    df['Inspection_Date2'] = pd.to_datetime(df['Inspection_Date2']).dt.date
    df['Relevant_Inspection_Date'] = df['Inspection_Date2']
    df.loc[df['Relevant_Inspection_Date'].isnull(), 'Relevant_Inspection_Date'] = df['Inspection_Date1']
    date_issue_df = df[df['FlightDate'] != df['Relevant_Inspection_Date']].copy()

    # empty folder issue
    empty_folder_issue_df = df[df['PhotoCount'] == 0].copy()

    # check if all dataframes are empty
    if (field_lat_issue_df.empty and folder_name_issue_df.empty and inspection_status_issue_df.empty
            and date_issue_df.empty and empty_folder_issue_df.empty):
        print_to_widget('No issues found.')
        return file_path
    else:
        print_to_widget("Issue(s) found. Creating \"Issues\" sheet...")

    # add 'Issues' column
    field_lat_issue_df['Issue'] = field_lat_issue_df['FieldLat']
    folder_name_issue_df['Issue'] = 'Incorrect Folder Name'
    inspection_status_issue_df['Issue'] = 'Missing/Incorrect GIS Notes/Fields'
    date_issue_df['Issue'] = 'Date Mismatch'
    empty_folder_issue_df['Issue'] = 'Empty Folder'

    # Concatenate the 5 dataframes
    issues_df = pd.concat([field_lat_issue_df, folder_name_issue_df, inspection_status_issue_df, date_issue_df,
                           empty_folder_issue_df])

    # Add the "QA Reviewer" column
    issues_df['QA Reviewer'] = ""

    # Add the "QA Date" column with the current date
    current_date = datetime.today().date().strftime('%Y-%m-%d')
    issues_df['QA Date'] = current_date

    # check which columns are present in the dataframe and select accordingly
    structure_col = 'Structure_' if 'Structure_' in issues_df.columns else 'FLOC'
    lat_col = 'Mapped_Lat' if 'Mapped_Lat' in issues_df.columns else 'y2'
    lon_col = 'Mapped_Lon' if 'Mapped_Lon' in issues_df.columns else 'x2'

    if structure_col == 'FLOC' and 'FLOC' not in issues_df.columns:
        print_to_widget('"FLOC" column not found.')
        return

    if lat_col == 'y2' and 'y2' not in issues_df.columns:
        print_to_widget('Neither "Mapped_Lat" nor "y2" column found.')
        return

    if lon_col == 'x2' and 'x2' not in issues_df.columns:
        print_to_widget('Neither "Mapped_Lon" nor "x2" column found.')
        return

    # Create a function to generate the note
    def generate_note(issue_row):
        if issue_row['Issue'] == 'Incorrect Folder Name':
            return f"Folder submitted as: {issue_row['FolderName']} (Correct ID: {issue_row[structure_col]})."
        elif issue_row['Issue'] == 'Missing/Incorrect GIS Notes/Fields':
            missing_fields = []
            if pd.isnull(issue_row['Inspection_Date1']) and pd.isnull(issue_row['Inspection_Date2']):
                missing_fields.append('Inspection Date')
            if pd.isnull(issue_row['Vendor_Status']):
                missing_fields.append('Vendor Status')
            return f"Missing {', '.join(missing_fields)}. Need to update GIS."
        elif issue_row['Issue'] == 'Nadir missing N':
            return ""
        elif issue_row['Issue'] == 'Metadata Issue':
            return "QA Note: GPS data missing."
        elif issue_row['Issue'] == 'Date Mismatch':
            return "Flight Date and GIS Inspection Date are not the same."
        else:
            return ""

    # Add the 'Notes/Feedback' column
    issues_df['Notes/Feedback'] = issues_df.apply(generate_note, axis=1)

    # Create a function to determine the division based on the flight team
    def determine_division(df_row):
        try:
            if df_row['Team_Number'].startswith('Team 1'):
                return "Transmission"
            elif df_row['Team_Number'].startswith('Team 5'):
                return "V360"
            else:
                return ""
        except Exception as e:
            print_to_widget(f"Error while processing {df_row['Team_Number']}, row {df_row.name + 1}")
            print_to_widget(str(e))
            return ""

    # Add the "Division" column
    issues_df['Division'] = issues_df.apply(determine_division, axis=1)

    # keep only the columns you care about, in the order you specified
    issues_df = issues_df[['FolderName', structure_col, lat_col, lon_col, 'Team_Number', 'Division', 'FlightDate',
                           'QA Reviewer', 'QA Date', 'Issue', 'Notes/Feedback']]

    # rename columns
    issues_df = issues_df.rename(columns={
        'FolderName': 'Folder Name',
        structure_col: 'Structure ID',
        lat_col: 'Latitude',
        lon_col: 'Longitude',
        'Team_Number': 'Flight Team',
        'FlightDate': 'Original Flight Date'
    })

    # convert 'FlightDate' to datetime and extract only the date part
    issues_df['Original Flight Date'] = pd.to_datetime(issues_df['Original Flight Date']).dt.date

    # Initiate the writer object
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
        # write the new sheet to the workbook
        issues_df.to_excel(writer, index=False, sheet_name='Issues')

    return file_path


ez_in_trans_df = None
ez_in_dist_df = None


def append_ez_aoc_sheet(file_path):
    # read FilteredExtract into a pandas DataFrame
    global ez_in_trans_df, ez_in_dist_df
    df = pd.read_excel(file_path, sheet_name='FilteredExtract')

    # filter ez poles
    ez_poles_df = df[df['Eq_ObjType'] == 'EZ_POLE'].reset_index(drop=True).copy()
    aoc_df = df[df['AOC'].notna()].reset_index(drop=True).copy()
    if 'EZ_in_Trans' in df.columns:
        ez_in_trans_df = df[df['EZ_in_Trans'] == 'EZ Pole on Trans Map'].reset_index(drop=True).copy()
        ez_in_dist_df = None
    elif 'EZ_in_Dist' in df.columns:
        ez_in_dist_df = df[df['EZ_in_Dist'] == 'In Distro Package'].reset_index(drop=True).copy()
        ez_in_trans_df = None

    # Define styles for highlighting
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    green_font = Font(color='006100')
    light_red_fill = PatternFill(start_color='FFFFC7CE', end_color='FFFFC7CE', fill_type='solid')
    dark_red_font = Font(color='9C0006')

    # Check which column exist in DataFrame
    ez_in_trans_exists = 'EZ_in_Trans' in ez_poles_df.columns
    ez_in_dist_exists = 'EZ_in_Dist' in ez_poles_df.columns

    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
        workbook = writer.book

        # Write the new sheets to the workbook
        if not ez_poles_df.empty:
            print_to_widget(f"\nEZ Poles found. Adding 'EZPoles' sheet...")
            ez_poles_df.to_excel(writer, index=False, sheet_name='EZPoles')
            worksheet_ez = workbook['EZPoles']
            # Apply styles
            for ez_idx, ez_row in ez_poles_df.iterrows():
                row_number = ez_idx + 2
                if ez_row['FLOC'] != ez_row['FolderName']:
                    for ez_col in range(1, 3):
                        cell = worksheet_ez.cell(row=row_number, column=ez_col)
                        cell.fill = light_red_fill
                        cell.font = dark_red_font
                if ez_in_trans_exists and ez_row['EZ_in_Trans'] == 'EZ Pole on Trans Map':
                    for ez_col in range(1, worksheet_ez.max_column + 1):
                        cell = worksheet_ez.cell(row=row_number, column=ez_col)
                        cell.fill = green_fill
                        cell.font = green_font
                elif ez_in_dist_exists and ez_row['EZ_in_Dist'] == 'In Distro Package':
                    for ez_col in range(1, worksheet_ez.max_column + 1):
                        cell = worksheet_ez.cell(row=row_number, column=ez_col)
                        cell.fill = green_fill
                        cell.font = green_font

            if ez_in_trans_df is not None and not ez_in_trans_df.empty:
                print_to_widget(f"\nEZ Pole in Trans found. Adding 'EZinTrans' sheet...")
                ez_in_trans_df.to_excel(writer, index=False, sheet_name='EZinTrans')
                worksheet_trans = workbook['EZinTrans']
                # Apply styles
                for trans_idx, trans_row in ez_in_trans_df.iterrows():
                    row_number = trans_idx + 2
                    if trans_row['FLOC'] != trans_row['FolderName']:
                        for trans_col in range(1, 3):
                            cell = worksheet_trans.cell(row=row_number, column=trans_col)
                            cell.fill = light_red_fill
                            cell.font = dark_red_font
            else:
                print_to_widget("\nNo EZ Poles in Trans found.")

            if ez_in_dist_df is not None and not ez_in_dist_df.empty:
                print_to_widget(f"\nEZ Pole in Distro found. Adding 'EZinDistro' sheet...")
                ez_in_dist_df.to_excel(writer, index=False, sheet_name='EZinDistro')
                worksheet_distro = workbook['EZinDistro']
                # Apply styles
                for distro_idx, distro_row in ez_in_dist_df.iterrows():
                    row_number = distro_idx + 2
                    if distro_row['FLOC'] != distro_row['FolderName']:
                        for distro_col in range(1, 3):
                            cell = worksheet_distro.cell(row=row_number, column=distro_col)
                            cell.fill = light_red_fill
                            cell.font = dark_red_font
            else:
                print_to_widget("\nNo EZ Poles in Distro found.")
        else:
            print_to_widget("\nNo EZ Poles found.")

        if not aoc_df.empty:
            print_to_widget(f"\nAOC structures found. Adding 'AOC' sheet...")
            aoc_df.to_excel(writer, index=False, sheet_name='AOC')
            worksheet_aoc = workbook['AOC']
            # Apply styles
            for aoc_idx, aoc_row in aoc_df.iterrows():
                row_number = aoc_idx + 2
                if aoc_row['FLOC'] != aoc_row['FolderName']:
                    for aoc_col in range(1, 3):
                        cell = worksheet_aoc.cell(row=row_number, column=aoc_col)
                        cell.fill = light_red_fill
                        cell.font = dark_red_font
        else:
            print_to_widget("\nNo AOC structures found.")

        # Save the workbook
        workbook.save(file_path)


def move_folders_based_on_issues(excel_file_path, directory_path):
    print_to_widget("\nSeparating issue and non-Issue structures...")
    issue_sheet_exists = True
    issue_ids = set()

    # Check if there is no Issues sheet
    try:
        issues_df = pd.read_excel(excel_file_path, sheet_name='Issues')
        issue_ids = set(issues_df['Folder Name'].astype(str))
    except (FileNotFoundError, ValueError, KeyError):
        issue_sheet_exists = False

    ez_poles_df = pd.read_excel(excel_file_path, sheet_name='EZPoles')
    # Create dictionaries for EZ Poles categories
    if 'EZ_in_Trans' in ez_poles_df.columns:
        ez_poles_map = dict(zip(ez_poles_df['FolderName'], ez_poles_df['EZ_in_Trans']))
        distro_ez_poles = True
    elif 'EZ_in_Dist' in ez_poles_df.columns:
        ez_poles_map = dict(zip(ez_poles_df['FolderName'], ez_poles_df['EZ_in_Dist']))
        distro_ez_poles = False
    else:
        raise ValueError("Neither 'EZ_in_Trans' nor 'EZ_in_Dist' column found in EZPoles sheet.")

    # Define all target folders with their names
    target_folders = {
        os.path.join(directory_path, 'Issue'): 'Issue',
        os.path.join(directory_path, 'Non-Issue', 'Regular Structures'): 'Regular Structures',
        os.path.join(directory_path, 'Non-Issue', 'EZ Poles in Both'): 'EZ Poles in Both',
        os.path.join(directory_path, 'Non-Issue', 'EZ Poles in Distro Only'): 'EZ Poles in Distro Only',
        os.path.join(directory_path, 'Non-Issue', 'EZ Poles in Trans Only'): 'EZ Poles in Trans Only'
    }
    if issue_sheet_exists:
        target_folders[os.path.join(directory_path, 'Issue')] = 'Issue'

    # Create the folders
    for folder_path in target_folders.keys():
        os.makedirs(folder_path, exist_ok=True)

    # Function to move folders
    def move_folder(original_path, target_path):
        if os.path.isdir(original_path):
            # Print the folder name being moved and its destination folder name
            print_to_widget(f"Moving '{os.path.basename(original_path)}' to '{os.path.basename(target_path)}'...")
            # Ensure the target subfolder exists
            os.makedirs(os.path.dirname(target_path), exist_ok=True)

            # Move the folder
            shutil.move(original_path, target_path)

    # Function to delete empty folders
    def delete_empty_folders(root_directory):
        deletion_occurred = True
        while deletion_occurred:
            deletion_occurred = False
            for dir_path, dir_names, filenames in os.walk(root_directory, topdown=False):
                if not dir_names and not filenames:
                    try:
                        print_to_widget(f"Folder {os.path.basename(dir_path)} is empty. Deleting folder...")
                        os.rmdir(dir_path)
                        deletion_occurred = True
                    except OSError as e:
                        print_to_widget(f"Error deleting {os.path.basename(dir_path)}: {e}")

    move_operations = []

    # Walk through the directory and move folders
    for root_dir, dirs, files in os.walk(directory_path, topdown=False):  # Use topdown=False to start from innermost
        for dir_name in dirs:
            current_path = os.path.join(root_dir, dir_name)
            parent_folder_path = os.path.dirname(current_path)

            # Skip if the folder is already in its correct target location
            if issue_sheet_exists and dir_name in issue_ids:
                if parent_folder_path == target_folders[os.path.join(directory_path, 'Issue')]:
                    continue
            elif dir_name not in issue_ids:
                if parent_folder_path in target_folders.values():
                    continue

            # Check if the folder is already in the correct target location
            if dir_name in issue_ids and parent_folder_path == target_folders[os.path.join(directory_path, 'Issue')]:
                continue  # Skip if the folder is already in 'Issue'
            elif dir_name not in issue_ids and parent_folder_path in target_folders.values():
                continue  # Skip if the folder is already in 'Non-Issue' or its subfolders
            # Check if the folder is innermost and not empty
            if not os.listdir(current_path) or any(
                    os.path.isdir(os.path.join(current_path, sub_dir)) for sub_dir in os.listdir(current_path)):
                continue
            # Determine the subfolder name
            if dir_name in issue_ids:
                subfolder_name = 'Issue'
            elif dir_name in ez_poles_map:
                if distro_ez_poles and ez_poles_map[dir_name] == 'EZ Pole on Trans Map':
                    subfolder_name = 'EZ Poles in Both'
                elif not distro_ez_poles and ez_poles_map[dir_name] == 'In Distro Package':
                    subfolder_name = 'EZ Poles in Both'
                elif not distro_ez_poles and ez_poles_map[dir_name] == 'Trans Only Structure':
                    subfolder_name = 'EZ Poles in Trans Only'
                else:
                    subfolder_name = 'EZ Poles in Distro Only'
            else:
                subfolder_name = 'Regular Structures'

            # Get the full target path from the dictionary
            target_subfolder_path = next((path for path, name in target_folders.items() if name == subfolder_name),
                                         None)

            # If the current folder is not in its intended target, add it to move operations
            if target_subfolder_path and os.path.dirname(current_path) != target_subfolder_path:
                move_operations.append((current_path, target_subfolder_path))
            else:
                print_to_widget(f"{os.path.basename(current_path)} is already in the correct folder.")

    # Perform the move operations
    for og_path, dest_path in move_operations:
        move_folder(og_path, dest_path)

    # Delete all empty folders
    delete_empty_folders(directory_path)
    print_to_widget('\nExtract filtered successfully!.')


def natural_sort_key(s):
    """
    Define a sort key function that can be used to sort filenames naturally.
    """
    return [int(x) if x.isdigit() else x for x in re.split(r'(\d+)', s)]


# Create a function to run when the "Rename Images" button is clicked
def rename_images_auto(dir_path):
    # Delete hidden Mac files
    print_to_widget("\nDeleting hidden Mac files...")
    found_mac_files = False
    for dirpath, dirnames, filenames in os.walk(dir_path):
        print_to_widget(f"Checking folder: {os.path.basename(dirpath)}")
        for filename2 in filenames:
            if filename2.startswith("._"):
                file_path = os.path.join(dirpath, filename2)
                os.remove(file_path)
                found_mac_files = True
                print_to_widget(f"{filename2} deleted.")

    if not found_mac_files:
        print_to_widget("No hidden Mac files found in the folder.")
    else:
        print_to_widget(f"All hidden Mac files have been deleted.")

    # Loop over all the directories and files under the root folder
    print_to_widget(f"\nRenaming images...")
    original_names = {}  # Initialize a dictionary to store the original names
    for dirpath, dirnames, filenames in os.walk(dir_path):
        # Initialize the counter
        counter = 10000
        for filename1 in sorted(filenames, key=natural_sort_key):
            # Check if the file is an image file
            if filename1.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp', '.gif')):
                file_path = os.path.join(dirpath, filename1)
                # Store the original name in the dictionary
                original_names[file_path] = filename1
                # Rename the file with a new name
                new_name = str(counter) + '.JPG'
                new_path = os.path.join(dirpath, new_name)
                os.rename(file_path, new_path)
                # Update the dictionary with the new file path
                original_names[new_path] = original_names.pop(file_path)
                # Increment the counter
                counter += 22

    # Loop over all the directories and files under the root folder
    for dirpath, dirnames, filenames in os.walk(dir_path):
        print_to_widget(f"\nProcessing folder {os.path.basename(dirpath)}")
        # Initialize the counter
        counter = 1
        for filename in sorted(filenames, key=natural_sort_key):
            # Check if the file is an image file
            if filename.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp', '.gif')):
                file_path = os.path.join(dirpath, filename)

                # Get the date taken from the image EXIF metadata
                date_taken = get_date_taken(file_path)

                # Rename the file with the specified format
                folder_name = os.path.basename(os.path.normpath(dirpath))
                new_filename = f"OH-{folder_name}_{date_taken}_{counter}.JPG"
                counter += 1
                new_file_path = os.path.join(dirpath, new_filename)
                os.rename(file_path, new_file_path)

                # Print the original name and the new name
                original_name = original_names[file_path]
                print_to_widget(f"{original_name} -> {new_filename}")

    print_to_widget(f"All images have been renamed.")


def rename_images_manual(dir_path):
    def is_valid_date(date_text, date_format='%Y%m%d'):
        try:
            datetime.strptime(date_text, date_format)
            return True
        except ValueError:
            return False

    # Delete hidden Mac files
    found_mac_files = False
    for dirpath, dirnames, filenames in os.walk(dir_path):
        for filename2 in filenames:
            if filename2.startswith("._"):
                file_path = os.path.join(dirpath, filename2)
                os.remove(file_path)
                found_mac_files = True
                print_to_widget(f"{filename2} deleted.")

    if not found_mac_files:
        print_to_widget("No hidden Mac files found in the folder.")
    else:
        print_to_widget(f"All hidden Mac files have been deleted.")

    # Loop over all the directories and files under the root folder
    print_to_widget(f"Renaming images...")
    for dirpath, dirnames, filenames in os.walk(dir_path):
        # Initialize the counter
        counter = 10000
        for filename1 in sorted(filenames, key=natural_sort_key):
            # Check if the file is an image file
            if filename1.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp', '.gif')):
                file_path = os.path.join(dirpath, filename1)
                # Rename the file with a new name
                new_name = str(counter) + '.JPG'
                new_path = os.path.join(dirpath, new_name)
                os.rename(file_path, new_path)
                # Increment the counter
                counter += 22

    # Open a Tkinter window to prompt the user for the date
    while True:
        # Open a Tkinter window to prompt the user for the date
        date_root = ctk.CTk()
        date_root.withdraw()
        date_str = tk.simpledialog.askstring("Enter Date", "Enter the date the images were taken (YYYYMMDD): ",
                                             parent=date_root)
        date_root.destroy()

        if date_str is None:
            print_to_widget("No date entered, please enter a date...")
        elif not is_valid_date(date_str):
            print_to_widget("Invalid date format, please enter a date in the format YYYYMMDD")
        else:
            break

    # Loop over all the directories and files under the root folder
    for dirpath, dirnames, filenames in os.walk(dir_path):
        # Initialize the counter
        counter = 1
        for filename in filenames:
            # Check if the file is an image file
            if filename.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp', '.gif')):
                file_path = os.path.join(dirpath, filename)

                # Rename the file with the specified format
                folder_name = os.path.basename(os.path.normpath(dirpath))
                new_filename = f"{folder_name}_{date_str}_{counter}.JPG"
                counter += 1
                new_file_path = os.path.join(dirpath, new_filename)
                os.rename(file_path, new_file_path)

    print_to_widget(f"All images have been renamed.")


# # function when watermark button is called
# def watermark_directory(src_dir, dst_dir):
#     print_to_widget(f"Watermarking images in {os.path.basename(src_dir)}...")
#     image_files = natsort.natsorted(glob.glob(os.path.join(src_dir, '**', '*.JPG'), recursive=True))
#     num_images = len(image_files)
#     start_time = time.time()
#     for image_i, image_path in enumerate(image_files):
#         try:
#             # read image
#             image = Image.open(image_path)
#             # read GPS image direction metadata and convert to directional letter
#             direction = get_gps_direction_letter(image_path)
#             # apply watermark
#             if direction is not None:
#                 # create destination directory if it doesn't exist
#                 rel_dirpath = os.path.relpath(os.path.dirname(image_path), src_dir)
#                 dst_subdir = os.path.join(dst_dir, rel_dirpath)
#                 os.makedirs(dst_subdir, exist_ok=True)
#                 # save watermarked image to destination directory
#                 dst_path = os.path.join(dst_subdir, os.path.basename(image_path))
#                 exif_data = piexif.load(image_path)
#                 exif_bytes = piexif.dump(exif_data)
#                 image_with_watermark = add_watermark(image, direction, exif_bytes)
#                 image_with_watermark.save(dst_path, exif=exif_bytes)
#                 print_to_widget(f"Watermarked {os.path.basename(image_path)} with {direction}. "
#                                 f"({image_i + 1}/{num_images})")
#             else:
#                 print_to_widget(
#                     f"No image direction data found for {os.path.basename(image_path)}. Skipping image. "
#                     f"({image_i + 1}/{num_images})")
#         except OSError as e:
#             print_to_widget(
#                 f"Error processing {os.path.basename(image_path)}: {e} Skipping image. ({image_i + 1}/{num_images})")
#     end_time = time.time()
#     elapsed_time = end_time - start_time
#     elapsed_min = int(elapsed_time // 60)
#     elapsed_sec = int(elapsed_time % 60)
#     print_to_widget(f"Done watermarking images in {os.path.basename(src_dir)}.")
#     print_to_widget(f"Elapsed time: {elapsed_min} min {elapsed_sec} sec")
#
#     # call gc.collect() to force garbage collection
#     gc.collect()


def complete_traveler_distro(source_file, dest_file, columns_mapping, qa_name="1", source_sheet=0, dest_sheet=0):
    print_to_widget("\nCopying data from source sheet...")
    # Read source Excel file as DataFrame
    source_df = pd.read_excel(source_file, sheet_name=source_sheet).reset_index(drop=True)

    # Compute the C2_ID column based on the last 4 characters of the block_id column
    block_id_col = next((source_col for source_col in source_df.columns if source_col.lower() == 'block_id'), None)

    # Check if the column exists
    if block_id_col:
        # Check for NaN or blank values in 'Block ID'
        mask = source_df[block_id_col].isna() | (source_df[block_id_col] == '')
        empty_block_ids = source_df[mask]

        if not empty_block_ids.empty:
            structure_col = next((source_col for source_col in source_df.columns
                                  if source_col.lower().strip().startswith('floc')), None)

            if structure_col:
                for empty_bid_index, empty_bid_source_row in empty_block_ids.iterrows():
                    print_to_widget(f"Warning: {empty_bid_source_row[structure_col]} has no Block ID.")
                # Fill NaN values with empty strings to continue
                source_df[block_id_col] = source_df[block_id_col].fillna('')

        # Compute 'C2_ID' column
        source_df['C2_ID'] = source_df[block_id_col].str[-4:].str.lstrip('0')

        # If 'C2_ID' is empty after stripping leading zeros, set it to '0'
        source_df['C2_ID'] = source_df['C2_ID'].replace('', 0).fillna(0).astype(int)

    else:
        print_to_widget("Column 'Block_ID' not found!")

    # Load destination Excel file with openpyxl
    dest_wb = load_workbook(dest_file)
    dest_ws = dest_wb[dest_wb.sheetnames[dest_sheet] if isinstance(dest_sheet, int) else dest_sheet]

    # Ensure only headers remain in the destination worksheet
    if dest_ws.max_row > 1:
        for source_df_row in dest_ws.iter_rows(min_row=2, max_row=dest_ws.max_row):
            for cell in source_df_row:
                cell.value = None

    # Find destination column indices
    print_to_widget("\nTransferring data to traveler sheet template...")
    dest_col_indices = {cell.value: dest_i for dest_i, cell in enumerate(dest_ws[1], start=1) if
                        cell.value in columns_mapping.values()}

    # Set alignment styles for the entire sheet
    center_alignment = Alignment(horizontal="center", vertical="center")
    dest_ws.sheet_view.showGridLines = False
    dest_ws.sheet_view.zoomScale = 100

    # Set left alignment for specific columns
    left_alignment = Alignment(horizontal="left", vertical="center")
    left_aligned_columns = ["Existing_Notes", "VENDOR_NOTES", "P1_Notes", "P2_Notes", "P3_Notes"]

    # Apply center alignment to all columns except those in left_aligned_columns
    for dest_ws_col in dest_ws.columns:
        if dest_ws_col[0].value == "C2_ID":
            for cell in dest_ws_col:
                cell.number_format = "#,##0"
        if dest_ws_col[0].value not in left_aligned_columns:
            for cell in dest_ws_col:
                cell.alignment = center_alignment

    # Copy columns from source DataFrame to destination Worksheet
    for source_index, source_row in source_df.iterrows():
        for source_col, dest_col in columns_mapping.items():
            if source_col in source_df.columns:
                dest_col_index = dest_col_indices[dest_col]
                if dest_col.lower() == "sce_struct":
                    if qa_name != "1":
                        cell_value = "OH-" + str(source_row[source_col])
                    else:
                        cell_value = str(source_row[source_col])
                    cell = dest_ws.cell(row=source_index + 2, column=dest_col_index, value=cell_value)
                elif dest_col.lower() == "flight_date":
                    cell = dest_ws.cell(row=source_index + 2, column=dest_col_index,
                                        value=pd.to_datetime(source_row[source_col]).date())
                    cell.number_format = "m/d/yyyy"
                else:
                    cell = dest_ws.cell(row=source_index + 2, column=dest_col_index, value=source_row[source_col])
                if dest_col in left_aligned_columns:
                    cell.alignment = left_alignment
                else:
                    cell.alignment = center_alignment
                if source_col == "Eq_ObjType" and source_row[source_col] != "EZ_POLE":
                    cell.value = None

    if qa_name != "1":
        # Identify the column index which will always have data (assuming it's the first column here)
        data_column_index = 4

        # Find the last row with data in the specified column
        last_row_with_data = max((dest_row for dest_row in range(1, dest_ws.max_row + 1) if
                                  dest_ws.cell(row=dest_row, column=data_column_index).value is not None), default=0)

        # Find the "QA Reviewer" column
        qa_reviewer_col = next((dest_i for dest_i, cell in enumerate(dest_ws[1], start=1)
                                if cell.value == "QA Reviewer"), None)

        if qa_reviewer_col:
            # Write the selected name in the "QA Reviewer" column for each row with data
            for source_df_row in range(2, last_row_with_data + 1):
                dest_ws.cell(row=source_df_row, column=qa_reviewer_col, value=qa_name)

        else:
            print_to_widget("Column 'QA Reviewer' not found!")
            return

        # Extract the date from the source file name
        date_match = re.search(r'\d{8}', os.path.basename(source_file))
        source_file_date = date_match.group(0) if date_match else datetime.now().strftime('%Y%m%d')

        # Construct the new filename
        qa_first_name = qa_name.split()[0]  # Assuming qa_name is in 'FirstName LastName' format
        new_filename = f"{qa_first_name}_D_C2_{source_file_date}.xlsx"

        # Get directory of the filtered extract file
        source_file_dir = os.path.dirname(source_file)
        new_file_path = os.path.join(source_file_dir, new_filename)

    else:
        # Construct the new filename
        new_filename = f"D_C2_{datetime.now().strftime('%Y%m%d')}.xlsx"

        # Get directory of the template file
        source_file_dir = os.path.dirname(dest_file)
        new_file_path = os.path.join(source_file_dir, new_filename)

    # Initialize a counter
    counter = 1

    # Split the new file path into directory, name, and extension
    file_dir, file_name = os.path.split(new_file_path)
    name, ext = os.path.splitext(file_name)

    # Check if the file exists and update the filename with a counter if needed
    while os.path.exists(new_file_path):
        new_file_name = f"{name} ({counter}){ext}"
        new_file_path = os.path.join(file_dir, new_file_name)
        counter += 1

    # Save the updated workbook to the new file path
    dest_wb.save(new_file_path)
    print_to_widget(f"\nFile saved as: {os.path.basename(new_file_path)}")


# def add_data_validation_distro(file_path):
#     # Load the workbook
#     wb = load_workbook(file_path)
#     ws = wb["C2-Distribution"]
#     dropdowns_ws = wb["DropDowns"]
#
#     # Identify the last row with data in SCE_Struct column
#     last_row = max((ws_row for ws_row in range(1, ws.max_row + 1) if ws.cell(row=ws_row, column=4).value is not None),
#                    default=1)
#
#     # Create named ranges for the dropdown references
#     print_to_widget("\nFinishing up...")
#     named_range_R = DefinedName(name="DropDowns_R", attr_text=f"{dropdowns_ws.title}!$B$2:$B$8")
#     named_range_S = DefinedName(name="DropDowns_S", attr_text=f"{dropdowns_ws.title}!$C$2:$C$23")
#     print(1)
#     wb.defined_names.append(named_range_R)
#     wb.defined_names.append(named_range_S)
#
#     # Create data validation objects
#     print(2)
#     with warnings.catch_warnings():
#         warnings.simplefilter("ignore")
#         dv_R = DataValidation(type="list", formula1="=DropDowns_R", allow_blank=True, showDropDown=None)
#         dv_S = DataValidation(type="list", formula1="=DropDowns_S", allowBlank=True, showDropDown=None)
#
#     print(3)
#     # Apply data validation to columns R and S up to the last row with entries
#     for ws_row in range(2, last_row + 1):
#         cell_r = ws["R{}".format(ws_row)]
#         cell_s = ws["S{}".format(ws_row)]
#
#         # Apply data validation to each cell
#         dv_R.add(cell_r)
#         dv_S.add(cell_s)
#
#     print(4)
#     # Add data validations to the worksheet
#     ws.add_data_validation(dv_R)
#     ws.add_data_validation(dv_S)
#
#     # Save the modified workbook
#     wb.save(file_path)
#     print_to_widget(f"\nFile saved as: {os.path.basename(file_path)}")


def complete_traveler_trans(source_file, dest_file, columns_mapping, qa_name="1", source_sheet=0, dest_sheet=0):
    print_to_widget("\nCopying data from source sheet...")
    # Read source Excel file as DataFrame
    source_df = pd.read_excel(source_file, sheet_name=source_sheet)

    # Check if the column exists
    block_id_col = next((source_col for source_col in source_df.columns if source_col.lower() == 'block_id'), None)

    if block_id_col:
        # Check for NaN or blank values in 'Block ID'
        mask = source_df[block_id_col].isna() | (source_df[block_id_col] == '')
        empty_block_ids = source_df[mask]

        if not empty_block_ids.empty:
            structure_col = next((source_col for source_col in source_df.columns
                                  if source_col.lower().strip().startswith('floc')), None)

            if structure_col:
                for empty_bid_index, empty_bid_source_row in empty_block_ids.iterrows():
                    print_to_widget(f"Warning: {empty_bid_source_row[structure_col]} has no Block ID.")
                # Fill NaN values with empty strings to continue
                source_df[block_id_col] = source_df[block_id_col].fillna('')
    else:
        print_to_widget("Column 'Block_ID' not found!")

    # Load destination Excel file with openpyxl
    dest_wb = load_workbook(dest_file)
    dest_ws = dest_wb[dest_wb.sheetnames[dest_sheet] if isinstance(dest_sheet, int) else dest_sheet]

    # Ensure only headers remain in the destination worksheet
    if dest_ws.max_row > 1:
        for empty_bid_source_row in dest_ws.iter_rows(min_row=2, max_row=dest_ws.max_row):
            for cell in empty_bid_source_row:
                cell.value = None

    # Find destination column indices
    print_to_widget("\nTransferring data to traveler sheet template...")
    dest_col_indices = {cell.value: dest_i for dest_i, cell in enumerate(dest_ws[1], start=1) if
                        cell.value in columns_mapping.values()}

    # Set alignment styles for the entire sheet
    center_alignment = Alignment(horizontal="center", vertical="center")
    dest_ws.sheet_view.showGridLines = False
    dest_ws.sheet_view.zoomScale = 85

    # Set left alignment for specific columns
    left_alignment = Alignment(horizontal="left", vertical="center")
    left_aligned_columns = ["Vendor_Note", "P1_Notes"]

    # Apply center alignment to all columns except those in left_aligned_columns
    for dest_ws_col in dest_ws.columns:
        if dest_ws_col[0].value not in left_aligned_columns:
            for cell in dest_ws_col:
                cell.alignment = center_alignment

    # Copy columns from source DataFrame to destination Worksheet
    for source_index, source_row in source_df.iterrows():
        for source_col, dest_col in columns_mapping.items():
            if source_col in source_df.columns:
                dest_col_index = dest_col_indices[dest_col]
                if dest_col.lower() == "sce_struct":
                    if qa_name != "1":
                        cell_value = "OH-" + str(source_row[source_col])
                    else:
                        cell_value = str(source_row[source_col])
                    cell = dest_ws.cell(row=source_index + 2, column=dest_col_index, value=cell_value)
                elif dest_col.lower() == "flight_date":
                    cell = dest_ws.cell(row=source_index + 2, column=dest_col_index,
                                        value=pd.to_datetime(source_row[source_col]).date())
                else:
                    cell = dest_ws.cell(row=source_index + 2, column=dest_col_index, value=source_row[source_col])
                if dest_col in left_aligned_columns:
                    cell.alignment = left_alignment
                else:
                    cell.alignment = center_alignment

    if qa_name != "1":
        # Identify the column index which will always have data (assuming it's the first column here)
        data_column_index = 1

        # Find the last row with data in the specified column
        last_row_with_data = max((dest_row for dest_row in range(1, dest_ws.max_row + 1) if
                                  dest_ws.cell(row=dest_row, column=data_column_index).value is not None), default=0)

        # Find the "QA Reviewer" column
        qa_reviewer_col = next((dest_i for dest_i, cell in enumerate(dest_ws[1], start=1)
                                if cell.value == "QA Reviewer"), None)

        if qa_reviewer_col:
            # Write the selected name in the "QA Reviewer" column for each row with data
            for source_df_row in range(2, last_row_with_data + 1):
                dest_ws.cell(row=source_df_row, column=qa_reviewer_col, value=qa_name)

        else:
            print_to_widget("Column 'QA Reviewer' not found!")
            return

        # Extract the date from the source file name
        date_match = re.search(r'\d{8}', os.path.basename(source_file))
        source_file_date = date_match.group(0) if date_match else datetime.now().strftime('%Y%m%d')

        # Construct the new filename
        qa_first_name = qa_name.split()[0]  # Assuming qa_name is in 'FirstName LastName' format
        new_filename = f"{qa_first_name}_T_C2_{source_file_date}.xlsx"

        # Get directory of the filtered extract file
        source_file_dir = os.path.dirname(source_file)
        new_file_path = os.path.join(source_file_dir, new_filename)

    else:
        # Construct the new filename
        new_filename = f"T_C2_{datetime.now().strftime('%Y%m%d')}.xlsx"

        # Get directory of the template file
        source_file_dir = os.path.dirname(dest_file)
        new_file_path = os.path.join(source_file_dir, new_filename)

    # Initialize a counter
    counter = 1

    # Split the new file path into directory, name, and extension
    file_dir, file_name = os.path.split(new_file_path)
    name, ext = os.path.splitext(file_name)

    # Check if the file exists and update the filename with a counter if needed
    while os.path.exists(new_file_path):
        new_file_name = f"{name} ({counter}){ext}"
        new_file_path = os.path.join(file_dir, new_file_name)
        counter += 1

    # Save the updated workbook to the new file path
    dest_wb.save(new_file_path)
    return new_file_path


# def add_data_validation_trans(file_path):
#     # Load the workbook
#     wb = load_workbook(file_path)
#     ws = wb["Known"]
#     dropdowns_ws = wb["DropDowns"]
#
#     # Identify the last row with data in SCE_Struct column
#     last_row = max((ws_row for ws_row in range(1, ws.max_row + 1) if ws.cell(row=ws_row, column=1).value is not None),
#                    default=1)
#
#     # Function to check and remove existing named range
#     def check_and_remove_named_range(wb, name):
#         for nr in wb.defined_names.definedName:
#             if nr.name == name:
#                 wb.defined_names.definedName.remove(nr)
#                 break
#
#     # Create named ranges for the dropdown references
#     print_to_widget("\nFinishing up...")
#
#     # Define your named ranges
#     named_range_R = DefinedName(name="DropDowns_R", attr_text=f"{dropdowns_ws.title}!$A$2:$A$8")
#     named_range_S = DefinedName(name="DropDowns_S", attr_text=f"{dropdowns_ws.title}!$B$2:$B$10")
#     named_range_T = DefinedName(name="DropDowns_T", attr_text=f"{dropdowns_ws.title}!$C$2:$C$30")
#
#     # Check and remove existing named ranges if they exist
#     check_and_remove_named_range(wb, "DropDowns_R")
#     check_and_remove_named_range(wb, "DropDowns_S")
#     check_and_remove_named_range(wb, "DropDowns_T")
#
#     # Append the new named ranges
#     wb.defined_names.append(named_range_R)
#     wb.defined_names.append(named_range_S)
#     wb.defined_names.append(named_range_T)
#
#     # Create data validation objects
#     with warnings.catch_warnings():
#         print(2)
#         warnings.simplefilter("ignore")
#         dv_R = DataValidation(type="list", formula1="=DropDowns_R", allow_blank=True, showDropDown=None)
#         dv_S = DataValidation(type="list", formula1="=DropDowns_S", allow_blank=True, showDropDown=None)
#         dv_T = DataValidation(type="list", formula1="=DropDowns_T", allow_blank=True, showDropDown=None)
#
#     # Apply data validation to columns S and T
#     print(3)
#     for ws_row in range(2, last_row + 1):
#         cell_r = ws["R{}".format(ws_row)]
#         cell_s = ws["S{}".format(ws_row)]
#         cell_t = ws["T{}".format(ws_row)]
#
#         # Apply data validation to each cell
#         dv_R.add(cell_r)
#         dv_S.add(cell_s)
#         dv_T.add(cell_t)
#
#     # Add data validations to the worksheet
#     print(4)
#     ws.add_data_validation(dv_R)
#     ws.add_data_validation(dv_S)
#     ws.add_data_validation(dv_T)
#
#     # Save the modified workbook
#     wb.save(file_path)
#     print_to_widget(f"\nFile saved as: {os.path.basename(file_path)}")


# Function to get the farthest image from the nadir
def get_farthest_from_nadir(root_directory):
    farthest_distances = {}  # Create a dictionary to store the farthest distances for each subfolder

    # Walk through the directory structure
    for dirpath, dirnames, filenames in os.walk(root_directory):
        n_image_path = None
        image_paths = []

        # Iterate through the files in the current directory
        for file in filenames:
            file_path = os.path.join(dirpath, file)

            if file_path.lower().endswith("n.jpg"):
                n_image_path = file_path
            elif file.lower().endswith((".jpg", ".jpeg")):
                image_paths.append(file_path)

        if n_image_path:
            n_coord = get_gps_from_image(n_image_path)

            if n_coord:
                max_distance = 0
                max_distance_img = None

                for img_path in image_paths:
                    coord = get_gps_from_image(img_path)
                    if coord:
                        dist_feet = distance_calculator(n_coord, coord)
                        if dist_feet > max_distance:
                            max_distance = dist_feet
                            max_distance_img = os.path.basename(img_path)
                    else:
                        img_name = os.path.basename(img_path)
                        print_to_widget(
                            f"Warning: No GPS data found on {img_name} from structure {os.path.basename(dirpath)}.")
                # Store the farthest distance in the dictionary
                farthest_distances[os.path.basename(dirpath)] = max_distance

            else:
                print_to_widget(f"Warning: No GPS data found on the nadir of {os.path.basename(dirpath)}.")
                farthest_distances[os.path.basename(dirpath)] = None
        elif image_paths:  # Check if the folder had images but no nadir
            print_to_widget(f"Warning: No nadir file found in {os.path.basename(dirpath)}.")
            farthest_distances[os.path.basename(dirpath)] = None

    return farthest_distances


# Main function to process image directories and generate output spreadsheet
def get_distances_from_nadir(root_directory, output_file):
    print_to_widget(f"\nCalculating distance of each image from the nadir...")
    print_to_widget(f"Adding \"Image Distances\" tab...")

    # Load existing workbook and select or create a new worksheet
    wb = load_workbook(output_file)
    ws = wb.create_sheet()
    ws.title = "Image Distances"

    # Write headers to the first row of the worksheet
    ws.cell(row=1, column=1, value="Folder Name")
    ws.cell(row=1, column=2, value="Image Name")
    ws.cell(row=1, column=3, value="Distance from Nadir (ft)")

    # Keep track of the current row and max distance for each folder
    current_row = 2
    max_distances = {}

    # Traverse the directory tree
    for dirpath, dirnames, filenames in os.walk(root_directory):
        folder = os.path.basename(dirpath)
        n_image_path = None
        image_paths = []

        # Iterate through the files in the current directory
        for file in filenames:
            file_path = os.path.join(dirpath, file)

            # Check if file is the nadir image (named with "n.jpg" suffix)
            if file_path.lower().endswith("n.jpg"):
                n_image_path = file_path
            # Check if file is a JPEG image
            elif file.lower().endswith((".jpg", ".jpeg")):
                image_paths.append(file_path)

        # Check if there is a nadir image in the folder
        if n_image_path:
            # Get GPS coordinates from nadir image
            n_coord = get_gps_from_image(n_image_path)

            # Check if nadir image has GPS data
            if n_coord:
                # Write nadir image data to worksheet
                ws.cell(row=current_row, column=1, value=folder)
                ws.cell(row=current_row, column=2, value=os.path.basename(n_image_path))
                ws.cell(row=current_row, column=3, value=0)
                ws.cell(row=current_row, column=3).number_format = '0.00'
                current_row += 1

                max_distance = 0  # initialize max distance to 0
                max_distance_indices = []  # initialize list of indices for max distance images

                # Iterate through the non-nadir images in the folder
                for img_idx, img_path in enumerate(image_paths):
                    # Get GPS coordinates from non-nadir image
                    coord = get_gps_from_image(img_path)

                    # Check if non-nadir image has GPS data
                    if coord:
                        # Calculate distance between nadir image and non-nadir image using geopy
                        dist_ft = distance_calculator(n_coord, coord)
                        # Write non-nadir image data to worksheet
                        ws.cell(row=current_row, column=1, value=folder)
                        ws.cell(row=current_row, column=2, value=os.path.basename(img_path))
                        ws.cell(row=current_row, column=3, value=dist_ft)
                        ws.cell(row=current_row, column=3).number_format = '0.00'
                        current_row += 1

                        # Update max distance and indices if necessary
                        if dist_ft > max_distance:
                            max_distance = dist_ft
                            max_distance_indices = [img_idx]
                        elif dist_ft == max_distance:
                            max_distance_indices.append(img_idx)
                    else:
                        # Handle the case when the coord is None
                        ws.cell(row=current_row, column=1, value=folder)
                        ws.cell(row=current_row, column=2, value=os.path.basename(img_path))
                        ws.cell(row=current_row, column=3, value="N/A")
                        current_row += 1

                # Store max distance for folder
                max_distances[folder] = max_distance

                # Format cells to highlight max distance
                for img_idx in max_distance_indices:
                    max_distance_cell = ws.cell(row=current_row - len(image_paths) + img_idx, column=3)
                    max_distance_cell.font = Font(bold=True)
                    max_distance_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00",
                                                         fill_type="solid")
                current_row += 1  # add a row space after the folder

            else:
                # Write error message to worksheet if nadir image does not have GPS data
                ws.cell(row=current_row, column=1, value=folder)
                ws.cell(row=current_row, column=2, value=os.path.basename(n_image_path))
                ws.cell(row=current_row, column=3, value="No GPS data found for os.path.basename(n_image_path)")
                current_row += 1
        else:
            # Check if any file in the folder has an image extension
            image_extensions = [".jpg", ".jpeg", ".png", ".bmp"]
            contains_images = any(file.lower().endswith(tuple(image_extensions)) for file in filenames)

            if not n_image_path and contains_images:
                # Write error message to worksheet if nadir image is not found in structure folder
                ws.cell(row=current_row, column=1, value=folder)
                ws.cell(row=current_row, column=2, value=f"No nadir image found in {folder}")
                ws.cell(row=current_row, column=3, value="")
                current_row += 1
            else:
                ws.cell(row=current_row, column=1, value=folder)

            current_row += 1

    # Apply formatting to max distance cells
    for ws_row in ws.iter_rows(min_row=2, min_col=1, max_col=3):
        folder = ws_row[0].value
        if folder in max_distances and ws_row[2].value == max_distances[folder]:
            max_distance_cell = ws_row[2]
            max_distance_cell.font = Font(bold=True)
            max_distance_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00",
                                                 fill_type="solid")

    # Save the workbook to the output file
    wb.save(output_file)
    return output_file


selected_name = None
traveler_sheet_paths = []
merge_ts_dest_file = None
merge_ts_source_sheet = None
merge_ts_dest_sheet = None
merge_ts_complete_traveler_func = complete_traveler_distro
merge_ts_columns_mapping = {}


def add_files():
    print_to_widget("\nSelect the traveler sheets you want to merge...")
    traveler_sheet_paths1 = filedialog.askopenfilenames(title="Select traveler sheets to merge",
                                                        filetypes=[("Excel files", "*.xlsx;*.xls")])
    if traveler_sheet_paths1:
        traveler_sheet_paths.extend(traveler_sheet_paths1)
        for path in traveler_sheet_paths1:
            file_name = os.path.basename(path)
            print_to_widget(f"Added traveler: {file_name}")


# Check if the user has provided both the Excel file and the source directory
def merge_travelers():
    global traveler_sheet_paths, merge_ts_dest_file, merge_ts_source_sheet, merge_ts_dest_sheet, \
        merge_ts_complete_traveler_func, merge_ts_columns_mapping
    if not traveler_sheet_paths:
        print_to_widget("\nNo traveler sheets selected.")
        return
    print_to_widget(f"\nSelect the traveler sheet template...")
    merge_ts_dest_file = filedialog.askopenfilename(title="Select the traveler sheet template",
                                                    filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
    print_to_widget(f"Selected traveler sheet template: {os.path.basename(merge_ts_dest_file)}")

    # Check if the filename contains specific keywords and set sheets accordingly
    file_name = os.path.basename(merge_ts_dest_file).lower()
    if "distribution" in file_name or "d_" in file_name:
        print_to_widget(f"\nMerging distribution traveler sheets...")
        merge_ts_source_sheet = "C2-Distribution"
        merge_ts_dest_sheet = "C2-Distribution"
        merge_ts_columns_mapping = {
            "C2_ID": "C2_ID",
            "BLOCK_ID": "BLOCK_ID",
            "Work Order No.": "Work Order No.",
            "SCE_STRUCT": "SCE_STRUCT",
            "HighFire": "HighFire",
            "AOC": "AOC",
            "EZ Pole": "EZ Pole",
            "Latitude": "Latitude",
            "Longitude": "Longitude",
            "Flight Team": "Flight Team",
            "Photo_Location": "Photo_Location",
            "FLIGHT_DATE": "FLIGHT_DATE",
            "Photo_Count": "Photo_Count",
            "FieldLAT": "FieldLAT",
            "FieldLong": "FieldLong",
            "VENDOR_STATUS": "VENDOR_STATUS",
            "VENDOR_CATEGORY": "VENDOR_CATEGORY",
            "VENDOR_NOTES": "VENDOR_NOTES",
            "QA Reviewer": "QA Reviewer",
            "EZ_in_Trans": "EZ_in_Trans"
        }
    elif "transmission" in file_name or "t_" in file_name:
        print_to_widget(f"\nMerging transmission traveler sheets...")
        merge_ts_source_sheet = "Known"
        merge_ts_dest_sheet = "Known"
        merge_ts_complete_traveler_func = complete_traveler_trans
        merge_ts_columns_mapping = {
            "REF_GLOBAL": "REF_GLOBAL",
            "BLOCK_ID": "BLOCK_ID",
            "DistrictNu": "DistrictNu",
            "SCE_STRUCT": "SCE_STRUCT",
            "CIRCUIT_NA": "CIRCUIT_NA",
            "CIRCUIT_FL": "CIRCUIT_FL",
            "VOLTAGE": "VOLTAGE",
            "Latitude": "Latitude",
            "Longitude": "Longitude",
            "SCE_Design": "SCE_Design",
            "Photo_Location": "Photo_Location",
            "Sub_Area": "Sub_Area",
            "Flight_Date": "Flight_Date",
            "Photo_Count": "Photo_Count",
            "LocationMethod": "LocationMethod",
            "FieldLAT": "FieldLAT",
            "FieldLong": "FieldLong",
            "Vendor_Status": "Vendor_Status",
            "Vendor_Category": "Vendor_Category",
            "Vendor_Notes": "Vendor_Notes",
            "QA Reviewer": "QA Reviewer",
            "Team_Number": "Team_Number",
            "EZ_in_Distro": "EZ_in_Distro"
        }
    else:
        print_to_widget("Please rename the template to contain 'D_' or 'distribution', or 'T_' or 'transmission' "
                        "in the filename.\n")
    mergetravelersheets()


dir_paths = []


def add_directory():
    dir_path = filedialog.askdirectory(title="Select directory to merge")
    if dir_path:
        dir_paths.append(dir_path)
        print_to_widget(f"Added directory: {dir_path}")


def remove_empty_folders(path):
    def has_files(dir_path):
        for _, _, filenames in os.walk(dir_path):
            if filenames:  # If there are any files, return True
                return True
        return False

    empty_folders = []
    for dirpath, dirnames, _ in os.walk(path, topdown=False):
        if all(not has_files(os.path.join(dirpath, d)) for d in dirnames) and not has_files(dirpath):
            empty_folders.append(dirpath)

    for folder in empty_folders:
        os.rmdir(folder)
        print_to_widget(f"Removed empty folder: {os.path.basename(folder)}")

    return path in empty_folders


def merge_directories():
    print_to_widget("Merging directories...")
    if not dir_paths:
        print_to_widget("No directories selected.")
        return

    merged_contents = {}

    def walk_directory(walk_dir_path):
        for dir_path, _, file_names in os.walk(walk_dir_path):
            for file_name in file_names:
                file_path = os.path.join(dir_path, filename)

                if os.path.isdir(filepath):
                    walk_directory(filepath)
                    continue

                dir_name = os.path.relpath(dir_path, walk_dir_path)

                if dir_name not in merged_contents:
                    merged_contents[dir_name] = []

                merged_contents[dir_name].append(filepath)

    for path in dir_paths:
        walk_directory(path)

    new_dir = os.path.join(os.path.commonpath(dir_paths), 'merged')
    if not os.path.exists(new_dir):
        os.makedirs(new_dir)

    for merge_dir_name, files in merged_contents.items():
        for filepath in files:
            filename = os.path.basename(filepath)
            new_path = os.path.join(new_dir, merge_dir_name, filename)

            if not os.path.exists(os.path.dirname(new_path)):
                os.makedirs(os.path.dirname(new_path))

            shutil.move(filepath, new_path)

    # Remove chosen directories if they and all their subdirectories are empty
    for path in dir_paths:
        if remove_empty_folders(path):
            continue
        else:
            print_to_widget(f"Directory not empty, not removed: {os.path.basename(path)}")

    print_to_widget("All empty folders are removed.")

    # Count the total number of images in all folders within each directory in dir_path and its subdirectories
    grand_total = 0
    for directory in os.listdir(new_dir):
        directory_path = os.path.join(new_dir, directory)
        if os.path.isdir(directory_path):
            total_image_count = count_images(directory_path)
            print_to_widget(f"{directory} contains {total_image_count} images.")
            grand_total += total_image_count
    print_to_widget(f"Total images in all directories: {grand_total}")
    dir_paths.clear()

    # Display a message box to indicate the process is complete
    print_to_widget(f'Directories merged successfully! New directory name: \"merged\"')
    messagebox.showinfo(title='Merge Directories', message='Directories merged successfully')


def extract_team_number(sheet_name):
    # Extract 3-digit team number from the sheet name
    match = re.search(r"\d{3}", sheet_name)
    if match:
        return int(match.group())
    else:
        return None


def gis_vs_uc(date_input, duc_file, gis_file):
    print_to_widget("Comparing data from GIS Extract and Upload Check...")
    date_input_obj = pd.to_datetime(date_input)  # Convert date_input to datetime

    # Read GIS file and filter data by date
    gis_df = pd.read_excel(gis_file)
    gis_df['Inspection_Date1'] = pd.to_datetime(gis_df['Inspection_Date1'])
    gis_df['Inspection_Date2'] = pd.to_datetime(gis_df['Inspection_Date2'])
    gis_df['Inspection_Date'] = gis_df['Inspection_Date2'].combine_first(gis_df['Inspection_Date1'])
    gis_df['Inspection_Date'] = gis_df['Inspection_Date'] - pd.Timedelta(hours=8)
    gis_selected_rows = gis_df[(gis_df['Inspection_Date'].dt.date == date_input_obj.date())]

    # Get list of team numbers from filtered GIS data
    team_numbers = gis_selected_rows['Team_Number'].astype(str).str.extract(r"(\d{3})")[0].unique()
    team_numbers = [int(float(team_number)) for team_number in map(str, team_numbers) if
                    team_number.replace('.', '', 1).isdigit()]

    # Creating the new Excel file in the same directory as the GIS file
    gis_dir = os.path.dirname(gis_file)
    date_input_converted = pd.to_datetime(date_input).strftime('%Y%m%d')

    counter = 1
    new_file = os.path.join(gis_dir,
                            f"GIS_vs_UploadCheck_{date_input_converted}.xlsx")  # Initial new file name

    # If the file exists, update the filename with a counter until we find one that doesn't exist
    while os.path.isfile(new_file):
        counter += 1
        new_file = os.path.join(gis_dir, f"GIS_vs_UploadCheck_{date_input_converted} ({counter}).xlsx")

    print_to_widget("Checking discrepancies...")
    with pd.ExcelFile(duc_file) as xls:
        for sheet_name in xls.sheet_names:
            team_number = extract_team_number(sheet_name)
            if team_number not in team_numbers:
                continue

            # Read the specific team's data from the Upload Check file
            df = pd.read_excel(xls, sheet_name=sheet_name, header=4)
            df['Flight Date'] = pd.to_datetime(df['Flight Date'])
            selected_rows = df[df['Flight Date'] == date_input_obj]

            structure_paths = selected_rows['SCE Structure Number or Folder Path'].str.split('\\').str[-1]
            df_structure_paths = pd.DataFrame(structure_paths.sort_values()).rename(
                columns={'SCE Structure Number or Folder Path': 'Folder Names'})

            gis_floc = gis_selected_rows[gis_selected_rows['Team_Number'] == team_number]['FLOC'].str.slice(start=3)
            df_gis_floc = pd.DataFrame(gis_floc.sort_values()).rename(columns={'FLOC': 'GIS ID'})

            df_structure_paths.reset_index(drop=True, inplace=True)
            df_gis_floc.reset_index(drop=True, inplace=True)

            df_combined = pd.concat([df_structure_paths, df_gis_floc], axis=1)

            if os.path.isfile(new_file):
                mode = 'a'  # append if the file exists
            else:
                mode = 'w'  # write a new file if it does not

            with pd.ExcelWriter(new_file, engine='openpyxl', mode=mode) as writer:
                df_combined.to_excel(writer, sheet_name=sheet_name, index=False)

            wb = load_workbook(new_file)
            sheet = wb[sheet_name]

            fill = PatternFill(start_color='FFFFC7CE', end_color='FFFFC7CE', fill_type='solid')
            font = Font(color='9C0006')

            value_counts = defaultdict(int)
            for sheet_row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=2):
                for cell in sheet_row:
                    value_counts[cell.value] += 1

            for sheet_row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=2):
                for cell in sheet_row:
                    if value_counts[cell.value] < 2:
                        cell.fill = fill
                        cell.font = font

            wb.save(new_file)
    print_to_widget(f"File saved as {os.path.basename(new_file)}")


def gis_vs_ts(gis_file, traveler_file):
    print_to_widget("Comparing data from GIS Extract and Traveler Sheet...")

    # Load the data from the Excel files
    ts_df = pd.read_excel(traveler_file, sheet_name=None)  # Read all sheets into a dictionary
    gis_df = pd.read_excel(gis_file)

    ts_key = 'SCE_STRUCT'
    if "C2-Distribution" in ts_df:
        ts_df = ts_df["C2-Distribution"]
        gis_key = 'FLOC'
        output_prefix = "D"
    else:
        ts_df = ts_df["Known"]
        ts_df[ts_key] = ts_df[ts_key].astype(str)
        gis_key = 'FLOC'  # Use 'FLOC' from GIS when "Known" sheet is used
        gis_df[gis_key] = gis_df[gis_key].apply(
            lambda x: x[3:] if pd.notnull(x) else x)  # remove first 3 characters from 'FLOC' in GIS
        output_prefix = "T"

    # Merge the two data frames based on IDs
    merged_df = pd.merge(ts_df, gis_df, how='left', left_on=ts_key, right_on=gis_key)

    # Create new column "Calculated"
    merged_df['Calculated'] = False

    # Define a function to calculate the geodesic distance
    def calculate_distance(coords_row):
        if pd.isnull(coords_row[gis_key]):
            if pd.notnull(coords_row['FieldLAT']) and pd.notnull(coords_row['FieldLong']):
                ts_coords = (coords_row['FieldLAT'], coords_row['FieldLong'])
                # Calculate the distance to all GIS locations
                distances = gis_df.apply(lambda gis_row: geodesic((gis_row['Mapped_Lat'], gis_row['Mapped_Lon']),
                                                                  ts_coords).feet if pd.notnull(
                    gis_row['Mapped_Lat']) and pd.notnull(gis_row['Mapped_Lon']) else float('NaN'), axis=1)
                # Get the index of the closest GIS location
                closest_index = distances.idxmin()
                # Assign the closest GIS location ID to the 'FLOC' column
                coords_row[gis_key] = gis_df.loc[closest_index, gis_key]
                coords_row['Mapped_Lon'] = gis_df.loc[closest_index, 'Mapped_Lon']
                coords_row['Mapped_Lat'] = gis_df.loc[closest_index, 'Mapped_Lat']
                coords_row['Calculated'] = True  # Add a column to mark the coords_row for highlighting
        return coords_row

    # Apply the function to each coords_row
    merged_df = merged_df.apply(calculate_distance, axis=1)

    # Drop rows where 'FieldLAT' or 'FieldLong' is NaN or empty
    merged_df = merged_df.dropna(subset=['FieldLAT', 'FieldLong'])

    # Ensure 'FieldLAT' and 'FieldLong' are float type
    merged_df['FieldLAT'] = pd.to_numeric(merged_df['FieldLAT'], errors='coerce')
    merged_df['FieldLong'] = pd.to_numeric(merged_df['FieldLong'], errors='coerce')

    # Drop any rows where 'FieldLAT' or 'FieldLong' could not be converted to numeric
    merged_df = merged_df.dropna(subset=['FieldLAT', 'FieldLong'])

    # Create new column "Distance"
    merged_df['Structure Distance from Coordinates (ft)'] = merged_df.apply(
        lambda distance_row:
        round(
            geodesic(
                (distance_row['FieldLAT'], distance_row['FieldLong']),
                (distance_row['Mapped_Lat'], distance_row['Mapped_Lon'])
            ).feet,
            2
        )
        if all(pd.notnull(distance_row[distance_col]) for distance_col in ['FieldLAT', 'FieldLong', 'Mapped_Lat',
                                                                           'Mapped_Lon'])
        else 'Refly',
        axis=1
    )
    # Keep only the columns you need
    final_df = merged_df[['SCE_STRUCT', gis_key, 'Calculated', 'Structure Distance from Coordinates (ft)']]

    # Rename the columns for the new file
    final_df.columns = ['TS IDs', 'GIS Matches', 'Calculated', 'Structure Distance from Coordinates (ft)']

    # Creating the new Excel file in the same directory as the GIS file
    gis_dir = os.path.dirname(gis_file)
    counter = 1
    new_file = os.path.join(gis_dir, f"{output_prefix}_GIS_vs_Traveler.xlsx")  # Initial new file name

    # If the file exists, update the filename with a counter until we find one that doesn't exist
    while os.path.isfile(new_file):
        counter += 1
        new_file = os.path.join(gis_dir, f"{output_prefix}_GIS_vs_Traveler ({counter}).xlsx")

    # Save the data frame to a new Excel file
    final_df.to_excel(new_file, index=False)

    # Open the workbook with openpyxl
    wb = load_workbook(new_file)

    # Select the active worksheet
    ws = wb.active

    # Define your styles
    fill = PatternFill(start_color='FFFFC7CE', end_color='FFFFC7CE', fill_type='solid')
    font = Font(color='9C0006')

    # Iterate through the rows of the worksheet and format issues
    for ws_row in ws.iter_rows(min_row=2):  # Start from the second coords_row, assuming first coords_row is the header
        if ws_row[2].value is True:  # Assuming 'Calculated' column is the third one
            ws_row[0].fill = fill  # Assuming 'TS IDs' column is the first one
            ws_row[0].font = font
            ws_row[1].fill = fill  # Assuming 'GIS Matches' column is the second one
            ws_row[1].font = font
        elif isinstance(ws_row[3].value, (int, float)) and ws_row[3].value > 200:
            ws_row[3].fill = fill
            ws_row[3].font = font

    # Drop the 'Calculated' column
    ws.delete_cols(3)  # Assuming 'Calculated' column is the third one

    # Save the workbook with the new formatting
    wb.save(new_file)

    print_to_widget(f"File saved as {os.path.basename(new_file)}")


def delta_report(gis_file, ts_file, uc_file):
    # Load GIS file
    print_to_widget("Filtering GIS...")
    gis_df = pd.read_excel(gis_file)
    gis_df['Inspection_Date1'] = gis_df['Inspection_Date1'].where(gis_df['Inspection_Date2'].isna(),
                                                                  gis_df['Inspection_Date2'])
    gis_df = gis_df[gis_df['Inspection_Date1'].notna()]
    gis_df['FLOC'] = gis_df.apply(lambda gis_row: gis_row['FLOC'], axis=1)
    gis_ids = pd.Series(gis_df['FLOC']).str.replace('OH-', '').sort_values().reset_index(drop=True)

    # Load TS file
    print_to_widget("Filtering Traveler Sheet...")
    ts_xls = pd.ExcelFile(ts_file)
    sheet_name = 'C2-Distribution' if 'C2-Distribution' in ts_xls.sheet_names else 'Sheet1'
    ts_df = ts_xls.parse(sheet_name)
    ts_df = ts_df[ts_df['FLIGHT_DATE'].notna()]
    ts_ids = pd.Series(ts_df['SCE_STRUCT']).astype(str).sort_values().reset_index(drop=True)

    # Load UC file
    print_to_widget("Filtering Upload Check...\n")
    uc_xls = pd.ExcelFile(uc_file)
    uc_dfs = []
    for sheet_name in uc_xls.sheet_names:
        if re.match(r'\d{3} -', sheet_name) and len(uc_xls.parse(
                sheet_name)) > 4:  # Check if sheet name starts with three digits and a dash, and has at least 5 rows
            uc_df = uc_xls.parse(sheet_name, header=4)
            uc_df = uc_df.dropna(subset=['SCE Structure Number or Folder Path'])
            uc_dfs.append(uc_df)

    # Assuming uc_dfs is a list of DataFrame objects
    cleaned_uc_dfs = []

    for df in uc_dfs:
        # Drop columns where all elements are NA
        df = df.dropna(axis=1, how='all')
        cleaned_uc_dfs.append(df)

    # Concatenate the cleaned DataFrames
    uc_df = pd.concat(cleaned_uc_dfs)
    uc_ids = pd.Series(uc_df['SCE Structure Number or Folder Path']).str.split('\\').str[-1]

    # Create pandas Series objects for easier handling
    gis_ids = pd.Series(gis_ids)
    ts_ids = pd.Series(ts_ids)
    uc_ids = pd.Series(uc_ids)

    # Convert to sets for faster membership checks
    gis_set = set(gis_ids.values)
    ts_set = set(ts_ids.values)
    uc_set = set(uc_ids.values)

    # Create DataFrame for results
    results_df = pd.DataFrame(columns=['GIS', 'TS', 'UC'])

    # Check if each GIS id exists in the other dataframes
    print_to_widget("Combining Dataframes...")
    # Check for GIS IDs
    for gis_id in gis_set:
        exists_in_ts = gis_id in ts_set
        exists_in_uc = gis_id in uc_set
        new_row = pd.DataFrame(
            {'GIS': [gis_id], 'TS': [gis_id if exists_in_ts else ''], 'UC': [gis_id if exists_in_uc else '']})
        results_df = pd.concat([results_df, new_row], ignore_index=True)

    # Check for TS IDs
    for ts_id in ts_set:
        if ts_id not in gis_set:
            exists_in_uc = ts_id in uc_set
            new_row = pd.DataFrame({'GIS': [''], 'TS': [ts_id], 'UC': [ts_id if exists_in_uc else '']})
            results_df = pd.concat([results_df, new_row], ignore_index=True)

    # Check for UC IDs
    for uc_id in uc_set:
        if uc_id not in gis_set and uc_id not in ts_set:
            new_row = pd.DataFrame({'GIS': [''], 'TS': [''], 'UC': [uc_id]})
            results_df = pd.concat([results_df, new_row], ignore_index=True)

    # Define a function that returns the match condition of a row
    print_to_widget("Checking Structure IDs missing in Traveler Sheet...")

    def get_match_condition(match_row):
        # Check for presence of each condition
        gis_present = match_row['GIS']
        ts_present = match_row['TS']
        uc_present = match_row['UC']

        # Determine the match condition
        if gis_present and ts_present and uc_present:
            return 'all_match'
        elif gis_present and ts_present:
            return 'GIS_TS_match'
        elif gis_present and uc_present:
            return 'GIS_UC_match'
        elif ts_present and uc_present:
            return 'TS_UC_match'
        elif gis_present:
            return 'GIS_only'
        elif ts_present:
            return 'TS_only'
        elif uc_present:
            return 'UC_only'
        else:
            return 'unknown'

    # Create the 'match_condition' column
    results_df['match_condition'] = results_df.apply(get_match_condition, axis=1)

    # Define the order
    order = ['all_match', 'GIS_TS_match', 'GIS_UC_match', 'TS_UC_match', 'GIS_only', 'TS_only', 'UC_only']

    # Convert 'match_condition' to a Categorical data type
    results_df['match_condition'] = pd.Categorical(results_df['match_condition'], categories=order, ordered=True)

    # Sort by the 'match_condition' column first, then within each condition, sort by 'GIS', 'TS', 'UC'
    results_df = results_df.sort_values(by=['match_condition', 'GIS', 'TS', 'UC'])

    # Drop the 'match_condition' column
    results_df = results_df.drop(columns='match_condition')

    # Subtract 8 hours to inspection date
    gis_df['Inspection_Date1'] = pd.to_datetime(gis_df['Inspection_Date1']).dt.date - pd.Timedelta(hours=8)

    # Create dictionaries mapping FLOC IDs (without 'OH-') to other columns
    id_to_date = dict(zip(gis_df['FLOC'].str.replace('OH-', ''), gis_df['Inspection_Date1']))
    id_to_team = dict(zip(gis_df['FLOC'].str.replace('OH-', ''), gis_df['Team_Number']))
    id_to_status = dict(zip(gis_df['FLOC'].str.replace('OH-', ''), gis_df['Vendor_Status']))
    id_to_notes = dict(zip(gis_df['FLOC'].str.replace('OH-', ''), gis_df['Vendor_Notes']))

    # Create a set for TS Missing IDs
    ts_missing_set = set()
    # Check for TS Missing IDs
    for gis_id in gis_set:
        exists_in_ts = gis_id in ts_set
        if not exists_in_ts:
            ts_missing_set.add(gis_id)
    # Convert the set to a DataFrame
    ts_missing_df = pd.DataFrame(list(ts_missing_set), columns=['Missing in TS'])

    # Create 'Inspection Date', 'Team_Number', 'Vendor_Status' columns by mapping 'Missing in TS' column to the dates
    # in the dictionary
    ts_missing_df['Flight Date'] = ts_missing_df['Missing in TS'].map(id_to_date)
    ts_missing_df['Flight Team'] = ts_missing_df['Missing in TS'].map(id_to_team)
    ts_missing_df['Vendor Status'] = ts_missing_df['Missing in TS'].map(id_to_status)
    ts_missing_df['Crew Notes'] = ts_missing_df['Missing in TS'].map(id_to_notes)
    ts_missing_df['In Upload Check?'] = ts_missing_df['Missing in TS'].apply(lambda x: 'Yes' if x in uc_set else 'No')

    # Create a set for GIS Missing IDs
    print_to_widget("Checking Structure IDs missing in GIS...")
    gis_missing_set = set()
    # Check for GIS Missing IDs
    for ts_id in ts_set:
        exists_in_gis = ts_id in gis_set
        if not exists_in_gis:
            gis_missing_set.add(ts_id)
    # Convert the set to a DataFrame
    gis_missing_df = pd.DataFrame(list(gis_missing_set), columns=['Need GIS Update'])

    # Create dictionaries mapping SCE_STRUCT IDs to other columns
    if 'Flight Team' in ts_df.columns:
        id_to_team = dict(zip(ts_df['SCE_STRUCT'].astype(str), ts_df['Flight Team']))
        gis_missing_df['Flight Team'] = gis_missing_df['Need GIS Update'].map(id_to_team)

    id_to_date = dict(zip(ts_df['SCE_STRUCT'].astype(str), ts_df['FLIGHT_DATE'].dt.date))
    gis_missing_df['Flight Date'] = gis_missing_df['Need GIS Update'].map(id_to_date)

    if 'CIRCUIT_FL' in ts_df.columns:
        id_to_type = dict(zip(ts_df['SCE_STRUCT'].astype(str), ts_df['CIRCUIT_FL']))
        gis_missing_df['Type'] = gis_missing_df['Need GIS Update'].map(id_to_type)

    # Get the prefix from the TS file
    prefix = os.path.basename(ts_file).split('_')[0] + "_" if os.path.basename(ts_file).startswith(('D_', 'T_')) else ""

    # Write to Excel
    output_file = os.path.join(os.path.dirname(gis_file), f'{prefix}Delta Report.xlsx')
    counter = 1
    while os.path.isfile(output_file):
        output_file = os.path.join(os.path.dirname(gis_file), f'{prefix}Delta Report ({counter}).xlsx')
        counter += 1

    with pd.ExcelWriter(output_file) as writer:
        results_df.to_excel(writer, sheet_name='GIS vs TS vs UC', index=False)
        ts_missing_df.to_excel(writer, sheet_name='TS Missing', index=False)
        gis_missing_df.to_excel(writer, sheet_name='GIS Missing', index=False)

    # Load the workbook
    book = load_workbook(output_file)

    # For each sheet in workbook
    for sheet_name in book.sheetnames:
        # Select the worksheet
        sheet = book[sheet_name]

        # Create a table
        max_col = sheet.max_column  # Get max columns in the worksheet
        max_row = sheet.max_row  # Get max row in the worksheet

        # Create a Table Style object
        tab = Table(displayName=f"Table_{sheet_name.replace(' ', '_')}",
                    ref=f"A1:{get_column_letter(max_col)}{max_row}")

        # Set table style to 'Table Style Light 1'
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleLight1",
                                            showFirstColumn=False,
                                            showLastColumn=False,
                                            showRowStripes=True,
                                            showColumnStripes=False)

        # Add table to the worksheet
        sheet.add_table(tab)

        # Wrap print_text for headers
        for sheet_col in range(1, max_col + 1):
            cell = sheet.cell(row=1, column=sheet_col)
            cell.alignment = Alignment(wrap_text=True)

    # Save the workbook
    book.save(output_file)
    print_to_widget(f"Results saved as \"{os.path.basename(output_file)}\".")


def create_team_sheets(teams_worked_today, start_date, end_date, duc_file, gis_file, output_file):
    print_to_widget("\nCreating new sheets for each team to compare inspected structures between GIS and Upload "
                    "Check...")
    start_date_obj = datetime.strptime(start_date, "%Y%m%d").date()
    end_date_obj = datetime.strptime(end_date, "%Y%m%d").date()  # Convert date_input to datetime

    # Read GIS file and filter data by teams and date
    gis_df = pd.read_excel(gis_file)
    gis_df['Inspection_Date1'] = pd.to_datetime(gis_df['Inspection_Date1'])
    gis_df['Inspection_Date2'] = pd.to_datetime(gis_df['Inspection_Date2'])
    gis_df['Inspection_Date'] = gis_df['Inspection_Date2'].combine_first(gis_df['Inspection_Date1'])
    gis_df['Inspection_Date'] = gis_df['Inspection_Date'] - pd.Timedelta(hours=8)
    gis_selected_rows = gis_df[
        (gis_df['Inspection_Date'].dt.date >= start_date_obj) & (gis_df['Inspection_Date'].dt.date <= end_date_obj)]

    with pd.ExcelFile(duc_file) as xls:
        for sheet_name in xls.sheet_names:
            if any(str(team) in sheet_name for team in teams_worked_today):
                team_number = [team for team in teams_worked_today if str(team) in sheet_name][0]
                df = pd.read_excel(xls, sheet_name=sheet_name,
                                   header=4)  # Read data from sheet, headers on 5th row (0-indexed)
                # Remove rows where 'Flight Date' is NaT or null
                df = df.dropna(subset=['Flight Date'])
                df['Flight Date'] = pd.to_datetime(df['Flight Date'])  # Convert 'Flight Date' to datetime
                selected_rows = df[(df['Flight Date'].dt.date >= start_date_obj) & (df['Flight Date'].dt.date <=
                                                                                    end_date_obj)].copy()
                # Convert the column to string type
                selected_rows['SCE Structure Number or Folder Path'] = selected_rows[
                    'SCE Structure Number or Folder Path'].astype(str)
                # Now apply your string operation
                folder_paths = selected_rows['SCE Structure Number or Folder Path'].str.split('\\').str[...]
                # Get the last part of 'SCE Structure Number or Folder Path' of selected rows
                folder_paths = selected_rows['SCE Structure Number or Folder Path'].str.split('\\').str[-1]
                # Create a DataFrame from folder_paths, sort and rename column to 'Folder Name'
                df_folder_paths = pd.DataFrame(folder_paths.sort_values()).rename(
                    columns={'SCE Structure Number or Folder Path': 'Folder Name'})

                # Processing GIS data for Current Team
                gis_team_data = gis_selected_rows[gis_selected_rows['Team_Number'] == team_number]
                gis_floc = gis_team_data['FLOC'].str.slice(start=3)
                df_gis_floc = pd.DataFrame(gis_floc.sort_values()).rename(columns={'FLOC': 'GIS ID'})

                # Extract FLOC values where Eq_ObjType is 'EZ_POLE'
                ez_pole_flocs_in_gis = gis_df[gis_df['Eq_ObjType'] == 'EZ_POLE']['FLOC'].str.slice(start=3).tolist()

                # Reset index of df_folder_paths and df_gis_floc
                df_folder_paths.reset_index(drop=True, inplace=True)
                df_gis_floc.reset_index(drop=True, inplace=True)

                # Merge df_folder_paths and df_gis_floc
                df_combined = pd.concat([df_folder_paths, df_gis_floc], axis=1)

                # Check if output_file already exists, if so, then append it, else create a new ExcelWriter object
                if os.path.exists(output_file):
                    with pd.ExcelWriter(output_file, engine='openpyxl', mode='a') as writer:
                        df_combined.to_excel(writer, sheet_name=sheet_name, index=False)
                else:
                    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                        df_combined.to_excel(writer, sheet_name=sheet_name, index=False)

                # Open output_file and select the sheet
                wb = load_workbook(output_file)
                sheet = wb[sheet_name]

                # Define fill style for unique values
                fill = PatternFill(start_color='FFFFC7CE', end_color='FFFFC7CE', fill_type='solid')
                font = Font(color='9C0006')
                ez_pole_font = Font(italic=True, bold=True)

                # Create a default dict for count of each value in columns A and B
                value_counts = defaultdict(int)

                for sheet_row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=2):
                    for cell in sheet_row:
                        value_counts[cell.value] += 1

                # Apply formatting to cells
                for sheet_row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=2):
                    for cell in sheet_row:
                        # Apply ez_pole_font to cells with values in ez_pole_flocs
                        if cell.value in ez_pole_flocs_in_gis and value_counts[cell.value] == 1:
                            cell.fill = fill
                            cell.font = ez_pole_font
                        elif cell.value in ez_pole_flocs_in_gis:
                            cell.font = ez_pole_font
                        # Apply fill and font to unique values
                        elif value_counts[cell.value] == 1:
                            cell.fill = fill
                            cell.font = font

                # Save the workbook
                wb.save(output_file)

    print_to_widget(f"Output file saved in: {output_file}")


def data_conformance_distro(GIS_file, EOD_file, DUC_file, template_file, start_date, end_date=None):
    # Parse the user input to a datetime.date object
    start_date_obj = datetime.strptime(start_date, "%Y%m%d").date()
    if end_date:
        end_date_obj = datetime.strptime(end_date, "%Y%m%d").date()
    else:
        end_date_obj = start_date_obj

    # Read the GIS file using pandas
    print_to_widget("Opening GIS Extract file...")
    GIS_df = pd.read_excel(GIS_file)

    # Read the EOD file using pandas
    print_to_widget("Opening End-Of-Day file...")
    EOD_df = pd.read_excel(EOD_file)

    # Read the DUC file using openpyxl
    print_to_widget("Opening Distribution Upload Check file...")
    DUC_wb = load_workbook(DUC_file)

    # Process EOD Extract
    print_to_widget("Processing End-Of-Day Extract...")
    # Convert the "Date" column to datetime objects
    EOD_df["Date"] = pd.to_datetime(EOD_df["Date"]).dt.date
    # Filter the DataFrame to get rows with the entered date and create a new DataFrame
    EOD_date_filtered_df = EOD_df[(EOD_df["Date"] >= start_date_obj) &
                                  (EOD_df["Date"] <= end_date_obj)].copy()
    # Convert 'Enter your Team #' and 'Total Structure Count' columns to integers
    EOD_date_filtered_df['Please enter your team number:'] = EOD_date_filtered_df[
        'Please enter your team number:'].astype(int)
    EOD_date_filtered_df['Total Structure Count'] = EOD_date_filtered_df['Total Structure Count'].astype(int)
    # Create a dictionary to store team numbers and their "Total Structure Count" values from EOD_file
    # Step 1: Sort by date in descending order
    EOD_date_filtered_df.sort_values(by=["Date"], ascending=False, inplace=True)
    # Step 2: Drop duplicates for same day same team, retaining the latest one
    EOD_date_filtered_df.drop_duplicates(subset=["Date", "Please enter your team number:"], keep="first",
                                         inplace=True)
    # Step 3: Group by team and sum the counts
    grouped_counts = EOD_date_filtered_df.groupby('Please enter your team number:')['Total Structure Count'].sum()
    # Step 4: Convert the grouped data to a dictionary
    team_structure_count = grouped_counts.to_dict()

    # Process GIS Extract
    print_to_widget("Processing GIS Extract...")
    # Convert the "Inspection_Date1" and "Inspection_Date2" columns to datetime objects and subtract 8 hours
    GIS_df["Inspection_Date1"] = pd.to_datetime(GIS_df["Inspection_Date1"]) - pd.Timedelta(hours=8)
    GIS_df["Inspection_Date2"] = pd.to_datetime(GIS_df["Inspection_Date2"]) - pd.Timedelta(hours=8)
    # Extract only the date part
    GIS_df["Inspection_Date1"] = GIS_df["Inspection_Date1"].dt.date
    GIS_df["Inspection_Date2"] = GIS_df["Inspection_Date2"].dt.date
    # Create a new column "Inspection_Date" that gives priority to "Inspection_Date2" over "Inspection_Date1"
    GIS_df["Inspection_Date"] = GIS_df["Inspection_Date2"].where(GIS_df["Inspection_Date2"].notna(),
                                                                 GIS_df["Inspection_Date1"])
    # Filter the DataFrame to get rows with the entered date
    date_filtered_df = GIS_df[(GIS_df["Inspection_Date"] >= start_date_obj) &
                              (GIS_df["Inspection_Date"] <= end_date_obj)]
    # Get the "Team_Number" column values as a list
    team_numbers = date_filtered_df["Team_Number"].tolist()
    # Count occurrences of each team number using Counter from collections
    team_number_count = Counter(team_numbers)
    # Count occurrences of each team number inspecting "EZ_POLE" on the specific date
    ez_pole_filtered_df = date_filtered_df[date_filtered_df["Eq_ObjType"] == "EZ_POLE"]
    ez_pole_team_counts = ez_pole_filtered_df["Team_Number"].value_counts().to_dict()

    # Process DUC file
    print_to_widget("Processing Distribution Upload Check...")
    # Initialize a dictionary to store the team numbers and their number of occurrences in the Upload Check file
    team_flight_date_count = {}
    # Iterate through each sheet in the DUC file
    for sheet_name in DUC_wb.sheetnames:
        try:
            # Extract the team number from the sheet name
            team_number = int(sheet_name[:3])
            print_to_widget(f"Processing team {team_number}'s upload check...")
        except ValueError:
            # Skip this sheet if it doesn't have the expected naming pattern
            continue

        # Get the sheet object
        sheet = DUC_wb[sheet_name]

        # Count the occurrences of dates within the range in the "Flight Date" column
        date_count = 0
        for sheet_row in range(6, sheet.max_row + 1):
            flight_date = sheet.cell(row=sheet_row, column=7).value
            if flight_date:
                # Parse the flight_date to a datetime object
                if isinstance(flight_date, str):
                    flight_date = parse(flight_date)
                # Check if flight_date is within the provided range
                if start_date_obj <= flight_date.date() <= end_date_obj:
                    date_count += 1

        # Add the team number and their corresponding number of occurrences to the dictionary
        team_flight_date_count[team_number] = date_count

    # Load the template using openpyxl
    print_to_widget("\nTransferring data to Data Conformance template...")
    template_wb = load_workbook(template_file)
    template_ws = template_wb['Pilot']

    # Write the Date or Date Range on E3
    if start_date_obj == end_date_obj:
        formatted_date = "Date: " + start_date_obj.strftime('%m/%d/%Y')
    else:
        formatted_date = "Date: " + start_date_obj.strftime('%m/%d/%Y') + " to " + end_date_obj.strftime(
            '%m/%d/%Y')

    template_ws['E3'] = formatted_date

    # Update the template with the number of structures inspected by each team
    for template_row in range(5, template_ws.max_row - 1):
        team_number = template_ws.cell(row=template_row, column=4).value

        if team_number in team_number_count:
            template_ws.cell(row=template_row, column=6).value = team_number_count[team_number]
            print_to_widget(f"\nGIS: Team {team_number} inspected {team_number_count[team_number]} structures")
        else:
            template_ws.cell(row=template_row, column=6).value = 0
            print_to_widget(f"\nGIS: Team {team_number} inspected 0 structures")

        if team_number in team_structure_count:
            template_ws.cell(row=template_row, column=5).value = team_structure_count[team_number]
            print_to_widget(f"EOD: Team {team_number} inspected {team_structure_count[team_number]} structures")
        else:
            template_ws.cell(row=template_row, column=5).value = 0
            print_to_widget(f"EOD: Team {team_number} inspected 0 structures")

        if team_number in team_flight_date_count:
            template_ws.cell(row=template_row, column=7).value = team_flight_date_count[team_number]
            print_to_widget(f"DUC: Team {team_number} inspected {team_flight_date_count[team_number]} structures")
        else:
            template_ws.cell(row=template_row, column=7).value = 0
            print_to_widget(f"DUC: Team {team_number} inspected 0 structures")

        # Update the template with the number of "EZ_POLE" inspected by each team
        if team_number in ez_pole_team_counts:
            template_ws.cell(row=template_row, column=10).value = ez_pole_team_counts[team_number]
            print_to_widget(f"GIS: Team {team_number} inspected {ez_pole_team_counts[team_number]} EZ Poles")
        else:
            template_ws.cell(row=template_row, column=10).value = 0
            print_to_widget(f"GIS: Team {team_number} inspected 0 EZ Poles")

        # Get the values in columns 3, 4, and 5
        eod_value = template_ws.cell(row=template_row, column=5).value
        gis_value = template_ws.cell(row=template_row, column=6).value
        duc_value = template_ws.cell(row=template_row, column=7).value

        # Compare the values
        if eod_value == gis_value == duc_value != 0:
            # Create a green fill and font for consistent values
            green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            green_font = Font(color='006100')
            # Format cells if values are consistent
            format_cell(template_ws.cell(row=template_row, column=5), eod_value, green_fill, green_font)
            format_cell(template_ws.cell(row=template_row, column=6), gis_value, green_fill, green_font)
            format_cell(template_ws.cell(row=template_row, column=7), duc_value, green_fill, green_font)
            template_ws.cell(row=template_row, column=22).value = 1

        elif eod_value != gis_value or gis_value != duc_value:
            # Create a red fill and font for inconsistent values
            red_fill = PatternFill(start_color='FFFFC7CE', end_color='FFFFC7CE', fill_type='solid')
            red_font = Font(color='9C0006')
            # Format cells if values are inconsistent
            format_cell(template_ws.cell(row=template_row, column=5), eod_value, red_fill, red_font)
            format_cell(template_ws.cell(row=template_row, column=6), gis_value, red_fill, red_font)
            format_cell(template_ws.cell(row=template_row, column=7), duc_value, red_fill, red_font)
            template_ws.cell(row=template_row, column=22).value = 0

    # Initialize sums and list
    eod_sum = 0
    gis_sum = 0
    duc_sum = 0
    ez_pole_sum = 0
    team_count = 0
    teams_worked_today: list[int] = []

    for template_row in range(5, template_ws.max_row - 1):  # max_row - 1, if "TOTAL" is the second-to-last template_row
        # Calculate the sum of each column
        eod_sum += template_ws.cell(row=template_row, column=5).value
        gis_sum += template_ws.cell(row=template_row, column=6).value
        duc_sum += template_ws.cell(row=template_row, column=7).value
        ez_pole_sum += template_ws.cell(row=template_row, column=10).value

        # Count the non-zero EOD, GIS, and DUC values
        if (template_ws.cell(row=template_row, column=5).value != 0) or \
                (template_ws.cell(row=template_row, column=6).value != 0) or \
                (template_ws.cell(row=template_row, column=7).value != 0):
            team_count += 1
            team_number = template_ws.cell(row=template_row, column=4).value
            if team_number not in teams_worked_today:  # To ensure no duplicate entries
                teams_worked_today.append(team_number)

    # Calculate the average structures per team per day
    num_days = (end_date_obj - start_date_obj).days + 1  # +1 to include both start and end dates

    # Ensure that team_count and num_days are not zero
    if team_count > 0 and num_days > 0:
        duc_avg = round(duc_sum / team_count / num_days)
    else:
        # Handle the case where division by zero would occur
        duc_avg = 0

    bold_font_sum = Font(bold=True)

    # Get the row number of the "TOTAL" row, assuming it's the second-to-last row
    total_row = template_ws.max_row - 1

    # Set the number of teams that worked (non-zero EOD, GIS, or DUC values) in the "Total" row
    template_ws.cell(row=total_row, column=4).value = f"{str(team_count)} Crews"
    template_ws.cell(row=total_row, column=4).font = bold_font_sum

    # Set the average number of structures inspected per team in the "Average" row
    template_ws.cell(row=total_row + 1, column=4).value = f"{str(duc_avg)} Structures"
    template_ws.cell(row=total_row + 1, column=4).font = bold_font_sum

    # Set the sum of each column in the "TOTAL" row
    template_ws.cell(row=total_row, column=5).value = eod_sum
    template_ws.cell(row=total_row, column=5).font = bold_font_sum
    template_ws.cell(row=total_row, column=6).value = gis_sum
    template_ws.cell(row=total_row, column=6).font = bold_font_sum
    template_ws.cell(row=total_row, column=7).value = duc_sum
    template_ws.cell(row=total_row, column=7).font = bold_font_sum

    # Set the sum in the "TOTAL" row for the "EZ Poles" column
    template_ws.cell(row=total_row, column=10).value = ez_pole_sum
    template_ws.cell(row=total_row, column=10).font = bold_font_sum

    # Check for multiple entries
    # Group by Date and Team, then count number of rows for each group
    grouped = EOD_date_filtered_df.groupby(["Date", "Please enter your team number:"]).size().reset_index(
        name='counts')

    # Filter for groups with counts greater than 1, indicating duplicate entries on the same date
    duplicates = grouped[grouped['counts'] > 1]

    if not duplicates.empty:
        # Get the unique teams with duplicates on the same date
        duplicate_teams = duplicates['Please enter your team number:'].unique()

        print_to_widget(f"\nWarning: The following team(s) have multiple EOD entries on the same date:"
                        f"\n{duplicate_teams}\nOnly the latest entry is recorded.\n")

    # Load "Inspector" sheet from template
    inspector_ws = template_wb['Inspector']

    # Write the Date or Date Range on E3
    if start_date_obj == end_date_obj:
        formatted_date = "Date: " + start_date_obj.strftime('%m/%d/%Y')
    else:
        formatted_date = "Date: " + start_date_obj.strftime('%m/%d/%Y') + " to " + end_date_obj.strftime(
            '%m/%d/%Y')

    inspector_ws['E3'] = formatted_date

    # Convert filtered GIS DataFrame to worksheet
    gis_ws = Workbook().active
    for gis_r_idx, gis_row in enumerate(dataframe_to_rows(date_filtered_df, index=False, header=True), 1):
        for gis_c_idx, value in enumerate(gis_row, 1):
            gis_ws.cell(row=gis_r_idx, column=gis_c_idx, value=value)

    # Create a dictionary to hold data for each inspector
    inspector_data = {}

    # Find indices of columns
    inspector_col_index = None
    p3_count_col_index = None
    p2_count_col_index = None
    p1_count_col_index = None

    for gis_col in range(1, gis_ws.max_column):
        header_value = gis_ws.cell(row=1, column=gis_col).value
        if header_value == 'Inspector':
            inspector_col_index = gis_col
        elif header_value == 'P3_Count':
            p3_count_col_index = gis_col
        elif header_value == 'P2_Count':
            p2_count_col_index = gis_col
        elif header_value == 'P1_Count':
            p1_count_col_index = gis_col

    # Check if all columns were found
    if None in [inspector_col_index, p3_count_col_index, p2_count_col_index, p1_count_col_index]:
        raise ValueError("Some required columns were not found in the GIS sheet.")

    # Now, you can use these indices in your loop:
    for gis_row in range(2, gis_ws.max_row + 1):  # Start from the fifth gis_row to skip header
        inspector = gis_ws.cell(row=gis_row, column=inspector_col_index).value
        p3_count = gis_ws.cell(row=gis_row, column=p3_count_col_index).value
        p3_count = 0 if pd.isna(p3_count) else 1
        p2_count = gis_ws.cell(row=gis_row, column=p2_count_col_index).value
        p2_count = 0 if pd.isna(p2_count) else 1
        p1_count = gis_ws.cell(row=gis_row, column=p1_count_col_index).value
        p1_count = 0 if pd.isna(p1_count) else 1

        if inspector not in inspector_data:
            inspector_data[inspector] = {'Total Inspections': 0, 'P3_Count': 0, 'P2_Count': 0, 'P1_Count': 0}

        inspector_data[inspector]['Total Inspections'] += 1
        inspector_data[inspector]['P3_Count'] += p3_count
        inspector_data[inspector]['P2_Count'] += p2_count
        inspector_data[inspector]['P1_Count'] += p1_count

    # Sort the inspector_data dictionary by the keys (i.e., the inspector names)
    sorted_inspectors = sorted(inspector_data.items(), key=lambda x: str(x[0]))

    # Populate the 'Inspector' tab of the template
    current_row = 5  # Starting at gis_row 5
    for inspector, data in sorted_inspectors:
        inspector_ws.cell(row=current_row, column=3).value = inspector
        inspector_ws.cell(row=current_row, column=4).value = data['Total Inspections']
        inspector_ws.cell(row=current_row, column=5).value = data['P3_Count']
        inspector_ws.cell(row=current_row, column=6).value = data['P2_Count']
        inspector_ws.cell(row=current_row, column=7).value = data['P1_Count']
        total_count = data['P3_Count'] + data['P2_Count'] + data['P1_Count']
        if total_count != 0:
            inspector_ws.cell(row=current_row, column=8).value = data['P3_Count'] / data['Total Inspections']
            inspector_ws.cell(row=current_row, column=9).value = data['P2_Count'] / data['Total Inspections']
            inspector_ws.cell(row=current_row, column=10).value = data['P1_Count'] / data['Total Inspections']
        else:
            inspector_ws.cell(row=current_row, column=8).value = 0
            inspector_ws.cell(row=current_row, column=9).value = 0
            inspector_ws.cell(row=current_row, column=10).value = 0

        if data['Total Inspections'] != 0:
            inspector_ws.cell(row=current_row, column=11).value = total_count / data['Total Inspections']
        else:
            inspector_ws.cell(row=current_row, column=11).value = 0

        for gis_col in [8, 9, 10, 11]:
            cell = inspector_ws.cell(row=current_row, column=gis_col)
            cell.number_format = '0%'

        current_row += 1

    for gis_col in [8, 9, 10, 11]:
        cell = inspector_ws.cell(row=current_row, column=gis_col)
        cell.number_format = '0%'

    # Generate output file name and save
    if start_date_obj == end_date_obj:
        output_file_name = f"Data_Conformance_Distribution_{start_date}.xlsx"
    else:
        output_file_name = f"Data_Conformance_Distribution_{start_date}_to_{end_date}.xlsx"

    # Get the directory path of the template file
    template_directory = os.path.dirname(template_file)

    # Join the directory path with the output file name
    output_file = os.path.join(template_directory, output_file_name)

    # Save the updated template to the output file
    template_wb.save(output_file)

    return teams_worked_today, output_file


def data_conformance_trans(GIS_file, EOD_file, TUC_file, template_file, start_date, end_date=None):
    # Parse the user input to a datetime.date object
    start_date_obj = datetime.strptime(start_date, "%Y%m%d").date()
    if end_date:
        end_date_obj = datetime.strptime(end_date, "%Y%m%d").date()
    else:
        end_date_obj = start_date_obj

    # Read the GIS file using pandas
    print_to_widget("Opening GIS Extract file...")
    GIS_df = pd.read_excel(GIS_file)

    # Read the EOD file using pandas
    print_to_widget("Opening End-Of-Day file...")
    EOD_df = pd.read_excel(EOD_file)

    # Read the DUC file using openpyxl
    print_to_widget("Opening Transmission Upload Check file...")
    TUC_wb = load_workbook(TUC_file)

    # Process EOD Extract
    print_to_widget("Processing End-Of-Day Extract")
    # Convert the "Flight Date" column to datetime objects
    EOD_df["Date"] = pd.to_datetime(EOD_df["Date"]).dt.date
    # Filter the DataFrame to get rows with the entered date and create a new DataFrame
    EOD_date_filtered_df = EOD_df[(EOD_df["Date"] >= start_date_obj) &
                                  (EOD_df["Date"] <= end_date_obj)].copy()
    # Convert 'Team Number' and 'Number of structures' columns to integers
    EOD_date_filtered_df['Team #'] = EOD_date_filtered_df['Team #'].astype(int)
    EOD_date_filtered_df['Number of structures'] = EOD_date_filtered_df['Number of structures'].astype(int)
    # Create a dictionary to store team numbers and their "Total Structure Count" values from EOD_file
    # Step 1: Sort by date in descending order
    EOD_date_filtered_df.sort_values(by=["Date"], ascending=False, inplace=True)
    # Step 2: Drop duplicates for same day same team, retaining the latest one
    EOD_date_filtered_df.drop_duplicates(subset=["Date", 'Team #'], keep="first",
                                         inplace=True)
    # Step 3: Group by team and sum the counts
    grouped_counts = EOD_date_filtered_df.groupby('Team #')['Number of structures'].sum()
    # Step 4: Convert the grouped data to a dictionary
    team_structure_count = grouped_counts.to_dict()

    # Process GIS Extract
    print_to_widget("Processing GIS Extract")
    # Convert the "Inspection_Date1" column to datetime objects and subtract 8 hours
    GIS_df["Inspection_Date1"] = pd.to_datetime(GIS_df["Inspection_Date1"]) - pd.Timedelta(hours=8)
    GIS_df["Inspection_Date2"] = pd.to_datetime(GIS_df["Inspection_Date2"]) - pd.Timedelta(hours=8)
    # Extract only the date part
    GIS_df["Inspection_Date1"] = GIS_df["Inspection_Date1"].dt.date
    GIS_df["Inspection_Date2"] = GIS_df["Inspection_Date2"].dt.date
    # Create a new column "Inspection_Date" that gives priority to "Inspection_Date2" over "Inspection_Date1"
    GIS_df["Inspection_Date"] = GIS_df["Inspection_Date2"].where(GIS_df["Inspection_Date2"].notna(),
                                                                 GIS_df["Inspection_Date1"])
    # Filter the DataFrame to get rows with the entered date
    date_filtered_df = GIS_df[(GIS_df["Inspection_Date"] >= start_date_obj) &
                              (GIS_df["Inspection_Date"] <= end_date_obj)]
    # Get the "Team_Number" column values as a list
    team_numbers = date_filtered_df["Team_Number"].tolist()
    # Count occurrences of each team number using Counter from collections
    team_number_count = Counter(team_numbers)
    # Count occurrences of each team number inspecting "EZ_POLE" on the specific date
    ez_pole_filtered_df = date_filtered_df[date_filtered_df["Eq_ObjType"] == "EZ_POLE"]
    ez_pole_team_counts = ez_pole_filtered_df["Team_Number"].value_counts().to_dict()

    # Process TUC file
    print_to_widget("Processing Transmission Upload Check...")
    # Initialize a dictionary to store the team numbers and their number of occurrences in the Upload Check file
    team_flight_date_count = {}
    # Iterate through each sheet in the DUC file
    for sheet_name in TUC_wb.sheetnames:
        try:
            # Extract the team number from the sheet name
            team_number = int(sheet_name[:3])
            print_to_widget(f"Processing team {team_number}'s upload check...")
        except ValueError:
            # Skip this sheet if it doesn't have the expected naming pattern
            continue

        # Get the sheet object
        sheet = TUC_wb[sheet_name]

        # Count the occurrences of the input date in the "Flight Date" column
        date_count = 0
        for sheet_row in range(6, sheet.max_row + 1):
            flight_date = sheet.cell(row=sheet_row, column=7).value
            if flight_date:
                # Parse the flight_date to a datetime object
                if isinstance(flight_date, str):
                    flight_date = parse(flight_date)
                # Check if flight_date is within the provided range
                if start_date_obj <= flight_date.date() <= end_date_obj:
                    date_count += 1

        # Add the team number and their corresponding number of occurrences to the dictionary
        team_flight_date_count[team_number] = date_count

    # Load the template using openpyxl
    print_to_widget("\nTransferring data to Data Conformance template...")
    template_wb = load_workbook(template_file)
    template_ws = template_wb['Pilot']

    # Write the Date or Date Range on E3
    if start_date_obj == end_date_obj:
        formatted_date = "Date: " + start_date_obj.strftime('%m/%d/%Y')
    else:
        formatted_date = "Date: " + start_date_obj.strftime('%m/%d/%Y') + " to " + end_date_obj.strftime(
            '%m/%d/%Y')

    template_ws['E3'] = formatted_date

    # Initialize sums and list
    eod_sum = 0
    gis_sum = 0
    duc_sum = 0
    ez_pole_sum = 0
    team_count = 0
    teams_worked_today: List[int] = []

    # Update the template with the number of structures inspected by each team
    for template_row in range(5, template_ws.max_row - 1):
        team_number = template_ws.cell(row=template_row, column=4).value

        if team_number in team_number_count:
            template_ws.cell(row=template_row, column=6).value = team_number_count[team_number]
            print_to_widget(f"\nGIS: Team {team_number} inspected {team_number_count[team_number]} structures")
        else:
            template_ws.cell(row=template_row, column=6).value = 0
            print_to_widget(f"\nGIS: Team {team_number} inspected 0 structures")

        if team_number in team_structure_count:
            template_ws.cell(row=template_row, column=5).value = team_structure_count[team_number]
            print_to_widget(f"EOD: Team {team_number} inspected {team_structure_count[team_number]} structures")
        else:
            template_ws.cell(row=template_row, column=5).value = 0
            print_to_widget(f"EOD: Team {team_number} inspected 0 structures")

        if team_number in team_flight_date_count:
            template_ws.cell(row=template_row, column=7).value = team_flight_date_count[team_number]
            print_to_widget(f"DUC: Team {team_number} inspected {team_flight_date_count[team_number]} structures")
        else:
            template_ws.cell(row=template_row, column=7).value = 0
            print_to_widget(f"DUC: Team {team_number} inspected 0 structures")

        # Update the template with the number of "EZ_POLE" inspected by each team
        if team_number in ez_pole_team_counts:
            template_ws.cell(row=template_row, column=10).value = ez_pole_team_counts[team_number]
            print_to_widget(f"GIS: Team {team_number} inspected {ez_pole_team_counts[team_number]} EZ Poles")
        else:
            template_ws.cell(row=template_row, column=10).value = 0
            print_to_widget(f"GIS: Team {team_number} inspected 0 EZ Poles")

        # Get the values in columns 5, 6, and 7
        eod_value = template_ws.cell(row=template_row, column=5).value
        gis_value = template_ws.cell(row=template_row, column=6).value
        duc_value = template_ws.cell(row=template_row, column=7).value

        # Compare the values
        if eod_value == gis_value == duc_value != 0:
            # Create a green fill and font for consistent values
            green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            green_font = Font(color='006100')
            # Format cells if values are consistent
            format_cell(template_ws.cell(row=template_row, column=5), eod_value, green_fill, green_font)
            format_cell(template_ws.cell(row=template_row, column=6), gis_value, green_fill, green_font)
            format_cell(template_ws.cell(row=template_row, column=7), duc_value, green_fill, green_font)
            template_ws.cell(row=template_row, column=22).value = 1
        elif eod_value != gis_value or gis_value != duc_value:
            # Create a red fill and font for inconsistent values
            red_fill = PatternFill(start_color='FFFFC7CE', end_color='FFFFC7CE', fill_type='solid')
            red_font = Font(color='9C0006')
            # Format cells if values are inconsistent
            format_cell(template_ws.cell(row=template_row, column=5), eod_value, red_fill, red_font)
            format_cell(template_ws.cell(row=template_row, column=6), gis_value, red_fill, red_font)
            format_cell(template_ws.cell(row=template_row, column=7), duc_value, red_fill, red_font)
            template_ws.cell(row=template_row, column=22).value = 0

        # Calculate the sum of each column
        eod_sum += template_ws.cell(row=row, column=5).value
        gis_sum += template_ws.cell(row=row, column=6).value
        duc_sum += template_ws.cell(row=row, column=7).value
        ez_pole_sum += template_ws.cell(row=row, column=10).value

        # Count the non-zero EOD, GIS, and DUC values
        if (template_ws.cell(row=row, column=5).value != 0) or \
                (template_ws.cell(row=row, column=6).value != 0) or \
                (template_ws.cell(row=row, column=7).value != 0):
            team_count += 1
            team_number = template_ws.cell(row=row, column=4).value
            if team_number not in teams_worked_today:  # To ensure no duplicate entries
                teams_worked_today.append(team_number)

    # Calculate the average structures per team per day
    num_days = (end_date_obj - start_date_obj).days + 1  # +1 to include both start and end dates

    # Ensure that team_count and num_days are not zero
    if team_count > 0 and num_days > 0:
        duc_avg = round(duc_sum / team_count / num_days)
    else:
        # Handle the case where division by zero would occur
        duc_avg = 0

    # Get the row number of the "TOTAL" row, assuming it's the second-to-last row
    total_row = template_ws.max_row - 1

    # Set the number of teams that worked (non-zero EOD values) in the "TOTAL" row
    template_ws.cell(row=total_row, column=4).value = f"{str(team_count)} Crews"
    template_ws.cell(row=total_row, column=4).font = Font(bold=True)

    # Set the average number of structures inspected per team in the "Average" row
    template_ws.cell(row=total_row + 1, column=4).value = f"{str(duc_avg)} Structures"
    template_ws.cell(row=total_row + 1, column=4).font = Font(bold=True)

    # Set the sum of each column in the "TOTAL" row
    template_ws.cell(row=total_row, column=5).value = eod_sum
    template_ws.cell(row=total_row, column=5).font = Font(bold=True)
    template_ws.cell(row=total_row, column=6).value = gis_sum
    template_ws.cell(row=total_row, column=6).font = Font(bold=True)
    template_ws.cell(row=total_row, column=7).value = duc_sum
    template_ws.cell(row=total_row, column=7).font = Font(bold=True)

    # Set the sum in the "TOTAL" row for the "EZ Poles" column
    template_ws.cell(row=total_row, column=10).value = ez_pole_sum
    template_ws.cell(row=total_row, column=10).font = Font(bold=True)

    # Check for multiple entries
    # Group by Date and Team, then count number of rows for each group
    grouped = EOD_date_filtered_df.groupby(["Date", 'Team #']).size().reset_index(
        name='counts')

    # Filter for groups with counts greater than 1, indicating duplicate entries on the same date
    duplicates = grouped[grouped['counts'] > 1]

    if not duplicates.empty:
        # Get the unique teams with duplicates on the same date
        duplicate_teams = duplicates['Team #'].unique()

        print_to_widget(f"\nWarning: The following team(s) have multiple EOD entries on the same date:"
                        f"\n{duplicate_teams}\nOnly the latest entry is recorded.\n")

    # Generate output file name and save
    if start_date_obj == end_date_obj:
        output_file_name = f"Data_Conformance_Transmission_{start_date}.xlsx"
    else:
        output_file_name = f"Data_Conformance_Transmission_{start_date}_to_{end_date}.xlsx"

    # Get the directory path of the template file
    template_directory = os.path.dirname(template_file)

    # Join the directory path with the output file name
    output_file = os.path.join(template_directory, output_file_name)

    # Save the updated template to the output file
    template_wb.save(output_file)

    return teams_worked_today, output_file


def daily_delta(traveler_file, template_file, selected_date):
    # Read current_date from the template (from the "Pilot" sheet)
    print_to_widget("\nAdding data to QA column in \"Pilot\" sheet...")

    # Read traveler sheet ("2024-Distribution-ALL")
    print_to_widget("\nReading traveler sheet...")
    try:
        traveler_df = pd.read_excel(traveler_file, sheet_name="2024-Distribution-ALL")
    except ValueError:
        traveler_df = pd.read_excel(traveler_file, sheet_name="C2-Distribution")

    # Filter by current_date
    print_to_widget("Filtering the traveler sheet by selected date...")
    traveler_df['FLIGHT_DATE'] = pd.to_datetime(traveler_df['FLIGHT_DATE'])
    filtered_df = traveler_df[traveler_df['FLIGHT_DATE'] == pd.Timestamp(selected_date)]

    # Create a copy of the filtered DataFrame to avoid SettingWithCopyWarning
    filtered_df_copy = filtered_df.copy()

    # Modify team names in the copied DataFrame and count occurrences
    print_to_widget("\nCounting the number of processed structures by each team...")
    filtered_df_copy['Flight Team'] = filtered_df_copy['Flight Team'].apply(
        lambda x: str(x).replace("Team ", ""))
    team_counts = filtered_df_copy['Flight Team'].value_counts()

    output_str = "Number of structures per team:\n"
    for team, count in team_counts.items():
        output_str += f"Team {team}: {count}\n"

    print_to_widget(output_str)

    # Open the template file with openpyxl for editing
    template_wb = load_workbook(template_file, data_only=True)
    pilot_sheet = template_wb["Pilot"]

    # Iterate through the rows and update 'QA' column
    for pilot_sheet_row in range(5, pilot_sheet.max_row - 1):  # Start from row 4
        team_cell_value = pilot_sheet[f'D{pilot_sheet_row}'].value  # Assuming 'Team' is in column D
        if team_cell_value is not None:
            team_cell_value = str(team_cell_value)
            count = team_counts.get(team_cell_value)
            if count is not None:
                pilot_sheet[f'H{pilot_sheet_row}'] = count  # Update 'QA' column with the count
            else:
                pilot_sheet[f'H{pilot_sheet_row}'] = 0  # Update 'QA' column with 0

    # Define styles for highlighting
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    green_font = Font(color='006100')
    red_fill = PatternFill(start_color='FFFFC7CE', end_color='FFFFC7CE', fill_type='solid')
    red_font = Font(color='9C0006')
    bold_italic_font = Font(bold=True, italic=True)

    print_to_widget("\nApplying highlighting effects to inconsistencies...")
    for pilot_sheet_row in range(5, pilot_sheet.max_row - 1):  # headers on row 4, start from row 5
        eod = pilot_sheet[f'E{pilot_sheet_row}'].value
        gis = pilot_sheet[f'F{pilot_sheet_row}'].value
        upload = pilot_sheet[f'G{pilot_sheet_row}'].value
        qa = pilot_sheet[f'H{pilot_sheet_row}'].value

        # Check if all values match and are not all zero
        if eod == gis == upload == qa and eod != 0:
            fill = green_fill
            font = green_font
        elif eod == 0 and gis == 0 and upload == 0 and qa == 0:
            fill = None  # No fill for all zeros
            font = None
        else:
            fill = red_fill
            font = red_font

        # Apply the fill to the cells
        if fill and font:
            for pilot_sheet_col in ['E', 'F', 'G', 'H']:
                pilot_sheet[f'{pilot_sheet_col}{pilot_sheet_row}'].fill = fill
                pilot_sheet[f'{pilot_sheet_col}{pilot_sheet_row}'].font = font

    # Process each sheet in the template
    print_to_widget("\nAdding QA column to each team sheet...")
    for sheet_name in template_wb.sheetnames:
        # Check if the sheet name follows the format "<number> - <name>"
        if " - " in sheet_name:
            team_number = sheet_name.split(" - ")[0]  # Extract team number
            team_sheet = template_wb[sheet_name]

            # Add a new column for Structure IDs if not already present
            if team_sheet.max_column < 3 or team_sheet.cell(row=1, column=3).value != "QA":
                team_sheet.cell(row=1, column=3).value = "QA"

            # Find matching structures for this team
            matching_structures = filtered_df[filtered_df_copy['Flight Team'] == team_number]['SCE_STRUCT']

            # Sort the structure IDs
            sorted_structures = sorted(matching_structures)

            # Record the sorted structure IDs in the team sheet
            for idx, structure_id in enumerate(sorted_structures, start=2):  # Start from row 2
                team_sheet.cell(row=idx, column=3).value = structure_id

            # Clear any remaining old data
            for idx in range(len(sorted_structures) + 2, team_sheet.max_row + 1):
                team_sheet.cell(row=idx, column=3).value = None

            # Define column indices for "Folder Name", "GIS ID", and "QA" in template team sheets
            folder_name_col = 'A'  # Example column index for "Folder Name"
            gis_id_col = 'B'  # Example column index for "GIS ID"
            qa_col = 'C'  # 'QA' column

            # Collect values from each column
            folder_names = [team_sheet[f'{folder_name_col}{folder_name_row}'].value for folder_name_row
                            in range(2, team_sheet.max_row + 1)]
            gis_ids = [team_sheet[f'{gis_id_col}{gid_id_row}'].value for gid_id_row in range(2, team_sheet.max_row + 1)]
            qas = [team_sheet[f'{qa_col}{gis_id_row}'].value for gis_id_row in range(2, team_sheet.max_row + 1)]

            # Create a mapping of structure IDs to their indices in filtered_df_copy
            structure_id_to_index = {sid_to_idx_row['SCE_STRUCT']: idx for idx, sid_to_idx_row
                                     in filtered_df_copy.iterrows()
                                     if sid_to_idx_row['Flight Team'] == team_number}

            # Iterate through the rows to apply highlighting
            for pilot_sheet_row in range(2, team_sheet.max_row + 1):
                structure_id = team_sheet[f'{qa_col}{pilot_sheet_row}'].value

                for pilot_sheet_col, val, other_vals1, other_vals2 in zip(
                        [folder_name_col, gis_id_col, qa_col],
                        [folder_names[pilot_sheet_row - 2], gis_ids[pilot_sheet_row - 2], qas[pilot_sheet_row - 2]],
                        [gis_ids, qas, folder_names],
                        [qas, folder_names, gis_ids]):
                    # Check if the value is not found in the other columns across all rows
                    if val is not None and (val not in other_vals1 or val not in other_vals2):
                        team_sheet[f'{pilot_sheet_col}{pilot_sheet_row}'].fill = red_fill
                        team_sheet[f'{pilot_sheet_col}{pilot_sheet_row}'].font = red_font

                # Check "EZ Pole" column in the traveler DataFrame and apply bold & italic to QA values
                if structure_id in structure_id_to_index:
                    ez_pole_index = structure_id_to_index[structure_id]
                    ez_pole_val = filtered_df_copy.at[ez_pole_index, 'EZ Pole']
                    if ez_pole_val == "EZ_POLE":
                        team_sheet[f'{qa_col}{pilot_sheet_row}'].font = bold_italic_font

            # Set the width for all specified columns to 16.29
            desired_width = 16.29
            team_sheet.column_dimensions[folder_name_col].width = desired_width
            team_sheet.column_dimensions[gis_id_col].width = desired_width
            team_sheet.column_dimensions[qa_col].width = desired_width

            # Define the table range. For example, if your table starts at row 2 and ends at row 20
            table_range = f"A1:C{team_sheet.max_row}"

            # Create the table
            table = Table(displayName=f"Team{team_number}Table", ref=table_range)

            # Add a default style with striped rows and banded columns
            style = TableStyleInfo(name="TableStyleLight15", showFirstColumn=False,
                                   showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            table.tableStyleInfo = style

            # Add the table to the worksheet
            team_sheet.add_table(table)

    # Format the selected_date into 'YYYYMMDD' format
    selected_date_obj = datetime.strptime(selected_date, '%Y%m%d')
    formatted_selected_date = selected_date_obj.strftime('%Y%m%d')

    # Extract the directory path from the template file
    directory_path = os.path.dirname(template_file)

    # Initialize the suffix and the filename
    suffix = 0
    new_filename = os.path.join(directory_path, f"Daily_Delta_{formatted_selected_date}.xlsx")

    # Check if the file exists and update the filename with a suffix if necessary
    while os.path.exists(new_filename):
        suffix += 1
        new_filename = os.path.join(directory_path, f"Daily_Delta_{formatted_selected_date} ({suffix}).xlsx")

    # Save the changes to the new file
    template_wb.save(new_filename)
    print_to_widget(f"\nDaily delta report saved as {new_filename}.")


def combine_extracts_distro(excel_file_1, excel_file_2):
    # Define a dictionary that maps the old column names to the new ones
    rename_dict = {
        "District_1": "District_Number",
        "District_N": "District_Name",
        "GlobalID": "REF_GlobalID",
        "GlobalID_2": "GlobalID",
        "TEAM NUMBER": "Team_Number",
        "PILOT": "Pilot",
        "INSPECTOR": "Inspector",
        "P1 COUNT": "P1_Count",
        "P2_COUNT": "P2_Count",
        "P3 COUNT": "P3_Count",
        "P1 NOTES": "P1_Notes",
        "P2 NOTES": "P2_Notes",
        "P3 NOTES": "P3_Notes",
        "VENDOR CATEGORY": "Vendor_Category",
        "VENDOR STATUS": "Vendor_Status",
        "INSPECTION DATE": "Inspection_Date1",
        "CREW NOTES": "Vendor_Notes",
        "x": "x2",
        "y": "y2"
    }

    # Read the Excel files into pandas DataFrames
    print_to_widget("Loading the extracts...")
    df1 = pd.read_excel(excel_file_1)
    df2 = pd.read_excel(excel_file_2)

    # Rename the columns in the second dataframe
    print_to_widget("Editing, adding, and reordering headers to match...")
    df2.rename(columns=rename_dict, inplace=True)

    # Add any missing columns to df2
    missing_columns = set(df1.columns) - set(df2.columns)
    for column in missing_columns:
        df2[column] = None

    # Reorder the columns in df2 to match df1
    df2 = df2[df1.columns]

    # Append the data from the second file to the first file
    print_to_widget("Combining the extracts...")
    df1 = df1.reset_index(drop=True)
    df2 = df2.reset_index(drop=True)
    combined_df = pd.concat([df1, df2], ignore_index=True)

    def fill_missing_values_distro(extract_row):
        if pd.isnull(extract_row['Work_Order']) and pd.notnull(extract_row['WO_1']):
            extract_row['Work_Order'] = extract_row['WO_1']
        elif pd.notnull(extract_row['Work_Order']) and pd.isnull(extract_row['WO_1']):
            extract_row['WO_1'] = extract_row['Work_Order']
        return extract_row

    # After concatenating the dataframes
    print_to_widget("Filling out missing values, if possible...")
    combined_df = combined_df.apply(fill_missing_values_distro, axis=1)
    combined_df["Mapped_Lat"] = combined_df["y2"]
    combined_df["Mapped_Lon"] = combined_df["x2"]

    # Check for empty 'Mapped_Lat' or 'Mapped_Lon' cells
    for mapped_coords_idx, mapped_coords_row in combined_df.iterrows():
        if pd.isnull(mapped_coords_row['Mapped_Lat']) or pd.isnull(mapped_coords_row['Mapped_Lon']):
            print_to_widget(f"'Mapped_Lat' or 'Mapped_Lon' is empty at index {mapped_coords_idx}")

    # Write the combined data to a new Excel file
    output_dir = os.path.dirname(excel_file_1)
    today = date.today().strftime('%Y%m%d')  # format date as YYYY_MM_DD
    output_name = f"ADI_Master_Structure_Combined_{today}.xlsx"
    output_file = os.path.join(output_dir, output_name)

    # Initialize the counter
    counter = 1

    # Check if the file already exists
    while os.path.exists(output_file):
        # Append the counter to the base file name
        modified_file_name = f"{output_name.split('.')[0]} ({counter}).{output_name.split('.')[-1]}"
        output_file = os.path.join(output_dir, modified_file_name)
        counter += 1

    combined_df.to_excel(output_file, index=False)

    # Open the workbook to make dataframe a table
    wb = load_workbook(output_file)
    ws = wb.active

    # Create a data table
    print_to_widget("Creating a data table...")
    tab = Table(displayName="Table1", ref=ws.dimensions)

    # Add a default table style
    table_style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                                 showLastColumn=False, showRowStripes=True)
    tab.tableStyleInfo = table_style

    # Add the table to the sheet
    ws.add_table(tab)

    # Save the workbook with the styling
    wb.save(output_file)
    print_to_widget(f"Extracts combined successfully! Output written to {os.path.basename(output_file)}")


def combine_extracts_trans(excel_file_1, excel_file_2):
    # Define a dictionary that maps the old column names to the new ones
    rename_dict = {
        "District_N": "District_Number",
        "GlobalID": "GlobalID_2",
        "TEAM NUMBER": "Team_Number",
        "PILOT": "Pilot",
        "INSPECTOR": "Inspector",
        "P1 COUNT": "P1_Count",
        "VENDOR CATEGORY": "Vendor_Category",
        "VENDOR STATUS": "Vendor_Status",
        "INSPECTION_DATE": "Inspection_Date1",
        "CREW NOTES": "Crew_Notes",
        "x": "x2",
        "y": "y2"
    }

    # Read the Excel files into pandas DataFrames
    print_to_widget("Loading the extracts...")
    df1 = pd.read_excel(excel_file_1)
    df2 = pd.read_excel(excel_file_2)

    # Rename the columns in the second dataframe
    print_to_widget("Editing, adding, and reordering headers to match...")
    df2.rename(columns=rename_dict, inplace=True)

    # Add any missing columns to df2
    missing_columns2 = set(df1.columns) - set(df2.columns)
    for column in missing_columns2:
        df2[column] = None

    # Add any missing columns to df1
    missing_columns1 = set(df2.columns) - set(df1.columns)
    for column in missing_columns1:
        df1[column] = None

    # Reorder the columns in df2 to match df1
    df2 = df2[df1.columns]

    # Append the data from the second file to the first file
    print_to_widget("Combining the extracts...")
    df1 = df1.reset_index(drop=True)
    df2 = df2.reset_index(drop=True)
    combined_df = pd.concat([df1, df2], ignore_index=True)

    def fill_missing_values_trans(extract_row):
        if pd.isnull(extract_row['FLOC']) and pd.isnull(extract_row['StructureN']):
            print_to_widget(f"Both values are missing at index {extract_row.name}")
        elif pd.isnull(extract_row['FLOC']) and pd.notnull(extract_row['StructureN']):
            extract_row['FLOC'] = 'OH-' + extract_row['StructureN']
        elif pd.notnull(extract_row['FLOC']) and pd.isnull(extract_row['StructureN']):
            extract_row['StructureN'] = extract_row['FLOC'].replace('OH-', '')
        return extract_row

    # After concatenating the dataframes
    print_to_widget("Filling out missing values, if possible...")
    combined_df = combined_df.apply(fill_missing_values_trans, axis=1)
    combined_df["Mapped_Lat"] = combined_df["y2"]
    combined_df["Mapped_Lon"] = combined_df["x2"]

    # Check for empty 'Mapped_Lat' or 'Mapped_Lon' cells
    for mapped_coords_idx, mapped_coords_row in combined_df.iterrows():
        if pd.isnull(mapped_coords_row['Mapped_Lat']) or pd.isnull(mapped_coords_row['Mapped_Lon']):
            print_to_widget(f"'Mapped_Lat' or 'Mapped_Lon' is empty at index {mapped_coords_idx}")

    # Write the combined data to a new Excel file
    output_dir = os.path.dirname(excel_file_1)
    today = date.today().strftime('%Y%m%d')  # format date as YYYY_MM_DD
    output_name = f"ATI_Master_Structure_Combined_{today}.xlsx"
    output_file = os.path.join(output_dir, output_name)

    # Initialize the counter
    counter = 1

    # Check if the file already exists
    while os.path.exists(output_file):
        # Append the counter to the base file name
        modified_file_name = f"{output_name.split('.')[0]} ({counter}).{output_name.split('.')[-1]}"
        output_file = os.path.join(output_dir, modified_file_name)
        counter += 1

    combined_df.to_excel(output_file, index=False)

    # Open the workbook to make dataframe a table
    wb = load_workbook(output_file)
    ws = wb.active

    # Create a data table
    print_to_widget("Creating a data table...")
    tab = Table(displayName="Table1", ref=ws.dimensions)

    # Add a default table style
    table_style = TableStyleInfo(name="TableStyleMedium3", showFirstColumn=False,
                                 showLastColumn=False, showRowStripes=True)
    tab.tableStyleInfo = table_style

    # Add the table to the sheet
    ws.add_table(tab)

    # Save the workbook with the styling
    wb.save(output_file)
    print_to_widget(f"Extracts combined successfully! Output written to {os.path.basename(output_file)}")


def issue_form():
    webbrowser.open('https://forms.office.com/r/tiFRWQcpX0')


def filter_helo_traveler(file_path, dir_path):
    # Delete hidden Mac files
    print_to_widget(f"\nDeleting hidden Mac files...")
    found_mac_files = False
    for dirpath, dirnames, filenames in os.walk(dir_path):
        for filename in filenames:
            if filename.startswith("._"):
                filepath = os.path.join(dirpath, filename)
                os.remove(filepath)
                print_to_widget(f"{filename} deleted.")
                found_mac_files = True
    if not found_mac_files:
        print_to_widget("No hidden Mac files found in the folder.")
    else:
        print_to_widget(f"All hidden Mac files have been deleted.")

    structure_names = []
    structure_paths = []
    structure_dict = {}

    print_to_widget("\nExtracting folder names and paths...")
    for helo_root, dirs, files in os.walk(dir_path):
        # Check if any of the files in the directory are image files
        if any(file.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp')) for file in files):
            structure_name = os.path.basename(helo_root)  # Name of the current helo_root
            structure_path = helo_root  # Full directory path

            # Add to the lists
            structure_names.append(structure_name)
            structure_paths.append(structure_path)

            # Add to the dictionary
            structure_dict[structure_name] = structure_path

    # Read the Excel file into a pandas dataframe
    print_to_widget(f"\nCounting images and extracting GPS data for each image...")
    df = pd.read_excel(file_path)

    # Check if the Structure_ID column is present in the dataframe
    if 'FLOC' not in df.columns:
        print_to_widget(
            "Error: \"FLOC\" column not found in Excel file. Make sure the correct file is "
            "selected. Rerun the script...")
        sys.exit(1)
    else:
        # Make a copy of the dataframe before filtering
        df_copy = df.copy()

        # Replace "OH-" with an empty string in the "FLOC" column
        df_copy["FLOC"] = df_copy["FLOC"].str.replace("OH-", "")

        # Create a dictionary to store the photo count for each folder
        photo_count = {}

        # Create a dictionary to store the latitude and longitude of image for each folder
        latlong_date_N = {}

        # Loop through each folder in the directory
        for folder in structure_paths:
            folder_name = os.path.basename(folder)  # Extract the name of the folder from the path
            # Get the list of files in the folder
            files = os.listdir(folder)
            # Count the number of files in the folder
            count = len(files)
            # Add the photo count to the dictionary
            photo_count[folder_name] = count
            # Initialize the lat and long variables to N/A.
            lat = 'N/A'
            lng = 'N/A'
            date_taken = 'N/A'
            # Initialize variable to track whether an "N" image was found
            n_found = False
            # Loop through each file in the folder
            print_to_widget(f"{folder_name} has {count} images.")
            for file in files:
                # Check if the file ends with "N" and is an image file
                if file.lower().endswith(("n.jpg", "n.jpeg", "n.png", "n.bmp", "n.JPG")):
                    # Get the full path of the image file
                    image_path = os.path.join(folder, file)
                    # Extract the date taken from the image file
                    date_taken = get_date_taken(image_path)
                    # Extract the GPS coordinates from the image file
                    coords = get_gps_from_image(image_path)
                    # Update the lat/long variables with the GPS coordinates, or blank if no GPS coordinates were found
                    if coords:
                        lat = coords[0]
                        lng = coords[1]
                    else:
                        lat = 'Metadata Issue'
                        lng = 'Metadata Issue'
                        print_to_widget(f"Warning: {folder_name}'s nadir has no GPS coordinates.\n")
                    # Set variable to True if an "N" image is found
                    n_found = True
            # Print a message if no "N" image was found
            if not n_found:
                print_to_widget(f"Warning: {folder_name} has no image that ends with \"N\".")
                lat = 'Nadir missing N'
                lng = 'Nadir missing N'

                # Loop again to find the first image with a date
                for file in files:
                    # Ensure we only consider image files (using the same extensions as before)
                    if file.lower().endswith((".jpg", ".jpeg", ".png", ".bmp", ".JPG")):
                        image_path = os.path.join(folder, file)
                        date_taken = get_date_taken(image_path)

                        # Check if a date was found
                        if date_taken not in ('N/A', '', None):
                            break

                else:
                    # No image with a date was found
                    print_to_widget(f"Warning: {folder_name} has no date taken metadata.\n")

            # Add the lat, long, and date taken to the dictionary
            latlong_date_N[folder_name] = (lat, lng, date_taken)

        # Find the closest match for each folder name in the dataframe
        # Dictionary to store matches
        matches = {}
        choices = list(df_copy['FLOC'])  # Convert to list for manipulation
        print_to_widget("\nVerifying if folder names exist in SCE Traveler (LOCAL COPY)...")
        for folder_name, folder_path in zip(structure_names, structure_paths):
            folder_name, closest_match, choices = find_closest_match(folder_name, folder_path, choices, df_copy)

            # Check if match already exists
            if closest_match in matches:
                if not isinstance(matches[closest_match], list):
                    matches[closest_match] = [matches[closest_match]]
                matches[closest_match].append(folder_name)
            else:
                matches[closest_match] = folder_name

        # Create a temporary duplicates dictionary for printing purposes
        duplicates = {key: value for key, value in matches.items() if isinstance(value, list)}

        if duplicates:
            output = "\nThe same match(es) found for the following folders:"
            for key, values in duplicates.items():
                output += f"\n{', '.join(values)} --> {key}"

            print_to_widget(f"{output}\nFinding new matches...")

            # Resolve duplicates
            available_options = choices
            matches, available_options = resolve_duplicates(matches, available_options, structure_dict, df_copy)

        # Filter the dataframe based on the closest matches
        closest_matches = list(matches.keys())
        filtered_df = df_copy[df_copy['FLOC'].isin(closest_matches)].copy()
        filtered_df = filtered_df.assign(
            closest_match=pd.Categorical(filtered_df['FLOC'], categories=closest_matches, ordered=True))
        filtered_df = filtered_df.sort_values('closest_match')
        filtered_df.drop('closest_match', axis=1, inplace=True)

        # Check for duplicates in the extract's 'FLOC' column
        duplicates = filtered_df.duplicated('FLOC', keep=False)
        columns_to_show = ['FLOC', 'Scope Type', 'Trans/Dist', 'FLOC Type', 'District/Region',
                           'Lat', 'Long']

        if any(duplicates):
            duplicate_values = filtered_df.loc[duplicates, 'FLOC'].unique()
            print_to_widget(
                f"Warning: The following Structure IDs are duplicated in the extract: {duplicate_values}."
                f" Please choose which row to keep.")
            for dup_value in duplicate_values:
                duplicate_rows = filtered_df[filtered_df['FLOC'] == dup_value]
                selected_index = display_duplicates_and_get_selection(duplicate_rows, columns_to_show)

                # Drop all rows except the selected one
                drop_indices = [idx for idx in duplicate_rows.index if idx != int(selected_index)]
                filtered_df.drop(drop_indices, inplace=True)

            print_to_widget("All duplicates have been effectively resolved.")

        print_to_widget("\nAdding new columns to the extract...")

        folder_names_update = list(matches.values())

        # Add a new column to the filtered dataframe indicating the folder name for each row's "FLOC" value
        filtered_df.insert(filtered_df.columns.get_loc('FLOC') + 1, 'FolderName', folder_names_update)

        # # Sort table in ascending order by Structure ID
        print_to_widget(f"\nSorting extract in ascending Structure ID order...")
        filtered_df.sort_values(by=["FLOC"], inplace=True)

        # Add a new column to the filtered dataframe indicating the photo count for each row's "FLOC" value
        filtered_df.insert(filtered_df.columns.get_loc('FLOC') + 2, 'PhotoCount',
                           filtered_df['FolderName'].apply(lambda x: photo_count[x]))

        # Add a new column to the filtered dataframe indicating the latitude for each row's "FLOC" value
        filtered_df.insert(filtered_df.columns.get_loc('Long') + 1, 'FieldLat',
                           filtered_df['FolderName'].apply(lambda x: latlong_date_N[x][0]))

        # Add a new column to the filtered dataframe indicating the longitude for each row's "FLOC" value
        filtered_df.insert(filtered_df.columns.get_loc('Long') + 2, 'FieldLong',
                           filtered_df['FolderName'].apply(lambda x: latlong_date_N[x][1]))

        # Add a new column to the filtered dataframe indicating the flight date for each row's "FLOC" value
        filtered_df['Date Captured'] = filtered_df['FolderName'].apply(
            lambda x: datetime.strptime(latlong_date_N[x][2], '%Y%m%d').date()
            if x in latlong_date_N and latlong_date_N[x][2] != 'N/A' else 'N/A')

        # Add a new column KnownAsset photo location for each row's "FLOC" value
        filtered_df.insert(filtered_df.columns.get_loc('PhotoCount') + 1, 'PhotoLoc',
                           filtered_df['Trans/Dist'].astype(str).str[0] + '23XX-XXXX_KnownAssets_' + pd.to_datetime(
                               filtered_df['Date Captured'], errors='coerce').dt.strftime('%Y%m%d').fillna(''))

        # Calculate distances and add them as a new column to the DataFrame
        print_to_widget("\nCalculating structure distance from GIS location...")
        distances = []
        for index, field_coords_row in filtered_df.iterrows():
            if not has_missing_values(field_coords_row):
                dist = distance_calculator((field_coords_row['Lat'], field_coords_row['Long']),
                                           (field_coords_row['FieldLat'], field_coords_row['FieldLong']))
                if dist is not None and not math.isnan(dist):
                    distances.append(round(dist, 2))
                else:
                    distances.append('N/A')
            else:
                distances.append('N/A')
        filtered_df.insert(filtered_df.columns.get_loc('FieldLong') + 1, 'DistFromGIS (ft)', distances)

        # Calculate the farthest image distance from the nadir as a new column to the DataFrame
        print_to_widget("\nCalculating farthest distance between images and the nadir...")
        nadir_distances = get_farthest_from_nadir(dir_path)
        filtered_df.insert(filtered_df.columns.get_loc('DistFromGIS (ft)') + 1, 'FarthestDistFromNadir (ft)',
                           filtered_df['FolderName'].apply(
                               lambda x: round(nadir_distances.get(x), 2) if nadir_distances.get(
                                   x) is not None else 'N/A'))

        # Extract the directory path from dir_path
        directory = os.path.dirname(file_path)

        # Extract the base name of the file without extension
        file_name = os.path.splitext(os.path.basename(file_path))[0]

        # Search for a date in the file name using regex
        match = re.search(r"_(\d{4}\d{2}\d{2})_|(\d{4}\d{2}\d{2})$", file_name)

        # If a date is found, parse it to a datetime object
        if match:
            file_date = match.group(0)

        else:
            # If no date is found, use the current date as a fallback
            file_date = datetime.now().strftime("%Y%m%d")

        # Now you can use file_date to name your new file
        new_file_name = f"Filtered_Helo_Traveler_{file_date}.xlsx"

        # Construct the new file path for saving
        new_file_path = os.path.join(directory, new_file_name)

        # Initialize the counter
        counter = 1

        # Check if the file already exists
        while os.path.exists(new_file_path):
            # Append the counter to the base file name
            modified_file_name = f"{new_file_name.split('.')[0]} ({counter}).{new_file_name.split('.')[-1]}"
            new_file_path = os.path.join(directory, modified_file_name)
            counter += 1

        # Write the filtered dataframe to the original Excel file
        filtered_df.to_excel(new_file_path, index=False, sheet_name='FilteredTraveler')

        # Open the Excel file with openpyxl and format the cells in the "Match" and "FLOC" columns that are False
        wb = load_workbook(new_file_path)
        ws = wb.active

        # Define a PatternFill object
        light_red_fill = PatternFill(start_color='FFFFC7CE', end_color='FFFFC7CE', fill_type='solid')
        dark_red_font = Font(color='9C0006')

        # Find the index of the 'Distance', 'FieldLat', and 'FieldLong' columns
        distance_column_index = filtered_df.columns.get_loc("DistFromGIS (ft)") + 1
        field_lat_column_index = filtered_df.columns.get_loc("FieldLat") + 1
        field_long_column_index = filtered_df.columns.get_loc("FieldLong") + 1
        nadir_max_distance_col = filtered_df.columns.get_loc('FarthestDistFromNadir (ft)') + 1
        photo_count_col_index = filtered_df.columns.get_loc('PhotoCount') + 1

        # Apply cell formatting to the appropriate cells based on the specified conditions
        print_to_widget("\nHighlighting potential issues...")
        for field_coords_row in range(2, ws.max_row + 1):
            distance_cell = ws.cell(row=field_coords_row, column=distance_column_index)
            field_lat_cell = ws.cell(row=field_coords_row, column=field_lat_column_index)
            field_long_cell = ws.cell(row=field_coords_row, column=field_long_column_index)
            farthest_cell = ws.cell(row=field_coords_row, column=nadir_max_distance_col)
            photo_count_cell = ws.cell(row=field_coords_row, column=photo_count_col_index)
            if distance_cell.value == 'N/A':
                format_cell(field_lat_cell, field_lat_cell.value, light_red_fill, dark_red_font)
                format_cell(field_long_cell, field_long_cell.value, light_red_fill, dark_red_font)
                format_cell(distance_cell, distance_cell.value, light_red_fill, dark_red_font)
            elif distance_cell.value > 200:
                format_cell(distance_cell, distance_cell.value, light_red_fill, dark_red_font)
            else:
                farthest_cell.alignment = Alignment(horizontal='right')
                distance_cell.alignment = Alignment(horizontal='right')
            if farthest_cell.value == 'N/A' or int(farthest_cell.value) > 300:
                format_cell(farthest_cell, farthest_cell.value, light_red_fill, dark_red_font)
            if photo_count_cell.value == 0:
                format_cell(photo_count_cell, photo_count_cell.value, light_red_fill, dark_red_font)

        # Highlight cells that are not a complete match in "FLOC" and "FolderName"
        for field_coords_row in ws.iter_rows(min_row=2, min_col=1, max_col=2):
            if field_coords_row[0].value != field_coords_row[1].value:
                field_coords_row[0].fill = light_red_fill
                field_coords_row[0].font = dark_red_font
                field_coords_row[1].fill = light_red_fill
                field_coords_row[1].font = dark_red_font
                field_coords_row[26].fill = light_red_fill
                field_coords_row[26].font = dark_red_font

        # Save the workbook
        wb.save(new_file_path)

    return new_file_path


# Create a queue
q = queue.Queue()


def filterextractdistro():
    def filterextractdistro_thread():
        # Ask for the Excel file
        print_to_widget(f"\nSelect Extract File.")
        file_path = filedialog.askopenfilename(title="Select the Extract file",
                                               filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        if file_path == "":
            print_to_widget("Operation canceled.")
            return
        print_to_widget(f"Selected File: {os.path.basename(file_path)}.")

        # Ask the user to choose the directory of folders
        print_to_widget(f"\nSelect Structure Folder Directory.")
        dir_path = filedialog.askdirectory(title="Select Structure Folder Directory")
        print_to_widget(f"Selected Folder Directory: {os.path.basename(dir_path)}.")

        # Check if the user has provided both the Excel file and the source directory
        if file_path and dir_path:
            try:
                new_file_path = filter_extract_distro(file_path, dir_path)
                output_file = get_distances_from_nadir(dir_path, new_file_path)
                final_file = append_issues_sheet(output_file)
                append_ez_aoc_sheet(final_file)
                move_folders_based_on_issues(final_file, dir_path)
                print_to_widget(f"Filtered extract saved as: {os.path.basename(output_file)}")
                # Display a message box to indicate the process is complete
                messagebox.showinfo(title='Filter Extract',
                                    message='Extract filtered and processed successfully. Please verify that the '
                                            'information is correct before proceeding to the next step.')
            except (FileNotFoundError, ValueError) as e:
                display_exception()
        else:
            print_to_widget("Operation canceled.\n")

    # Create a new thread for the filterextractdistro_thread() function
    filterextractdistrothread = threading.Thread(target=filterextractdistro_thread)
    filterextractdistrothread.start()


def filterextracttrans():
    def filterextracttrans_thread():
        # Ask for the Excel file
        print_to_widget(f"\nSelect Extract File.")
        file_path = filedialog.askopenfilename(title="Select the Extract file",
                                               filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        if file_path == "":
            print_to_widget("Operation canceled.")
            return
        print_to_widget(f"Selected File: {os.path.basename(file_path)}.")

        # Ask the user to choose the directory of folders
        print_to_widget(f"\nSelect Structure Folder Directory.")
        dir_path = filedialog.askdirectory(title="Select Structure Folder Directory")
        print_to_widget(f"Selected Folder Directory: {os.path.basename(dir_path)}.")

        # Check if the user has provided both the Excel file and the source directory
        if file_path and dir_path:
            try:
                new_file_path = filter_extract_trans(file_path, dir_path)
                output_file = get_distances_from_nadir(dir_path, new_file_path)
                final_file = append_issues_sheet(output_file)
                append_ez_aoc_sheet(final_file)
                print_to_widget(f"Filtered extract saved as: {os.path.basename(output_file)}")
                # Display a message box to indicate the process is complete
                messagebox.showinfo(title='Filter Extract',
                                    message='Extract filtered and processed successfully. Please verify that the '
                                            'information is correct before proceeding to the next step.')
            except (FileNotFoundError, ValueError) as e:
                display_exception()
        else:
            print_to_widget("Operation canceled.\n")

    filterextracttransthread = threading.Thread(target=filterextracttrans_thread)
    filterextracttransthread.start()


def renameimagesauto():
    def renameimagesauto_thread():
        # Ask the user to choose the directory of folders
        print_to_widget(f"\nSelect Structure Folder Directory.")
        dir_path = filedialog.askdirectory(title="Select Structure Folder Directory")
        print_to_widget(f"Selected Folder Directory: {os.path.basename(dir_path)}.")

        # Check if the user has provided both the Excel file and the source directory
        if dir_path:
            try:
                rename_images_auto(dir_path)
                # Display a message box to indicate the process is complete
                messagebox.showinfo(title='Rename Image', message='Image renaming process completed')
            except (FileNotFoundError, ValueError) as e:
                display_exception()
        else:
            print_to_widget("Operation canceled.\n")

    renameimagesautothread = threading.Thread(target=renameimagesauto_thread)
    renameimagesautothread.start()


def renameimagesmanual():
    def renameimagesmanual_thread():
        # Ask the user to choose the directory of folders
        print_to_widget(f"\nSelect Structure Folder Directory.")
        dir_path = filedialog.askdirectory(title="Select Structure Folder Directory")
        print_to_widget(f"Selected Folder Directory: {os.path.basename(dir_path)}.")

        # Check if the user has provided both the Excel file and the source directory
        if dir_path:
            try:
                rename_images_manual(dir_path)
                # Display a message box to indicate the process is complete
                messagebox.showinfo(title='Rename Image', message='Image renaming process completed successfully')
            except (FileNotFoundError, ValueError) as e:
                display_exception()
        else:
            print_to_widget("Operation canceled.\n")

    renameimagesmanualthread = threading.Thread(target=renameimagesmanual_thread)
    renameimagesmanualthread.start()


def filterandrenamedistro():
    def filterandrenamedistro_thread():
        # Ask for the Excel file
        print_to_widget(f"\nSelect the Extract File.")
        file_path = filedialog.askopenfilename(title="Select the Extract File",
                                               filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        if file_path == "":
            print_to_widget("Operation canceled.")
            return
        print_to_widget(f"Selected File: {os.path.basename(file_path)}.")

        # Ask the user to choose the directory of folders
        print_to_widget(f"\nSelect Structure Folder Directory.")
        dir_path = filedialog.askdirectory(title="Select Structure Folder Directory")
        print_to_widget(f"Selected Folder Directory: {os.path.basename(dir_path)}.")

        # Check if the user has provided both the Excel file and the source directory
        if file_path and dir_path:
            try:
                new_file_path = filter_extract_distro(file_path, dir_path)
                output_file = get_distances_from_nadir(dir_path, new_file_path)
                final_file = append_issues_sheet(output_file)
                append_ez_aoc_sheet(final_file)
                print_to_widget(f"Filtered extract saved as: {os.path.basename(final_file)}")
                rename_images_auto(dir_path)
                # Display a message box to indicate the process is complete
                messagebox.showinfo(title='Data Prep', message='Data prep process completed')
            except (FileNotFoundError, ValueError) as e:
                display_exception()
        else:
            print_to_widget("Operation canceled.\n")

    filterandrenamedistrothread = threading.Thread(target=filterandrenamedistro_thread)
    filterandrenamedistrothread.start()


def filterandrenametrans():
    def filterandrenametrans_thread():
        # Ask for the Excel file
        print_to_widget(f"\nSelect Extract File.")
        file_path = filedialog.askopenfilename(title="Select the Excel file",
                                               filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        if file_path == "":
            print_to_widget("Operation canceled.")
            return
        print_to_widget(f"Selected File: {os.path.basename(file_path)}.")

        # Ask the user to choose the directory of folders
        print_to_widget(f"\nSelect Structure Folder Directory.")
        dir_path = filedialog.askdirectory(title="Select Structure Folder Directory")
        print_to_widget(f"Selected Folder Directory: {os.path.basename(dir_path)}.")

        # Check if the user has provided both the Excel file and the source directory
        if file_path and dir_path:
            try:
                new_file_path = filter_extract_trans(file_path, dir_path)
                output_file = get_distances_from_nadir(dir_path, new_file_path)
                final_file = append_issues_sheet(output_file)
                append_ez_aoc_sheet(final_file)
                print_to_widget(f"Filtered extract saved as: {os.path.basename(final_file)}")
                rename_images_auto(dir_path)
                # Display a message box to indicate the process is complete
                messagebox.showinfo(title='Data Prep', message='Data prep process completed')
            except (FileNotFoundError, ValueError) as e:
                display_exception()
        else:
            print_to_widget("Operation canceled.\n")

    filterandrenametransthread = threading.Thread(target=filterandrenametrans_thread)
    filterandrenametransthread.start()


def packagedata():
    def packagedata_thread():
        # Ask for the traveler sheet file
        print_to_widget(f"\nSelect the completed Traveler Sheet.")
        file_path = filedialog.askopenfilename(title="Select the completed Traveler Sheet.",
                                               filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        if file_path == "":
            print_to_widget("Operation canceled.")
            return
        print_to_widget(f"Selected File: {os.path.basename(file_path)}.")

        # Ask the user to choose the directory of folders
        print_to_widget(f"\nSelect Structure Folder Directory.")
        dir_path = filedialog.askdirectory(title="Select Structure Folder Directory")
        print_to_widget(f"Selected Folder Directory: {os.path.basename(dir_path)}.")

        # Check if the user has provided both the Excel file and the source directory
        if file_path and dir_path:
            try:
                package_data(file_path, dir_path)
                # Display a message box to indicate the process is complete
                messagebox.showinfo(title='Package Data', message='Data packaging process completed')
            except (FileNotFoundError, ValueError) as e:
                display_exception()
        else:
            print_to_widget("Operation canceled.\n")

    # Create a new thread for the filterextractdistro_thread() function
    packagedatathread = threading.Thread(target=packagedata_thread)
    packagedatathread.start()


def packagehelodata():
    def packagehelodata_thread():
        # Ask for the traveler sheet file
        print_to_widget(f"\nSelect the filtered Helo Traveler sheet.")
        file_path = filedialog.askopenfilename(title="Select the Filtered Helo Traveler Sheet",
                                               filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        if file_path == "":
            print_to_widget("Operation canceled.")
            return
        print_to_widget(f"Selected File: {os.path.basename(file_path)}.")

        # Ask the user to choose the directory of folders
        print_to_widget(f"\nSelect Structure Folder Directory.")
        dir_path = filedialog.askdirectory(title="Select Structure Folder Directory")
        print_to_widget(f"Selected Folder Directory: {os.path.basename(dir_path)}.")

        # Check if the user has provided both the Excel file and the source directory
        if file_path and dir_path:
            try:
                package_helo_data(file_path, dir_path)
                # Display a message box to indicate the process is complete
                messagebox.showinfo(title='Package Data', message='Data packaging process completed')
            except (FileNotFoundError, ValueError) as e:
                display_exception()
        else:
            print_to_widget("Operation canceled.\n")

    # Create a new thread for the filterextractdistro_thread() function
    packagehelodatathread = threading.Thread(target=packagehelodata_thread)
    packagehelodatathread.start()


def undopackagedata():
    def undopackagedata_thread():
        # Ask the user to choose the directory of folders
        print_to_widget(f"\nSelect Structure Folder Directory.")
        dir_path = filedialog.askdirectory(title="Select Structure Folder Directory")
        print_to_widget(f"Selected Folder Directory: {os.path.basename(dir_path)}.")

        # Check if the user has provided both the Excel file and the source directory
        if dir_path:
            try:
                sort_images_by_structure_id(dir_path)
                # Display a message box to indicate the process is complete
                messagebox.showinfo(title='Undo Packaging', message='Images sorted by Structure ID')
            except (FileNotFoundError, ValueError) as e:
                display_exception()
        else:
            print_to_widget("Operation canceled.\n")

    # Create a new thread for the filterextractdistro_thread() function
    undopackagedatathread = threading.Thread(target=undopackagedata_thread)
    undopackagedatathread.start()


def watermarkprep():
    def watermarkprep_thread():
        # Ask for the source directory
        print_to_widget(f"\nSelect source directory to watermark.")
        src_dir = filedialog.askdirectory(title='Select source directory to watermark')
        print_to_widget(f"Selected Folder Directory: {os.path.basename(src_dir)}.")

        # Check if the user has provided both the Excel file and the source directory
        if src_dir:
            try:
                watermark_prep(src_dir)
                # Display a message box to indicate the process is complete
                messagebox.showinfo(title='Watermark Prep', message='Images are now ready to watermark.')
            except (FileNotFoundError, ValueError) as e:
                display_exception()
        else:
            print_to_widget("Operation canceled.\n")

    watermarkprepthread = threading.Thread(target=watermarkprep_thread)
    watermarkprepthread.start()


def completetravelerdistro():
    def completetravelerdistro_thread():
        # Ask for the extract file
        print_to_widget(f"\nSelect the filtered extract.")
        source_file = filedialog.askopenfilename(title="Select the filtered extract",
                                                 filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        if source_file == "":
            print_to_widget("Operation canceled.")
            return
        print_to_widget(f"Selected File: {os.path.basename(source_file)}.")

        # Ask for the traveler sheet file
        print_to_widget(f"\nSelect the traveler sheet template.")
        dest_file = filedialog.askopenfilename(title="Select the traveler sheet template",
                                               filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        print_to_widget(f"Selected File: {os.path.basename(dest_file)}.")

        source_sheet = "FilteredExtract"
        dest_sheet = "C2-Distribution"
        columns_mapping = {
            "C2_ID": "C2_ID",
            "FLOC": "SCE_STRUCT",
            "Block_ID": "BLOCK_ID",
            "Work_Order": "Work Order No.",
            "HighFire_T": "HighFire",
            "AOC": "AOC",
            "Eq_ObjType": "EZ Pole",
            "Mapped_Lat": "Latitude",
            "Mapped_Lon": "Longitude",
            "PhotoCount": "Photo_Count",
            "FieldLat": "FieldLAT",
            "FieldLong": "FieldLong",
            "Team_Number": "Flight Team",
            "PhotoLoc": "Photo_Location",
            "FlightDate": "FLIGHT_DATE",
            "Vendor_Notes": "VENDOR_NOTES",
            "EZ_in_Trans": "EZ_in_Trans",
            "Existing_Notes": "Existing_Notes",
            "P1_Notes": "P1_Notes",
            "P2_Notes": "P2_Notes",
            "P3_Notes": "P3_Notes"
        }

        # Check if the user has provided both the Excel file and the source directory
        if source_file and columns_mapping and dest_file and selected_name:
            try:
                complete_traveler_distro(source_file, dest_file, columns_mapping, selected_name, source_sheet,
                                         dest_sheet)
                # add_data_validation_distro(new_dest_file)
                # Display a message box to indicate the process is complete
                print_to_widget(f"\nTraveler sheet completed.")
                messagebox.showinfo(title='Complete Traveler Sheet', message='Traveler sheet completed')
            except (FileNotFoundError, ValueError) as e:
                display_exception()
        else:
            print_to_widget("Operation canceled.\n")

    completetravelerdistrothread = threading.Thread(target=completetravelerdistro_thread)
    completetravelerdistrothread.start()


def completetravelertrans():
    def completetravelertrans_thread():
        # Ask for the extract file
        print_to_widget(f"\nSelect the filtered extract.")
        source_file = filedialog.askopenfilename(title="Select the filtered extract",
                                                 filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        if source_file == "":
            print_to_widget("Operation canceled.")
            return
        print_to_widget(f"Selected File: {os.path.basename(source_file)}.")

        # Ask for the traveler sheet file
        print_to_widget(f"\nSelect the traveler sheet template.")
        dest_file = filedialog.askopenfilename(title="Select the traveler sheet template",
                                               filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        print_to_widget(f"Selected File: {os.path.basename(dest_file)}.")

        source_sheet = "FilteredExtract"  # or 0 for the first sheet
        dest_sheet = "Known"
        columns_mapping = {
            "FLOC": "SCE_STRUCT",
            "District_Number": "DistrictNu",
            "Circuit_Name": "CIRCUIT_NA",
            "Circuit_Fl": "CIRCUIT_FL",
            "Voltage": "VOLTAGE",
            "HighFire_T": "SCE_Design",
            "Block_ID": "BLOCK_ID",
            "Mapped_Lat": "Latitude",
            "Mapped_Lon": "Longitude",
            "PhotoCount": "Photo_Count",
            "FieldLat": "FieldLAT",
            "FieldLong": "FieldLong",
            "PhotoLoc": "Photo_Location",
            "FlightDate": "Flight_Date",
            "Vendor_Notes": "Vendor_Note",
            "Team_Number": "Team_Number",
            "EZ_in_Dist": "EZ_in_Distro",
            "Mile_Tower": "Mile_Tower",
            "P1_Notes": "P1_Notes",
        }

        # Check if the user has provided both the Excel file and the source directory
        if source_file and columns_mapping and dest_file and selected_name:
            try:
                complete_traveler_trans(source_file, dest_file, columns_mapping, selected_name, source_sheet,
                                        dest_sheet)
                # add_data_validation_trans(new_dest_file)
                # Display a message box to indicate the process is complete
                print_to_widget(f"\nTraveler sheet completed.")
                messagebox.showinfo(title='Complete Traveler Sheet', message='Traveler sheet completed')
            except (FileNotFoundError, ValueError) as e:
                display_exception()
        else:
            print_to_widget("Operation canceled.\n")

    completetravelertransthread = threading.Thread(target=completetravelertrans_thread)
    completetravelertransthread.start()


def mergetravelersheets():
    def mergetravelersheets_thread():
        global traveler_sheet_paths, merge_ts_dest_file, merge_ts_source_sheet, merge_ts_dest_sheet, \
            merge_ts_complete_traveler_func, merge_ts_columns_mapping
        try:
            # Read and concatenate the sheets from the Excel files
            data_frames = [pd.read_excel(file, sheet_name=merge_ts_source_sheet) for file in traveler_sheet_paths if
                           merge_ts_source_sheet
                           in pd.ExcelFile(file).sheet_names]
            merged_data = pd.concat(data_frames)

            # Create a temporary file for the output
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file:
                temp_filename: str = temp_file.name

                # Save the merged data to the temporary file
                merged_data.to_excel(temp_filename, index=False, sheet_name=merge_ts_source_sheet)

                # Call complete_traveler_func with the temporary file as the source_file argument
                merge_ts_complete_traveler_func(
                    temp_filename, merge_ts_dest_file, merge_ts_columns_mapping, '1', merge_ts_source_sheet,
                    merge_ts_dest_sheet)

            # Clean up the temporary file
            os.remove(temp_filename)
            traveler_sheet_paths.clear()

            # Display a message box to indicate the process is complete
            print_to_widget("\nTraveler sheets merged.")
            messagebox.showinfo(title='Merge Traveler Sheet', message='Traveler sheets merged successfully')

            traveler_sheet_paths = []
            merge_ts_dest_file = None
            merge_ts_source_sheet = None
            merge_ts_dest_sheet = None
            merge_ts_complete_traveler_func = complete_traveler_distro
            merge_ts_columns_mapping = {}

        except (FileNotFoundError, ValueError) as e:
            display_exception()

    mergetravelersheetsthread = threading.Thread(target=mergetravelersheets_thread)
    mergetravelersheetsthread.start()


# def mergealldirectories():
#     try:
#         merge_all_directories()
#     except (FileNotFoundError, ValueError) as e:
#         display_exception()


def gisvsuc():
    def gisvsuc_thread():
        # Fetch date from the queue
        gis_uc_date_input = q.get()

        # Check if the user has provided both the GIS file and the Upload Check file
        if date_input and uc_file and gis_file:
            try:
                gis_vs_uc(date_input, uc_file, gis_file)
                # Display a message box to indicate the process is complete
                messagebox.showinfo(title='GIS vs Upload Check', message='GIS vs Upload Check report complete.')
            except (FileNotFoundError, ValueError) as e:
                display_exception()
        else:
            print_to_widget("Operation canceled.\n")

    # Ask the user to choose GIS and Upload Check files
    print_to_widget(f"\nSelect GIS Extract File.")
    gis_file = filedialog.askopenfilename(title="Select GIS Extract File")
    if gis_file == "":
        print_to_widget("Operation canceled.")
        return
    print_to_widget(f"Selected File: {os.path.basename(gis_file)}.")

    print_to_widget(f"\nSelect Upload Check File.")
    uc_file = filedialog.askopenfilename(title="Select Upload Check File")
    if uc_file == "":
        print_to_widget("Operation canceled.")
        return
    print_to_widget(f"Selected File: {os.path.basename(uc_file)}.")

    # Prompt the user to enter the date (repeatedly until a valid format is inputted)
    while True:
        # Open a dialog box to prompt the user for input
        date_input = tk.simpledialog.askstring("Enter Date", "Enter the date (YYYYMMDD): ")

        try:
            # Attempt to convert the input to a datetime object
            datetime.strptime(date_input, "%Y%m%d")
            break  # Valid format, exit the loop
        except ValueError:
            # Invalid format, prompt the user again for a valid date format
            messagebox.showerror("Invalid Format",
                                 "Incorrect date format. Please enter the date in the format (YYYYMMDD).")

    # Put date_input into queue from the main thread
    q.put(date_input)

    gisvsucthread = threading.Thread(target=gisvsuc_thread)
    gisvsucthread.start()


def gisvsts():
    def gisvsts_thread():

        # Check if the user has provided both the GIS file and the Upload Check file
        if ts_file and gis_file:
            try:
                gis_vs_ts(gis_file, ts_file)
                # Display a message box to indicate the process is complete
                messagebox.showinfo(title='GIS vs Traveler Sheet', message='GIS vs Traveler Sheet report complete.')
            except (FileNotFoundError, ValueError) as e:
                display_exception()
        else:
            print_to_widget("Operation canceled.\n")

    # Ask the user to choose GIS and Upload Check files
    print_to_widget(f"\nSelect GIS Extract File.")
    gis_file = filedialog.askopenfilename(title="Select GIS Extract File")
    if gis_file == "":
        print_to_widget("Operation canceled.")
        return
    print_to_widget(f"Selected File: {os.path.basename(gis_file)}.")

    print_to_widget(f"\nSelect Traveler Sheet File.")
    ts_file = filedialog.askopenfilename(title="Select Traveler Sheet File")
    if ts_file == "":
        print_to_widget("Operation canceled.")
        return
    print_to_widget(f"Selected File: {os.path.basename(ts_file)}.")

    gisvststhread = threading.Thread(target=gisvsts_thread)
    gisvststhread.start()


def deltareport():
    def deltareport_thread():
        # Check if the user has provided both the GIS file and the Upload Check file
        if gis_file and ts_file and uc_file:
            try:
                delta_report(gis_file, ts_file, uc_file)
                # Display a message box to indicate the process is complete
                messagebox.showinfo(title='Delta Report', message='Delta report complete.')
            except (FileNotFoundError, ValueError) as e:
                display_exception()
        else:
            print_to_widget("Operation canceled.\n")

    # Ask the user to choose GIS, Traveler Sheet, and Upload Check files
    print_to_widget(f"\nSelect GIS Extract File.")
    gis_file = filedialog.askopenfilename(title="Select GIS Extract File")
    if gis_file == "":
        print_to_widget("Operation canceled.")
        return
    print_to_widget(f"Selected File: {os.path.basename(gis_file)}.")

    print_to_widget(f"\nSelect Traveler Sheet File.")
    ts_file = filedialog.askopenfilename(title="Select Traveler Sheet File")
    if ts_file == "":
        print_to_widget("Operation canceled.")
        return
    print_to_widget(f"Selected File: {os.path.basename(ts_file)}.")

    print_to_widget(f"\nSelect Upload Check File.")
    uc_file = filedialog.askopenfilename(title="Select Upload Check File")
    if uc_file == "":
        print_to_widget("Operation canceled.")
        return
    print_to_widget(f"Selected File: {os.path.basename(uc_file)}.")

    deltareportthread = threading.Thread(target=deltareport_thread)
    deltareportthread.start()


def dailydeltadistro():
    def dailydeltadistro_thread():
        # Fetch date range from the queue
        daily_d_start_date = q.get()
        # Check if the user has provided both the GIS file and the Upload Check file
        if traveler_file and template_file and gis_file and eod_file and uc_file and daily_d_start_date:
            try:
                teams_worked, output_file = data_conformance_distro(gis_file, eod_file, uc_file, template_file,
                                                                    daily_d_start_date, daily_d_start_date)
                create_team_sheets(teams_worked, daily_d_start_date, daily_d_start_date, uc_file, gis_file, output_file)
                daily_delta(traveler_file, output_file, daily_d_start_date)
                # Display a message box to indicate the process is complete
                messagebox.showinfo(title='Daily Delta', message='Daily delta report complete.')
            except (FileNotFoundError, ValueError) as e:
                display_exception()
        else:
            print_to_widget("Operation canceled.\n")

    # Ask the user to choose necessary files
    print_to_widget(f"\nSelect GIS Extract File.")
    gis_file = filedialog.askopenfilename(title="Select GIS Extract File")
    if gis_file == "":
        print_to_widget("Operation canceled.")
        return
    print_to_widget(f"Selected File: {os.path.basename(gis_file)}.")

    print_to_widget(f"\nSelect End-of-Day File.")
    eod_file = filedialog.askopenfilename(title="Select End-of-Day File")
    if eod_file == "":
        print_to_widget("Operation canceled.")
        return
    print_to_widget(f"Selected File: {os.path.basename(eod_file)}.")

    print_to_widget(f"\nSelect Upload Check File.")
    uc_file = filedialog.askopenfilename(title="Select Upload Check File")
    if uc_file == "":
        print_to_widget("Operation canceled.")
        return
    print_to_widget(f"Selected File: {os.path.basename(uc_file)}.")

    print_to_widget(f"\nSelect Traveler Sheet File.")
    traveler_file = filedialog.askopenfilename(title="Select Traveler Sheet File")
    if traveler_file == "":
        print_to_widget("Operation canceled.")
        return
    print_to_widget(f"Selected File: {os.path.basename(traveler_file)}.")

    print_to_widget(f"\nSelect Data Conformance Template.")
    template_file = filedialog.askopenfilename(title="Select Data Conformance Template")
    if template_file == "":
        print_to_widget("Operation canceled.")
        return
    print_to_widget(f"Selected File: {os.path.basename(template_file)}.")

    # Prompt the user to enter the date (repeatedly until a valid format is inputted)
    while True:
        delta_start_date = simpledialog.askstring("Enter Start Date", "Enter the start date (YYYYMMDD): ",
                                                  parent=root)
        # Check if Cancel was pressed or dialog was closed
        if delta_start_date is None:
            print_to_widget("\nOperation canceled.")
            return

        # Check if an empty string was entered
        if delta_start_date == '':
            messagebox.showwarning("No Date Entered", "Please enter a start date.", parent=root)
            continue  # Restart the loop, prompting for the start date again

        try:
            datetime.strptime(delta_start_date, "%Y%m%d")
            break
        except ValueError:
            messagebox.showerror("Invalid Format",
                                 "Incorrect date format. Please enter the date in the format (YYYYMMDD).",
                                 parent=root)

    # Put date_input into queue from the main thread
    q.put(delta_start_date)

    dailydeltadistrothread = threading.Thread(target=dailydeltadistro_thread())
    dailydeltadistrothread.start()


def dailydeltatrans():
    def dailydeltatrans_thread():
        # Fetch date range from the queue
        daily_t_start_date = q.get()
        # Check if the user has provided both the GIS file and the Upload Check file
        if traveler_file and template_file and gis_file and eod_file and uc_file:
            try:
                teams_worked, output_file = data_conformance_trans(gis_file, eod_file, uc_file, template_file,
                                                                   daily_t_start_date, daily_t_start_date)
                create_team_sheets(teams_worked, daily_t_start_date, daily_t_start_date, uc_file, gis_file, output_file)
                daily_delta(traveler_file, template_file, daily_t_start_date)
                # Display a message box to indicate the process is complete
                messagebox.showinfo(title='Daily Delta', message='Daily delta report complete.')
            except (FileNotFoundError, ValueError) as e:
                display_exception()
        else:
            print_to_widget("Operation canceled.\n")

    # Ask the user to choose necessary files
    print_to_widget(f"\nSelect GIS Extract File.")
    gis_file = filedialog.askopenfilename(title="Select GIS Extract File")
    if gis_file == "":
        print_to_widget("Operation canceled.")
        return
    print_to_widget(f"Selected File: {os.path.basename(gis_file)}.")

    print_to_widget(f"\nSelect End-of-Day File.")
    eod_file = filedialog.askopenfilename(title="Select End-of-Day File")
    if eod_file == "":
        print_to_widget("Operation canceled.")
        return
    print_to_widget(f"Selected File: {os.path.basename(eod_file)}.")

    print_to_widget(f"\nSelect Upload Check File.")
    uc_file = filedialog.askopenfilename(title="Select Upload Check File")
    if uc_file == "":
        print_to_widget("Operation canceled.")
        return
    print_to_widget(f"Selected File: {os.path.basename(uc_file)}.")

    print_to_widget(f"\nSelect Traveler Sheet File.")
    traveler_file = filedialog.askopenfilename(title="Select Traveler Sheet File")
    if traveler_file == "":
        print_to_widget("Operation canceled.")
        return
    print_to_widget(f"Selected File: {os.path.basename(traveler_file)}.")

    print_to_widget(f"\nSelect Data Conformance File.")
    template_file = filedialog.askopenfilename(title="Select Data Conformance File")
    if template_file == "":
        print_to_widget("Operation canceled.")
        return
    print_to_widget(f"Selected File: {os.path.basename(template_file)}.")

    # Prompt the user to enter the date (repeatedly until a valid format is inputted)
    while True:
        delta_start_date = simpledialog.askstring("Enter Start Date", "Enter the start date (YYYYMMDD): ",
                                                  parent=root)
        # Check if Cancel was pressed or dialog was closed
        if delta_start_date is None:
            print_to_widget("\nOperation canceled.")
            return

        # Check if an empty string was entered
        if delta_start_date == '':
            messagebox.showwarning("No Date Entered", "Please enter a start date.", parent=root)
            continue  # Restart the loop, prompting for the start date again

        try:
            datetime.strptime(delta_start_date, "%Y%m%d")
        except ValueError:
            messagebox.showerror("Invalid Format",
                                 "Incorrect date format. Please enter the date in the format (YYYYMMDD).",
                                 parent=root)

    # Put date_input into queue from the main thread
    q.put(delta_start_date)

    dailydeltatransthread = threading.Thread(target=dailydeltatrans_thread())
    dailydeltatransthread.start()


def combineextractsdistro():
    def combineextractsdistro_thread():
        # Check if the user has provided both the GIS file and the Upload Check file
        if extract1 and extract2:
            try:
                combine_extracts_distro(extract1, extract2)
                # Display a message box to indicate the process is complete
                messagebox.showinfo(title='Combine Extracts', message='Extracts combined successfully!')
            except (FileNotFoundError, ValueError) as e:
                display_exception()
        else:
            print_to_widget("Operation canceled.\n")

    # Ask the user to choose GIS, Traveler Sheet, and Upload Check files
    print_to_widget(f"\nSelect Original Extract File.")
    extract1 = filedialog.askopenfilename(title="Select Original Extract File")
    if extract1 == "":
        print_to_widget("Operation canceled.")
        return
    print_to_widget(f"Selected File: {os.path.basename(extract1)}.")

    print_to_widget(f"\nSelect New Extract File.")
    extract2 = filedialog.askopenfilename(title="Select New Extract File")
    if extract2 == "":
        print_to_widget("Operation canceled.")
        return
    print_to_widget(f"Selected File: {os.path.basename(extract2)}.")

    combineextractsdistrothread = threading.Thread(target=combineextractsdistro_thread)
    combineextractsdistrothread.start()


def combineextractstrans():
    def combineextractstrans_thread():
        # Check if the user has provided both the GIS file and the Upload Check file
        if extract1 and extract2:
            try:
                combine_extracts_trans(extract1, extract2)
                # Display a message box to indicate the process is complete
                messagebox.showinfo(title='Combine Extracts', message='Extracts combined successfully!')
            except (FileNotFoundError, ValueError) as e:
                display_exception()
        else:
            print_to_widget("Operation canceled.\n")

    # Ask the user to choose GIS, Traveler Sheet, and Upload Check files
    print_to_widget(f"\nSelect Original Extract File.")
    extract1 = filedialog.askopenfilename(title="Select Original Extract File")
    if extract1 == "":
        print_to_widget("Operation canceled.")
        return
    print_to_widget(f"Selected File: {os.path.basename(extract1)}.")

    print_to_widget(f"\nSelect New Extract File.")
    extract2 = filedialog.askopenfilename(title="Select New Extract File")
    if extract2 == "":
        print_to_widget("Operation canceled.")
        return
    print_to_widget(f"Selected File: {os.path.basename(extract2)}.")

    combineextractstransthread = threading.Thread(target=combineextractstrans_thread)
    combineextractstransthread.start()


def filterhelotraveler():
    def filterhelotraveler_thread():
        # Ask for the Excel file
        print_to_widget(f"\nSelect Helo Traveler File.")
        file_path = filedialog.askopenfilename(title="Select the Helo Traveler File",
                                               filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        if file_path == "":
            print_to_widget("Operation canceled.")
            return
        print_to_widget(f"Selected File: {os.path.basename(file_path)}.")

        # Ask the user to choose the directory of folders
        print_to_widget(f"Select Structure Folder Directory.")
        dir_path = filedialog.askdirectory(title="Select Structure Folder Directory")
        print_to_widget(f"Selected Folder Directory: {os.path.basename(dir_path)}.")

        # Check if the user has provided both the Excel file and the source directory
        if file_path and dir_path:
            try:
                new_file_path = filter_helo_traveler(file_path, dir_path)
                print_to_widget(f"Filtered helo traveler saved as: {os.path.basename(new_file_path)}")
                # Display a message box to indicate the process is complete
                messagebox.showinfo(title='Filter Helo Traveler Sheet', message='Traveler sheet filtered and '
                                                                                'processed successfully')
            except (FileNotFoundError, ValueError) as e:
                display_exception()
        else:
            print_to_widget("Operation canceled.\n")

    # Create a new thread for the filterextractdistro_thread() function
    filterhelotravelerthread = threading.Thread(target=filterhelotraveler_thread)
    filterhelotravelerthread.start()


# Define the color palette for dark mode
dark_bg = "gray10"
dark_fg = "gray84"
accent_color = '#242526'
hover_color = '#3A3B3C'

# Define the color palette for custom tk dark mode
ctk.set_appearance_mode("Dark")  # Set the appearance to dark mode
ctk.set_default_color_theme("dark-blue")  # Set the default color theme


def eod_distro():
    webbrowser.open('https://c2groupoffice.sharepoint.com/:x:/r/sites/C2Group-SCEFieldOperations/Shared%20Documents'
                    '/General/Distribution%20-%20End%20of%20Day.xlsx?d=w21d4cfccae4544199766c4d10bbe8e37&csf=1&web=1'
                    '&e=hZmVaK')


def eod_trans():
    webbrowser.open('https://c2groupoffice.sharepoint.com/:x:/r/sites/C2Group-SCEFieldOperations/Shared%20Documents'
                    '/General/Transmission%20-%20End%20of%20Day.xlsx?d=w3393c26dc1e04d389793299f71e122a4&csf=1&web=1'
                    '&e=C4NVoQ')


def uc_distro():
    webbrowser.open('https://c2groupoffice-my.sharepoint.com/:x:/r/personal/c2docs_c2groupoffice_onmicrosoft_com'
                    '/Documents/C2%20Operations/C2%20Projects/SCE/22-032%20UAV/03%20-%20Field%20Ops/05%20-%20Upload'
                    '%20Check/2024_Distribution_Upload_Check.xlsx?d=w917f410190524758bb0978585fe2c98f&csf=1&web=1&e'
                    '=2THD5d')


def uc_trans():
    webbrowser.open('https://c2groupoffice-my.sharepoint.com/:x:/r/personal/c2docs_c2groupoffice_onmicrosoft_com'
                    '/Documents/C2%20Operations/C2%20Projects/SCE/22-032%20UAV/03%20-%20Field%20Ops/05%20-%20Upload'
                    '%20Check/2024_Transmission_Upload_Check.xlsx?d=w8e41a40801c84f048a3f6fa207fc30c3&csf=1&web=1&e'
                    '=LgI5hN')


def upload_distro():
    webbrowser.open('https://c2groupoffice-my.sharepoint.com/:f:/r/personal/c2drone_c2groupoffice_onmicrosoft_com'
                    '/Documents/UAV%20Projects/SCE/Field%20Uploads/2024/Distribution?csf=1&web=1&e=15i4Ya')


def upload_trans():
    webbrowser.open('https://c2groupoffice-my.sharepoint.com/:f:/r/personal/c2drone_c2groupoffice_onmicrosoft_com'
                    '/Documents/UAV%20Projects/SCE/Field%20Uploads/2024/Transmission?csf=1&web=1&e=WyUbxd')


def extract_distro():
    webbrowser.open('https://c2groupoffice-my.sharepoint.com/:f:/r/personal/c2docs_c2groupoffice_onmicrosoft_com'
                    '/Documents/C2%20Operations/C2%20Projects/SCE/22-032%20UAV/04%20-%20QA/02%20Daily%20Extracts'
                    '%20From%20GIS/Distribution?csf=1&web=1&e=07qI4C')


def extract_trans():
    webbrowser.open('https://c2groupoffice-my.sharepoint.com/:f:/r/personal/c2docs_c2groupoffice_onmicrosoft_com'
                    '/Documents/C2%20Operations/C2%20Projects/SCE/22-032%20UAV/04%20-%20QA/02%20Daily%20Extracts'
                    '%20From%20GIS/Transmission?csf=1&web=1&e=CITguE')


def gis_distro():
    webbrowser.open('https://conekt2.maps.arcgis.com/apps/mapviewer/index.html?webmap=2f137240cab044399017cbe93ddc3a93')


def gis_trans():
    webbrowser.open('https://conekt2.maps.arcgis.com/apps/mapviewer/index.html?webmap=098e0cb6502640b9bcccd130b8119b68')


def master_traveler_distro():
    webbrowser.open('https://c2groupoffice-my.sharepoint.com/:x:/r/personal/c2docs_c2groupoffice_onmicrosoft_com'
                    '/Documents/C2%20Operations/C2%20Projects/SCE/22-032%20UAV/04%20-%20QA/06%20Tracking/2024%20SCE'
                    '%20Distribution%20Master.xlsx?d=w41c2b01abacf4b159406c3febeac1f50&csf=1&web=1&e=eOVNhd')


def master_traveler_trans():
    webbrowser.open('https://c2groupoffice-my.sharepoint.com/:x:/r/personal/c2docs_c2groupoffice_onmicrosoft_com'
                    '/Documents/C2%20Operations/C2%20Projects/SCE/22-032%20UAV/04%20-%20QA/06%20Tracking/2024%20SCE'
                    '%20Transmission%20Master.xlsx?d=w3220f4ca700c44b68f2e11ba1d71d932&csf=1&web=1&e=d2X1dF')


def on_name_selected(*args):
    if name_var.get() == "Other":
        name_option.pack(padx=10, pady=20)
        custom_name_label.pack(padx=10, pady=(10, 0))  # Show the label
        custom_name_entry.pack(padx=10, pady=(0, 20))  # Show the entry
        submit_button.pack_forget()
        submit_button.pack(padx=10, pady=30)
    else:
        name_option.pack(padx=10, pady=(20, 105))
        submit_button.pack(padx=10, pady=30)
        custom_name_label.pack_forget()  # Hide the label
        custom_name_entry.pack_forget()  # Hide the entry


def show_merge_traveler_frame(event=None):
    # Hide all frames
    for each_frame in frames.values():
        each_frame.grid_remove()
    # Show the merge_traveler_frame
    option_frame.grid(row=0, column=0, columnspan=3, padx=0, pady=10)
    merge_traveler_frame.grid()


def show_merge_directory_frame(event=None):
    # Hide all frames
    for each_frame in frames.values():
        each_frame.grid_remove()
    # Show the merge_directory_frame
    option_frame.grid(row=0, column=0, columnspan=3, padx=0, pady=10)
    merge_directory_frame.grid()


# Function to show the corresponding frame based on the dropdown selection
def show_selected_frame(event=None):
    global last_shown_frame
    selected_option = option_menu.get()

    # Hide all frames first
    for each_frame in frames.values():
        each_frame.grid_remove()

    # Check if the selected option is a valid key in the frames dictionary
    option_frame.grid(row=0, column=0, columnspan=3, padx=0, pady=10)
    if selected_option in frames:
        # Show the selected frame
        selected_frame = frames[selected_option]
        selected_frame.grid(row=1, column=0, columnspan=3, padx=20, pady=(30, 0), sticky="nsew")
        selected_frame.grid_columnconfigure(0, weight=1)
        last_shown_frame = selected_frame


def on_submit():
    global selected_name
    selected_name = custom_name_entry.get().strip() if name_var.get() == "Other" else name_var.get().strip()

    if not selected_name:
        # User has not entered a name, show an alert or update a label to prompt for a name
        print_to_widget("Please enter your first and last name.")
        return  # Stop further execution of this function

    # Assuming selected_name is a string like "John Doe"
    first_name = selected_name.split()[0]
    print_to_widget(f"Hello, {first_name}!")
    qa_frame.grid_remove()
    option_frame.grid(row=0, column=0, columnspan=3, padx=0, pady=10)
    show_selected_frame()


def back_to_previous_frame(event=None):
    global traveler_sheet_paths, merge_ts_dest_file, merge_ts_source_sheet, merge_ts_dest_sheet, \
        merge_ts_complete_traveler_func, merge_ts_columns_mapping
    # Hide the current frame
    merge_traveler_frame.grid_remove()
    merge_directory_frame.grid_remove()
    # Show the last active frame before merge_traveler_frame was shown
    if last_shown_frame is not None:
        last_shown_frame.grid()
        option_frame.grid()
        traveler_sheet_paths = []
        merge_ts_dest_file = None
        merge_ts_source_sheet = None
        merge_ts_dest_sheet = None
        merge_ts_complete_traveler_func = complete_traveler_distro
        merge_ts_columns_mapping = {}


# Create an instance of the standard ctk.CTk class
root = ctk.CTk()
root.title("QA Scripts")
root.grid_rowconfigure(1, weight=1)

# Create a dropdown list for names
name_list = ["Austin Tinnell", "Evan Boeh", "Karen Moreno", "Luis Haro", "Miguel Portugal", "Nick Santana",
             "Pretan Tabag", "Taylor Addams", "Other"]

# Create a StringVar for tracking the selection
name_var = tk.StringVar(root)

# Set the initial value of the OptionMenu
name_var.set(name_list[0])

# Define the bold font
bold_font = ctk.CTkFont(family="Roboto", size=15, weight="bold")

# QA NAME Frame and its contents
qa_frame = ctk.CTkFrame(root, fg_color="gray10")
name_option_label = ctk.CTkLabel(qa_frame, text="Select your name from the list:")
name_option_label.pack(padx=10, pady=(10, 0), anchor='center')
name_option = ctk.CTkOptionMenu(qa_frame, variable=name_var, values=name_list, button_color="#4B5E81",
                                fg_color="#343638", button_hover_color="#35425A")
name_option.pack(padx=10, pady=(20, 105))
# Create an entry widget for custom name input
custom_name_label = ctk.CTkLabel(qa_frame, text="Type your full name below:")
custom_name_label.pack(padx=10, pady=(10, 0), anchor='center')
custom_name_label.pack_forget()  # Hide the label
custom_name_entry = ctk.CTkEntry(qa_frame, state="normal")
custom_name_entry.pack(padx=10, pady=(0, 20), anchor='center')
custom_name_entry.pack_forget()  # Hide the entry
# Create a Submit button
submit_button = ctk.CTkButton(qa_frame, text="Submit", command=on_submit, fg_color="#4B5E81",
                              hover_color="#35425A")
submit_button.pack(padx=10, pady=30, anchor='center')

# Set a callback to the StringVar
name_var.trace("w", lambda *args: on_name_selected(*args))

# Version Label
version_label = ctk.CTkLabel(root, text=f"Version {get_current_version()}", cursor='hand2')
version_label.place(relx=0.99, rely=1.0, anchor="se")
version_label.bind("<Button-1>", open_version_history)

# Define the frames
data_val_frame = ctk.CTkFrame(root, fg_color="gray10")
links_frame = ctk.CTkFrame(root, fg_color="gray10")
dist_frame = ctk.CTkFrame(root, fg_color="gray10")
trans_frame = ctk.CTkFrame(root, fg_color="gray10")
helo_frame = ctk.CTkFrame(root, fg_color="gray10")
tools_frame = ctk.CTkFrame(root, fg_color="gray10")
merge_traveler_frame = ctk.CTkFrame(root, fg_color="gray10")
merge_directory_frame = ctk.CTkFrame(root, fg_color="gray10")
option_frame = ctk.CTkFrame(root, fg_color="gray10")

last_shown_frame = tools_frame

# Create a dictionary of frame configurations
frame_config = {
    "Data Validation": {
        "colors": ("#7B5E7B", "#5C475C"),
        "buttons": [
            ("GIS vs Upload Check", gisvsuc),
            ("GIS vs Traveler", gisvsts),
            ("Delta Report", deltareport),
            ("Daily Delta Distro", dailydeltadistro),
            ("Daily Delta Trans", dailydeltatrans)
        ]},
    "Quick Links": {
        "colors": ("#695E93", "#504870"),
        "buttons": [
            ("DISTRIBUTION:", None),
            ("End-of-Day", eod_distro),
            ("Upload Check", uc_distro),
            ("Field Uploads", upload_distro),
            ("GIS Extract", extract_distro),
            ("GIS Map", gis_distro),
            ("Master Traveler", master_traveler_distro),
            ("TRANSMISSION:", None),
            ("End-of-Day", eod_trans),
            ("Upload Check", uc_trans),
            ("Field Uploads", upload_trans),
            ("GIS Extract", extract_trans),
            ("GIS Map", gis_trans),
            ("Master Traveler", master_traveler_trans)
        ]},
    "Distribution": {
        "colors": ("#3c6e71", "#234143"),
        "buttons": [
            ("Filter Extract", filterextractdistro),
            ("Filter and Rename", filterandrenamedistro),
            ("Complete Traveler", completetravelerdistro),
            ("Rename (Auto)", renameimagesauto),
            ("Package Data", packagedata),
            ("Merge Extracts", combineextractsdistro)
        ]},
    "Transmission": {
        "colors": (None, None),
        "buttons": [
            ("Filter Extract", filterextracttrans),
            ("Filter and Rename", filterandrenametrans),
            ("Complete Traveler", completetravelertrans),
            ("Watermark Prep", watermarkprep),
            ("Rename (Auto)", renameimagesauto),
            ("Package Data", packagedata),
            ("Merge Extracts", combineextractstrans)
        ]},
    "Helo": {
        "colors": ("#67595E", "#41393C"),
        "buttons": [
            ("Filter Helo", filterhelotraveler),
            ("Package Helo", packagehelodata),
            ("Undo Package", undopackagedata)
        ]},
    "Other Tools": {
        "colors": ("#65630F", "#353408"),
        "buttons": [
            ("Rename (Manual)", renameimagesmanual),
            ("Merge Travelers", show_merge_traveler_frame),
            ("Merge Directories", show_merge_directory_frame),
            ("Undo Package", undopackagedata)
        ]}
}

# Create a dictionary to map the frame names to the frame objects
frames = {
    "Data Validation": data_val_frame,
    "Quick Links": links_frame,
    "Distribution": dist_frame,
    "Transmission": trans_frame,
    "Helo": helo_frame,
    "Other Tools": tools_frame,
    "merge_traveler_frame": merge_traveler_frame,
    "merge_directory_frame": merge_directory_frame,
    "option_frame": option_frame
}

# Create buttons for each frame
for frame_key, config in frame_config.items():
    row = 0
    col = 0
    max_cols = 3
    frame = frames[frame_key]  # Access frame object using the dictionary
    frame.grid(row=1, column=0, columnspan=3, padx=0, pady=10)
    fg_color, hover_color = config["colors"]

    for text, command in config["buttons"]:
        # Check for special cases where you want to add labels
        if frame_key == "Quick Links":
            if text == "DISTRIBUTION:":
                label = ctk.CTkLabel(frame, text=text)
                label.configure(font=bold_font)
                label.grid(row=0, column=0, padx=10, sticky="ew")
                row += 1
                col -= 1
            elif text == "TRANSMISSION:":
                label = ctk.CTkLabel(frame, text=text)
                label.configure(font=bold_font)
                label.grid(row=3, column=0, padx=10, pady=(20, 0), sticky="ew")
                row += 1
                col -= 1
            else:
                button = ctk.CTkButton(frame, text=text, command=command, fg_color=fg_color, hover_color=hover_color)
                button.grid(row=row, column=col, padx=10, pady=10, sticky="ew")
        else:
            # Create buttons as usual
            button = ctk.CTkButton(frame, text=text, command=command, fg_color=fg_color, hover_color=hover_color)
            button.grid(row=row, column=col, padx=10, pady=20, sticky="ew")

        col += 1
        if col == max_cols:
            col = 0
            row += 1

## MERGE TRAVELER SHEETS Frame and its contents
merge_traveler_frame.grid_columnconfigure(0, weight=1)
# Create the Merge Directories label widget
merge_travelers_label = ctk.CTkLabel(merge_traveler_frame, text="MERGE TRAVELER SHEETS:")
merge_travelers_label.grid(row=0, column=0, columnspan=4, padx=20, pady=(10, 30))
merge_travelers_label.configure(font=bold_font)
# Create the Add Traveler Sheets button widget
addmultiple_button = ctk.CTkButton(merge_traveler_frame, text="Add Traveler Sheets", command=add_files,
                                   fg_color="#65630F", hover_color="#353408")
addmultiple_button.grid(row=1, column=0, columnspan=4, padx=20, pady=10)
# create_tooltip(addmultiple_button, "Select multiple traveler sheets all at once.")
# Create the Merge Traveler Sheets button widget
merge_travelers_button = ctk.CTkButton(merge_traveler_frame, text="START MERGE", command=merge_travelers,
                                       fg_color="#65630F", hover_color="#353408")
merge_travelers_button.grid(row=2, column=0, columnspan=4, padx=20, pady=10)
# create_tooltip(merge_travelers_button, "Press to start merging traveler sheets.")
# Add a back button to the merge_traveler_frame
back_button = ctk.CTkButton(merge_traveler_frame, text="Back", command=back_to_previous_frame,
                            fg_color="#565B5E", hover_color="#3a3a3a")
back_button.grid(row=3, column=3, padx=10, pady=(80, 0))

## MERGE DIRECTORIES Frame and its contents
merge_directory_frame.grid_columnconfigure(0, weight=1)
# Create the Merge Directories label widget
merge_directories_label = ctk.CTkLabel(merge_directory_frame, text="MERGE DIRECTORIES:")
merge_directories_label.grid(row=0, column=0, columnspan=4, padx=20, pady=(10, 30))
merge_directories_label.configure(font=bold_font)
# Create the Add Directory button widget
add_directory_button = ctk.CTkButton(merge_directory_frame, text="Add Directory", command=add_directory,
                                     fg_color="#65630F", hover_color="#353408")
add_directory_button.grid(row=1, column=0, columnspan=4, padx=20, pady=10)
# create_tooltip(add_directory_button, "Add a folder to combine")
# Create the Merge Directory button widget
merge_directories_button = ctk.CTkButton(merge_directory_frame, text="START MERGE", command=merge_directories,
                                         fg_color="#65630F", hover_color="#353408")
merge_directories_button.grid(row=2, column=0, columnspan=4, padx=20, pady=10)
# create_tooltip(merge_directories_button, "Start merging directories")
# Add a back button to the merge_directory_frame
back_button = ctk.CTkButton(merge_directory_frame, text="Back", command=back_to_previous_frame,
                            fg_color="#565B5E", hover_color="#3a3a3a")
back_button.grid(row=3, column=3, padx=10, pady=(80, 0))

# Options for the dropdown menu
# option_frame = ctk.CTkFrame(root, fg_color="gray10")
option_frame.grid_columnconfigure(0, weight=1)
options = ["Data Validation", "Quick Links", "Distribution", "Transmission", "Helo", "Other Tools"]
option_menu = ctk.CTkOptionMenu(option_frame, values=options, command=show_selected_frame, width=140,
                                button_color="#565B5E", fg_color="#343638", button_hover_color="#3a3a3a")
option_menu.set("Data Validation")  # Set the default option
option_menu.grid(row=0, column=0, columnspan=3, padx=20, pady=(30, 10))

# Initially hide all frames
for frame in frames.values():
    frame.grid(row=1, column=0, columnspan=3, padx=20, sticky="nsew")
    frame.grid_remove()

qa_frame.grid(row=0, column=0, rowspan=14, columnspan=3, padx=20, pady=20)

# Add the print_text widget after other widgets in the root window
text_space = ctk.CTkTextbox(root, wrap=tk.WORD, width=570, height=450)
text_space.configure(fg_color="gray20", font=('Segoe UI', 13))
text_space.grid(row=0, column=3, rowspan=14, columnspan=2, pady=10, padx=(0, 20), sticky=tk.W + tk.E)


# Create a function to copy the text_space content
def copy_text_space_content():
    content = text_space.get(1.0, tk.END)
    root.clipboard_clear()
    root.clipboard_append(content)
    tooltip.show_tip()
    root.after(1500, tooltip.hide_tip)


# Create the "Copy" button widget
copy_button = ctk.CTkButton(root, text="Copy Text", command=copy_text_space_content, fg_color="#565B5E",
                            hover_color="#3a3a3a")
copy_button.grid(row=14, column=4, padx=(0, 20), pady=(0, 30), sticky='e')
tooltip = ToolTip(copy_button, "Copied to clipboard")

# Create the "Issue Tracker" button widget
issue_button = ctk.CTkButton(root, text="Issue Tracker", command=issue_form, fg_color="#565B5E",
                             hover_color="#3a3a3a")
issue_button.grid(row=14, column=3, pady=(0, 30), sticky="w")

# Configure the column and row widths to evenly space the buttons
for i in range(3):
    root.grid_columnconfigure(i, weight=1)

# Set the size of the window
root.geometry("1100x500")

# Display the window
root.deiconify()

# Check for update
check_for_updates()

# Start the main event loop
root.mainloop()
